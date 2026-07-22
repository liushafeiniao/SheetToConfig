"""Automatic Protobuf schema inference tests.

These tests intentionally omit the PROTO worksheet.  The automatic mode must
derive the schema from the existing four-line data-table header and CODE row.
"""

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from sheet_to_config.utils.exporter.converter import ExcelConverter
from sheet_to_config.utils.exporter.exporters.protobuf_exporter import ProtobufExporter
from sheet_to_config.utils.exporter.protobuf_schema import extract_manifest, ProtoSchemaParser
from sheet_to_config.utils.exporter.template import TypeDefinitionTemplate

from tests.workbook_factory import write_workbook


class AutomaticProtobufTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)
        self.root = Path(self.temp_dir.name)
        self.tables = self.root / "tables"
        self.client = self.root / "client"
        self.server = self.root / "server"
        self.tables.mkdir()
        TypeDefinitionTemplate.ensure_exists(str(self.tables))

    @staticmethod
    def _decode(schema, payload):
        message = ProtobufExporter(schema)._message_class()
        message.ParseFromString(payload)
        return message

    def test_infers_scalars_lists_and_platform_fields_without_proto_sheet(self):
        path = write_workbook(
            self.tables / "Item.xlsx",
            fields=(
                ("id", "int", "CS"),
                ("weight", "float", "CS"),
                ("name", "string", "CS"),
                ("alias", "str", "CS"),
                ("enabled", "bool", "CS"),
                ("payload", "bytes", "CS"),
                ("tags", "intList", "CS"),
                ("labels", "strList", "CS"),
                ("matrix", "intList2", "CS"),
                ("grid", "strList2", "CS"),
                ("client_note", "string", "C"),
                ("server_note", "string", "S"),
            ),
            data_rows=(
                (7, 1.5, "Sword", "S", True, "bin", "1#2", "a#b", "1#2|3#4", "a#b|c", "only c", "only s"),
            ),
            code_rows=(("Item", "Item.pb", "cs"),),
        )
        logs = []
        result = ExcelConverter(logs.append).export_all(
            str(self.tables), str(self.client), str(self.server), "cs"
        )
        self.assertTrue(result["success"], logs)
        self.assertTrue((self.client / "Item.proto").is_file())
        self.assertTrue((self.client / "Item.pb").is_file())
        self.assertEqual(
            (self.client / "Item.proto").read_bytes(),
            (self.server / "Item.proto").read_bytes(),
        )

        schema = ProtoSchemaParser.parse_workbook(str(path), "Item")
        client_row = self._decode(schema, (self.client / "Item.pb").read_bytes()).rows[0]
        server_row = self._decode(schema, (self.server / "Item.pb").read_bytes()).rows[0]
        self.assertEqual(7, client_row.id)
        self.assertAlmostEqual(1.5, client_row.weight, places=5)
        self.assertEqual("Sword", client_row.name)
        self.assertTrue(client_row.enabled)
        self.assertEqual(b"bin", client_row.payload)
        self.assertEqual([1, 2], list(client_row.tags))
        self.assertEqual([[1, 2], [3, 4]], [list(v.values) for v in client_row.matrix])
        self.assertEqual([["a", "b"], ["c"]], [list(v.values) for v in client_row.grid])
        self.assertEqual("only c", client_row.client_note)
        self.assertEqual("", client_row.server_note)
        self.assertEqual("", server_row.client_note)
        self.assertEqual("only s", server_row.server_note)

    def test_auto_messages_follow_output_filename_not_worksheet_name(self):
        write_workbook(
            self.tables / "Source.xlsx",
            fields=(
                ("id", "int", "CS"),
                ("matrix", "intList2", "CS"),
            ),
            data_rows=((1, "1#2|3#4"),),
            code_rows=(("任务配置", "QuestConfig.pb", "c"),),
            sheet="任务配置",
        )
        logs = []
        result = ExcelConverter(logs.append).export_all(
            str(self.tables), str(self.client), str(self.server), "c"
        )
        self.assertTrue(result["success"], logs)

        proto_text = (self.client / "QuestConfig.proto").read_text(encoding="utf-8")
        self.assertIn("message QuestConfigTable", proto_text)
        self.assertIn("message QuestConfigRow", proto_text)
        self.assertIn("message QuestConfigmatrixLevel1", proto_text)
        self.assertNotIn("message ___Table", proto_text)
        self.assertNotIn("message ___Row", proto_text)

    def test_reference_uses_the_target_field_scalar_type(self):
        write_workbook(
            self.tables / "Catalog.xlsx",
            fields=(("id", "string", "CS"), ("name", "string", "CS")),
            data_rows=(("item-a", "Alpha"),),
            code_rows=(("Item", "Catalog.json", "c"),),
            sheet="Item",
        )
        write_workbook(
            self.tables / "Source.xlsx",
            fields=(("id", "int", "CS"), ("catalog_id", "catalogRef", "CS")),
            data_rows=((1, "item-a"),),
            code_rows=(("Source", "Source.pb", "c"),),
            sheet="Source",
        )
        definition = load_workbook(self.tables / "TypeDefinition.xlsx")
        definition["CODE"].append(("catalogRef", "find(Catalog, Items, id)", ""))
        definition.save(self.tables / "TypeDefinition.xlsx")
        definition.close()

        logs = []
        result = ExcelConverter(logs.append).export_all(
            str(self.tables), str(self.client), str(self.server), "c"
        )

        self.assertTrue(result["success"], logs)
        proto_text = (self.client / "Source.proto").read_text(encoding="utf-8")
        self.assertIn("string catalog_id", proto_text)

    def test_manifest_keeps_numbers_when_columns_are_inserted_and_reserves_deleted_field(self):
        path = write_workbook(
            self.tables / "Item.xlsx",
            fields=(("id", "int", "CS"), ("name", "string", "CS"), ("old", "int", "CS")),
            data_rows=((1, "Sword", 9),),
            code_rows=(("Item", "Item.pb", "c"),),
        )
        logs = []
        self.assertTrue(
            ExcelConverter(logs.append).export_all(str(self.tables), str(self.client), str(self.server), "c")["success"],
            logs,
        )
        first = extract_manifest((self.client / "Item.proto").read_text(encoding="utf-8"))
        first_fields = {f["name"]: f["number"] for m in first["messages"] for f in m["fields"]}

        book = load_workbook(path)
        ws = book["Item"]
        ws.insert_cols(2)
        ws.cell(1, 2, "new_field")
        ws.cell(2, 2, "int")
        ws.cell(3, 2, "CS")
        ws.cell(5, 2, 42)
        # Remove the old field from the data table while retaining its history.
        old_col = 4
        for row in range(1, ws.max_row + 1):
            ws.cell(row, old_col).value = None
        book.save(path)
        book.close()

        logs.clear()
        self.assertTrue(
            ExcelConverter(logs.append).export_all(str(self.tables), str(self.client), str(self.server), "c")["success"],
            logs,
        )
        second = extract_manifest((self.client / "Item.proto").read_text(encoding="utf-8"))
        second_fields = {f["name"]: f["number"] for m in second["messages"] for f in m["fields"]}
        self.assertEqual(first_fields["id"], second_fields["id"])
        self.assertEqual(first_fields["name"], second_fields["name"])
        row_manifest = next(m for m in second["messages"] if m["name"].endswith("Row"))
        self.assertEqual(first_fields["old"], row_manifest["reserved_numbers"][0])
        self.assertIn("old", row_manifest["reserved_names"])
        self.assertIn("new_field", second_fields)

    def test_explicit_rebuild_allows_scalar_to_list_protocol_change(self):
        path = write_workbook(
            self.tables / "Stage.xlsx",
            fields=(("id", "int", "CS"), ("RewardAwardID", "int", "CS")),
            data_rows=((1, 1001),),
            code_rows=(("Stage", "Stage.pb", "c"),),
            sheet="Stage",
        )
        logs = []
        converter = ExcelConverter(logs.append)
        self.assertTrue(
            converter.export_all(str(self.tables), str(self.client), str(self.server), "c")["success"],
            logs,
        )

        book = load_workbook(path)
        ws = book["Stage"]
        ws.cell(2, 2, "intList")
        ws.cell(5, 2, "1001#1002")
        book.save(path)
        book.close()

        logs.clear()
        self.assertFalse(
            converter.export_all(str(self.tables), str(self.client), str(self.server), "c")["success"]
        )
        self.assertTrue(any("破坏性变更" in line for line in logs), logs)

        logs.clear()
        result = converter.export_all(
            str(self.tables), str(self.client), str(self.server), "c",
            allow_breaking_proto_change=True,
        )
        self.assertTrue(result["success"], logs)
        proto = (self.client / "Stage.proto").read_text(encoding="utf-8")
        self.assertIn("repeated int32 RewardAwardID = 3;", proto)
        schema = ProtoSchemaParser.parse_workbook(str(path), "Stage")
        row = self._decode(schema, (self.client / "Stage.pb").read_bytes()).rows[0]
        self.assertEqual([1001, 1002], list(row.RewardAwardID))

    def test_complex_dict_type_fails_with_actionable_error(self):
        write_workbook(
            self.tables / "Award.xlsx",
            fields=(("id", "int", "CS"), ("rewards", "award", "CS")),
            data_rows=((1, "1001,1,2,50"),),
            code_rows=(("Award", "Award.pb", "c"),),
            sheet="Award",
        )
        logs = []
        result = ExcelConverter(logs.append).export_all(
            str(self.tables), str(self.client), str(self.server), "c"
        )
        self.assertFalse(result["success"])
        self.assertTrue(any("award" in line.lower() or "PROTO" in line for line in logs))
        self.assertFalse((self.client / "Award.pb").exists())
        self.assertFalse((self.client / "Award.proto").exists())

    def test_proto_metadata_only_customizes_package_and_namespace(self):
        path = write_workbook(
            self.tables / "Custom.xlsx",
            fields=(("id", "int", "CS"), ("name", "str", "CS")),
            data_rows=((3, "Blade"),),
            code_rows=(("Item", "Custom.pb", "c"),),
            sheet="Item",
        )
        book = load_workbook(path)
        proto = book.create_sheet("PROTO")
        proto.cell(1, 1, "package")
        proto.cell(1, 2, "game.config")
        proto.cell(2, 1, "csharp_namespace")
        proto.cell(2, 2, "Game.Generated")
        book.save(path)
        book.close()

        logs = []
        result = ExcelConverter(logs.append).export_all(
            str(self.tables), str(self.client), str(self.server), "c"
        )
        self.assertTrue(result["success"], logs)
        proto_text = (self.client / "Custom.proto").read_text(encoding="utf-8")
        self.assertIn("package game.config;", proto_text)
        self.assertIn('option csharp_namespace = "Game.Generated";', proto_text)
        self.assertIn("message CustomTable", proto_text)
        self.assertIn("message CustomRow", proto_text)

    def test_type_definition_aliases_and_constraints_keep_excel_type_format(self):
        definition = load_workbook(self.tables / "TypeDefinition.xlsx")
        ws = definition["CODE"]
        ws.append(("属性列表", "split_list(int)", "项目自定义名称"))
        ws.append(("物品列表2", "split_list2(string)", "项目自定义名称"))
        definition.save(self.tables / "TypeDefinition.xlsx")
        definition.close()
        logs = []
        write_workbook(
            self.tables / "Alias.xlsx",
            fields=(
                ("id", "int", "CS"),
                ("attrs", "属性列表", "CS"),
                ("names", "strList", "CS"),
                ("matrix", "物品列表2", "CS"),
            ),
            data_rows=((1, "10#20", "a#b", "x#y|z"),),
            code_rows=(("Alias", "Alias.pb", "c"),),
            sheet="Alias",
        )
        result = ExcelConverter(logs.append).export_all(
            str(self.tables), str(self.client), str(self.server), "c"
        )
        self.assertTrue(result["success"], logs)


if __name__ == "__main__":
    unittest.main()
