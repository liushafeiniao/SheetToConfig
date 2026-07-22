import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from tests.workbook_factory import write_workbook
from sheet_to_config.utils.exporter.converter import ExcelConverter
from sheet_to_config.utils.exporter.template import TypeDefinitionTemplate


def export_same_workbook_reference(reference_value, *, reference_type="catalog_ref",
                                   extra_type_definitions=(),
                                   include_independent_definition=True,
                                   additional_candidate_type=None,
                                   catalog_reference_conversion=(
                                       "find_id(Catalog,DisplayOnly,target_id)"
                                   )):
    with tempfile.TemporaryDirectory() as temp_dir:
        root = Path(temp_dir)
        tables = root / "tables"
        tables.mkdir()
        TypeDefinitionTemplate.ensure_exists(str(tables))
        definition = load_workbook(tables / "TypeDefinition.xlsx")
        definition["CODE"].append((
            "catalog_ref",
            catalog_reference_conversion,
            "",
        ))
        for type_definition in extra_type_definitions:
            definition["CODE"].append(type_definition)
        definition.save(tables / "TypeDefinition.xlsx")
        definition.close()

        if include_independent_definition:
            code_rows = [
                ("Definitions", "Definitions.json", "c"),
                ("References", "References.json", "c"),
            ]
            if additional_candidate_type:
                code_rows.append(("NonScalar", "NonScalar.json", "c"))
            workbook_path = write_workbook(
                tables / "Catalog.xlsx",
                fields=(
                    ("row_id", "int", "CS"),
                    ("target_id", "int", "CS"),
                ),
                data_rows=((1, 1),),
                code_rows=code_rows,
                sheet="Definitions",
            )
            workbook = load_workbook(workbook_path)
            references = workbook.create_sheet("References")
            references.append(("row_id", "target_id"))
            references.append(("int", reference_type))
            references.append(("CS", "CS"))
            references.append(("", ""))
            references.append((2, reference_value))
            if additional_candidate_type:
                non_scalar = workbook.create_sheet("NonScalar")
                non_scalar.append(("row_id", "target_id"))
                non_scalar.append(("int", additional_candidate_type))
                non_scalar.append(("CS", "CS"))
                non_scalar.append(("", ""))
                non_scalar.append((3, "999999"))
            workbook.save(workbook_path)
            workbook.close()
        else:
            write_workbook(
                tables / "Catalog.xlsx",
                fields=(
                    ("row_id", "int", "CS"),
                    ("target_id", reference_type, "CS"),
                ),
                data_rows=((2, reference_value),),
                code_rows=(("References", "References.json", "c"),),
                sheet="References",
            )

        return ExcelConverter().export_all(
            str(tables), str(root / "client"), str(root / "server"),
            "c", validation_only=True,
        )


def export_zero_references(target_int, target_float):
    with tempfile.TemporaryDirectory() as temp_dir:
        root = Path(temp_dir)
        tables = root / "tables"
        tables.mkdir()
        TypeDefinitionTemplate.ensure_exists(str(tables))
        definition = load_workbook(tables / "TypeDefinition.xlsx")
        definition["CODE"].append((
            "int_ref", "find_id(Target,DisplayOnly,int_id)", ""
        ))
        definition["CODE"].append((
            "float_ref", "find_id(Target,DisplayOnly,float_id)", ""
        ))
        definition.save(tables / "TypeDefinition.xlsx")
        definition.close()
        write_workbook(
            tables / "Target.xlsx",
            fields=(
                ("row_id", "int", "CS"),
                ("int_id", "int", "CS"),
                ("float_id", "float", "CS"),
            ),
            data_rows=((1, target_int, target_float),),
            code_rows=(("Target", "Target.json", "c"),),
            sheet="Target",
        )
        write_workbook(
            tables / "Source.xlsx",
            fields=(
                ("row_id", "int", "CS"),
                ("int_id", "int_ref", "CS"),
                ("float_id", "float_ref", "CS"),
            ),
            data_rows=((9, None, None), (10, 0, 0.0)),
            code_rows=(("Source", "Source.json", "c"),),
            sheet="Source",
        )

        return ExcelConverter().export_all(
            str(tables), str(root / "client"), str(root / "server"),
            "c", validation_only=True,
        )


class ConverterValidationTests(unittest.TestCase):
    def test_validation_collects_bad_rows_and_uses_converted_first_column_as_key(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tables = root / "tables"
            client = root / "client"
            server = root / "server"
            tables.mkdir()
            write_workbook(
                tables / "A.xlsx",
                fields=(("key", "int", "CS"), ("count", "int+required()", "CS")),
                data_rows=((None, 5), ("bad", "oops"), (1, 1), ("1", 2)),
                code_rows=(("Item", "A.json", "c"),),
            )
            write_workbook(
                tables / "B.xlsx",
                fields=(("code", "int", "CS"), ("enabled", "bool", "CS")),
                data_rows=((2, "maybe"),),
                code_rows=(("Item", "B.json", "c"),),
            )

            result = ExcelConverter().export_all(
                str(tables), str(client), str(server), "c", validation_only=True
            )

            self.assertFalse(result["success"])
            codes = {issue["code"] for issue in result["issues"]}
            self.assertIn("MISSING_PRIMARY_KEY", codes)
            self.assertIn("CONVERSION_ERROR", codes)
            self.assertIn("DUPLICATE_VALUE", codes)
            duplicate = next(
                issue for issue in result["issues"]
                if issue["code"] == "DUPLICATE_VALUE"
            )
            self.assertIn("主键必须唯一", duplicate["message"])
            self.assertIn("A8", duplicate["message"])
            self.assertIn("A7", duplicate["message"])
            self.assertEqual(
                set(result["issues"][0]),
                {"code", "message", "file", "sheet", "row", "column", "field", "path", "rawValue"},
            )
            self.assertFalse(client.exists())

    def test_formula_without_cached_value_is_reported_and_writes_nothing(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tables = root / "tables"
            client = root / "client"
            server = root / "server"
            tables.mkdir()
            path = write_workbook(
                tables / "Formula.xlsx",
                fields=(("id", "int", "CS"), ("value", "int", "CS")),
                data_rows=((1, 2),),
                code_rows=(("Item", "Formula.json", "c"),),
            )
            workbook = load_workbook(path)
            workbook["Item"]["B5"] = "=1+1"
            workbook.save(path)
            workbook.close()

            result = ExcelConverter().export_all(
                str(tables), str(client), str(server), "c"
            )

            issue = next(
                item for item in result["issues"]
                if item["code"] == "FORMULA_NO_CACHED_VALUE"
            )
            self.assertEqual((issue["row"], issue["column"], issue["field"]), (5, 2, "value"))
            self.assertFalse((client / "Formula.json").exists())

    def test_asset_root_validates_path_existence_and_rejects_traversal(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tables = root / "tables"
            assets = root / "assets"
            client = root / "client"
            tables.mkdir()
            (assets / "icons").mkdir(parents=True)
            (assets / "icons" / "sword.png").write_bytes(b"png")
            TypeDefinitionTemplate.ensure_exists(str(tables))
            workbook = load_workbook(tables / "TypeDefinition.xlsx")
            workbook["CODE"].append(("asset", "path(icons/,.png)", ""))
            workbook.save(tables / "TypeDefinition.xlsx")
            workbook.close()

            def export_value(value, validation_only=True):
                write_workbook(
                    tables / "Asset.xlsx",
                    fields=(("id", "int", "CS"), ("icon", "asset", "CS")),
                    data_rows=((1, value),),
                    code_rows=(("Item", "Asset.json", "c"),),
                )
                return ExcelConverter().export_all(
                    str(tables), str(client), str(root / "server"),
                    "c", validation_only=validation_only, asset_root=str(assets),
                )

            self.assertTrue(export_value("sword")["success"])
            blank = export_value("   ", validation_only=False)
            self.assertTrue(blank["success"], blank["issues"])
            payload = json.loads(
                (client / "Asset.json").read_text(encoding="utf-8")
            )
            self.assertEqual("", payload["1"]["icon"])
            missing = export_value("missing")
            self.assertFalse(missing["success"])
            self.assertIn("文件不存在", missing["issues"][0]["message"])
            traversal = export_value("../../outside")
            self.assertFalse(traversal["success"])
            self.assertTrue(any(
                "越过资源根目录" in issue["message"] for issue in traversal["issues"]
            ))

    def test_missing_asset_root_skips_path_checks_without_warning(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tables = root / "tables"
            tables.mkdir()
            TypeDefinitionTemplate.ensure_exists(str(tables))
            workbook = load_workbook(tables / "TypeDefinition.xlsx")
            workbook["CODE"].append(("asset", "path(icons/,.png)", ""))
            workbook.save(tables / "TypeDefinition.xlsx")
            workbook.close()
            write_workbook(
                tables / "Item.xlsx",
                fields=(("id", "int", "CS"), ("icon", "asset", "CS")),
                data_rows=((1, "missing"),),
                code_rows=(("Item", "Item.json", "c"),),
            )
            logs = []
            result = ExcelConverter(logs.append).export_all(
                str(tables), str(root / "client"), str(root / "server"),
                "c", validation_only=True,
            )
            self.assertTrue(result["success"], result["issues"])
            self.assertEqual(sum("assetRoot" in line for line in logs), 0)
            self.assertFalse((root / "client").exists())
            self.assertFalse((root / "server").exists())

    def test_reference_validation_canonicalizes_aliases_and_keeps_dict_shape(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tables = root / "tables"
            client = root / "client"
            tables.mkdir()
            TypeDefinitionTemplate.ensure_exists(str(tables))
            definition = load_workbook(tables / "TypeDefinition.xlsx")
            definition["CODE"].append(("ID", "int", "integer alias"))
            definition["CODE"].append((
                "asset_path", "path(icons/,.png)", "asset path alias"
            ))
            definition["CODE"].append((
                "id_ref", "find_id(Target,Target,target_id)", "ID reference"
            ))
            definition["CODE"].append((
                "path_ref", "find_id(Target,Target,asset_key)", "path reference"
            ))
            definition.save(tables / "TypeDefinition.xlsx")
            definition.close()

            write_workbook(
                tables / "Target.xlsx",
                fields=(
                    ("id", "int", "CS"),
                    ("target_id", "ID", "CS"),
                    ("asset_key", "asset_path", "CS"),
                ),
                data_rows=((1, "001", "sword"),),
                code_rows=(("Target", "Target.json", "c"),),
                sheet="Target",
            )
            write_workbook(
                tables / "Source.xlsx",
                fields=(
                    ("id", "int", "CS"),
                    ("target_id", "id_ref", "CS"),
                    ("asset_key", "path_ref", "CS"),
                ),
                data_rows=((1, 1, "sword"),),
                code_rows=(("Source", "Source.json", "c"),),
                sheet="Source",
            )

            result = ExcelConverter().export_all(
                str(tables), str(client), str(root / "server"), "c"
            )

            self.assertTrue(result["success"], result["issues"])
            payload = json.loads(
                (client / "Source.json").read_text(encoding="utf-8")
            )["1"]
            self.assertIsInstance(payload["target_id"], dict)
            self.assertEqual(1, payload["target_id"]["target_id"])
            self.assertIsInstance(payload["asset_key"], dict)
            self.assertEqual("sword", payload["asset_key"]["asset_key"])

    def test_code_bytes_and_reference_failures_have_specific_issue_codes(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tables = root / "tables"
            tables.mkdir()
            TypeDefinitionTemplate.ensure_exists(str(tables))
            workbook = load_workbook(tables / "TypeDefinition.xlsx")
            workbook["CODE"].append(("missing_ref", "find_id(Missing,Missing,id)", ""))
            workbook.save(tables / "TypeDefinition.xlsx")
            workbook.close()
            write_workbook(
                tables / "Bytes.xlsx",
                fields=(("id", "int", "CS"), ("payload", "bytes", "CS")),
                data_rows=((1, "raw"),), code_rows=(("Item", "Bytes.json", "c"),),
            )
            write_workbook(
                tables / "Reference.xlsx",
                fields=(("id", "int", "CS"), ("target", "missing_ref", "CS")),
                data_rows=((1, 99),), code_rows=(("Item", "Reference.json", "c"),),
            )
            write_workbook(
                tables / "NoCode.xlsx",
                fields=(("id", "int", "CS"),), data_rows=((1,),), code_rows=None,
            )
            write_workbook(
                tables / "BadFormat.xlsx",
                fields=(("id", "int", "CS"),), data_rows=((1,),),
                code_rows=(("Item", "BadFormat.csv", "c"),),
            )

            result = ExcelConverter().export_all(
                str(tables), str(root / "client"), str(root / "server"),
                "c", validation_only=True,
            )

            codes = {issue["code"] for issue in result["issues"]}
            self.assertTrue({
                "BYTES_FORMAT_ERROR", "REFERENCE_TABLE_ERROR",
                "MISSING_CODE", "UNKNOWN_OUTPUT_FORMAT",
            }.issubset(codes), result["issues"])

    def test_primary_key_must_be_scalar_and_exported_to_every_target(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tables = root / "tables"
            tables.mkdir()
            write_workbook(
                tables / "Platform.xlsx",
                fields=(("id", "int", "C"), ("value", "str", "CS")),
                data_rows=((None, None), (1, "kept")),
                code_rows=(("Item", "Platform.json", "cs"),),
            )
            write_workbook(
                tables / "ListKey.xlsx",
                fields=(("ids", "intList", "CS"), ("value", "str", "CS")),
                data_rows=(("1#2", "bad key"),),
                code_rows=(("Item", "ListKey.json", "c"),),
            )

            result = ExcelConverter().export_all(
                str(tables), str(root / "client"), str(root / "server"),
                "cs", validation_only=True,
            )

            codes = {issue["code"] for issue in result["issues"]}
            self.assertIn("PRIMARY_KEY_NOT_EXPORTED", codes)
            self.assertIn("INVALID_PRIMARY_KEY", codes)
            self.assertNotIn("MISSING_PRIMARY_KEY", codes)
            invalid_type = next(
                issue for issue in result["issues"]
                if issue["code"] == "INVALID_PRIMARY_KEY"
            )
            self.assertEqual((invalid_type["row"], invalid_type["column"]), (2, 1))
            self.assertIn("intList", invalid_type["message"])
            self.assertIn("A2", invalid_type["message"])
            self.assertIn("int/string", invalid_type["message"])

    def test_repeated_row_errors_are_reported_once_per_field(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tables = root / "tables"
            tables.mkdir()
            write_workbook(
                tables / "Repeated.xlsx",
                fields=(
                    ("id", "int", "CS"),
                    ("count", "int", "CS"),
                    ("price", "int", "CS"),
                ),
                data_rows=tuple(
                    (row_id, f"bad-count-{row_id}", f"bad-price-{row_id}")
                    for row_id in range(1, 11)
                ),
                code_rows=(("Item", "Repeated.json", "c"),),
            )

            result = ExcelConverter().export_all(
                str(tables), str(root / "client"), str(root / "server"),
                "c", validation_only=True,
            )

            conversion_issues = [
                issue for issue in result["issues"]
                if issue["code"] == "CONVERSION_ERROR"
            ]
            self.assertEqual(len(conversion_issues), 2, result["issues"])
            self.assertEqual(
                {(issue["field"], issue["row"]) for issue in conversion_issues},
                {("count", 5), ("price", 5)},
            )

    def test_different_constraints_on_one_field_remain_separate(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tables = root / "tables"
            tables.mkdir()
            write_workbook(
                tables / "Constraints.xlsx",
                fields=(
                    ("id", "int", "CS"),
                    ("value", "int+range(1,2)+regex(^[A-Z]+$)", "CS"),
                ),
                data_rows=((1, 3), (2, 1)),
                code_rows=(("Item", "Constraints.json", "c"),),
            )

            result = ExcelConverter().export_all(
                str(tables), str(root / "client"), str(root / "server"),
                "c", validation_only=True,
            )

            issues = [
                issue for issue in result["issues"]
                if issue["code"] == "CONSTRAINT_ERROR"
            ]
            self.assertEqual(len(issues), 2, result["issues"])
            self.assertEqual({issue["row"] for issue in issues}, {5, 6})
            self.assertTrue(any("闭区间" in issue["message"] for issue in issues))
            self.assertTrue(any("正则表达式" in issue["message"] for issue in issues))

    def test_repeated_missing_and_duplicate_primary_keys_report_first_only(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tables = root / "tables"
            tables.mkdir()
            write_workbook(
                tables / "Primary.xlsx",
                fields=(("id", "int", "CS"), ("value", "str", "CS")),
                data_rows=(
                    (None, "missing one"),
                    (None, "missing two"),
                    *((1, f"duplicate {index}") for index in range(20)),
                ),
                code_rows=(("Item", "Primary.json", "c"),),
            )

            result = ExcelConverter().export_all(
                str(tables), str(root / "client"), str(root / "server"),
                "c", validation_only=True,
            )

            missing = [
                issue for issue in result["issues"]
                if issue["code"] == "MISSING_PRIMARY_KEY"
            ]
            duplicate = [
                issue for issue in result["issues"]
                if issue["code"] == "DUPLICATE_VALUE"
            ]
            self.assertEqual(len(missing), 1, result["issues"])
            self.assertEqual(missing[0]["row"], 5)
            self.assertEqual(len(duplicate), 1, result["issues"])
            self.assertEqual(duplicate[0]["row"], 8)

    def test_repeated_missing_references_report_first_row_only(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tables = root / "tables"
            tables.mkdir()
            TypeDefinitionTemplate.ensure_exists(str(tables))
            workbook = load_workbook(tables / "TypeDefinition.xlsx")
            workbook["CODE"].append(
                ("item_ref", "find_id(Target,Target,id)", "")
            )
            workbook.save(tables / "TypeDefinition.xlsx")
            workbook.close()
            write_workbook(
                tables / "Target.xlsx",
                fields=(("id", "int", "CS"),),
                data_rows=((1,),),
                code_rows=(("Target", "Target.json", "c"),),
                sheet="Target",
            )
            write_workbook(
                tables / "Source.xlsx",
                fields=(("id", "int", "CS"), ("target_id", "item_ref", "CS")),
                data_rows=((10, 99), (11, 98), (12, 97)),
                code_rows=(("Source", "Source.json", "c"),),
                sheet="Source",
            )

            result = ExcelConverter().export_all(
                str(tables), str(root / "client"), str(root / "server"),
                "c", validation_only=True,
            )

            issues = [
                issue for issue in result["issues"]
                if issue["code"] == "REFERENCE_NOT_FOUND"
            ]
            self.assertEqual(len(issues), 1, result["issues"])
            self.assertEqual(
                (issues[0]["field"], issues[0]["row"], issues[0]["column"]),
                ("target_id", 5, 2),
            )
            self.assertIn("='99'", issues[0]["message"])

    def test_same_workbook_reference_column_cannot_define_target_ids(self):
        result = export_same_workbook_reference(999999)

        issues = [
            issue for issue in result["issues"]
            if issue["code"] == "REFERENCE_NOT_FOUND"
        ]
        self.assertFalse(result["success"], result["issues"])
        self.assertEqual(len(issues), 1, result["issues"])
        self.assertEqual(
            (issues[0]["sheet"], issues[0]["field"], issues[0]["row"]),
            ("References", "target_id", 5),
        )
        self.assertIn("='999999'", issues[0]["message"])

    def test_same_workbook_reference_to_independent_id_passes(self):
        result = export_same_workbook_reference(1)

        self.assertTrue(result["success"], result["issues"])
        self.assertEqual([], result["issues"])

    def test_indirect_reference_alias_cannot_define_target_ids(self):
        result = export_same_workbook_reference(
            999999,
            reference_type="indirect_catalog_ref",
            extra_type_definitions=((
                "indirect_catalog_ref", "catalog_ref", "indirect reference alias"
            ),),
        )

        self.assertFalse(result["success"], result["issues"])
        self.assertIn(
            "REFERENCE_NOT_FOUND",
            {issue["code"] for issue in result["issues"]},
        )

    def test_reference_spacing_cannot_bypass_id_source_filter(self):
        result = export_same_workbook_reference(
            999999,
            catalog_reference_conversion=(
                "find_id (Catalog,DisplayOnly,target_id)"
            ),
        )

        self.assertFalse(result["success"], result["issues"])
        self.assertIn(
            "REFERENCE_NOT_FOUND",
            {issue["code"] for issue in result["issues"]},
        )

    def test_reference_target_without_independent_scalar_definition_is_an_error(self):
        result = export_same_workbook_reference(
            1, include_independent_definition=False
        )

        errors = [
            issue for issue in result["issues"]
            if issue["code"] == "REFERENCE_TABLE_ERROR"
        ]
        self.assertFalse(result["success"], result["issues"])
        self.assertEqual(1, len(errors), result["issues"])
        self.assertIn("没有独立标量定义列", errors[0]["message"])

    def test_string_reference_chain_without_independent_definition_is_an_error(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tables = root / "tables"
            tables.mkdir()
            TypeDefinitionTemplate.ensure_exists(str(tables))
            definition = load_workbook(tables / "TypeDefinition.xlsx")
            definition["CODE"].append((
                "other_ref", "find_id(Other,DisplayOnly,id)", ""
            ))
            definition["CODE"].append((
                "catalog_ref", "find_id(Catalog,DisplayOnly,target_id)", ""
            ))
            definition.save(tables / "TypeDefinition.xlsx")
            definition.close()
            write_workbook(
                tables / "Other.xlsx",
                fields=(("id", "string", "CS"),),
                data_rows=(("alpha",),),
                code_rows=(("Other", "Other.json", "c"),),
                sheet="Other",
            )
            write_workbook(
                tables / "Catalog.xlsx",
                fields=(
                    ("row_id", "int", "CS"),
                    ("target_id", "other_ref", "CS"),
                ),
                data_rows=((1, "alpha"),),
                code_rows=(("Catalog", "Catalog.json", "c"),),
                sheet="Catalog",
            )
            write_workbook(
                tables / "Source.xlsx",
                fields=(
                    ("row_id", "int", "CS"),
                    ("target_id", "catalog_ref", "CS"),
                ),
                data_rows=((1, "alpha"),),
                code_rows=(("Source", "Source.json", "c"),),
                sheet="Source",
            )

            result = ExcelConverter().export_all(
                str(tables), str(root / "client"), str(root / "server"),
                "c", validation_only=True,
            )

            errors = [
                issue for issue in result["issues"]
                if issue["code"] == "REFERENCE_TABLE_ERROR"
            ]
            self.assertFalse(result["success"], result["issues"])
            self.assertEqual(1, len(errors), result["issues"])
            self.assertIn("没有独立标量定义列", errors[0]["message"])

    def test_non_scalar_same_named_column_cannot_define_target_ids(self):
        result = export_same_workbook_reference(
            999999, additional_candidate_type="intList"
        )

        self.assertFalse(result["success"], result["issues"])
        self.assertIn(
            "REFERENCE_NOT_FOUND",
            {issue["code"] for issue in result["issues"]},
        )

    def test_reference_column_to_another_target_cannot_define_ids(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tables = root / "tables"
            tables.mkdir()
            TypeDefinitionTemplate.ensure_exists(str(tables))
            definition = load_workbook(tables / "TypeDefinition.xlsx")
            definition["CODE"].append((
                "catalog_ref", "find_id(Catalog,DisplayOnly,target_id)", ""
            ))
            definition["CODE"].append((
                "other_ref", "find_id(Other,DisplayOnly,id)", ""
            ))
            definition.save(tables / "TypeDefinition.xlsx")
            definition.close()
            write_workbook(
                tables / "Other.xlsx",
                fields=(("id", "int", "CS"),),
                data_rows=((999999,),),
                code_rows=(("Other", "Other.json", "c"),),
                sheet="Other",
            )
            workbook_path = write_workbook(
                tables / "Catalog.xlsx",
                fields=(
                    ("row_id", "int", "CS"),
                    ("target_id", "int", "CS"),
                ),
                data_rows=((1, 1),),
                code_rows=(
                    ("Definitions", "Definitions.json", "c"),
                    ("References", "References.json", "c"),
                    ("OtherReferences", "OtherReferences.json", "c"),
                ),
                sheet="Definitions",
            )
            workbook = load_workbook(workbook_path)
            references = workbook.create_sheet("References")
            references.append(("row_id", "target_id"))
            references.append(("int", "catalog_ref"))
            references.append(("CS", "CS"))
            references.append(("", ""))
            references.append((2, 999999))
            other_references = workbook.create_sheet("OtherReferences")
            other_references.append(("row_id", "target_id"))
            other_references.append(("int", "other_ref"))
            other_references.append(("CS", "CS"))
            other_references.append(("", ""))
            other_references.append((3, 999999))
            workbook.save(workbook_path)
            workbook.close()

            result = ExcelConverter().export_all(
                str(tables), str(root / "client"), str(root / "server"),
                "c", validation_only=True,
            )

            self.assertFalse(result["success"], result["issues"])
            missing = [
                issue for issue in result["issues"]
                if issue["code"] == "REFERENCE_NOT_FOUND"
                and issue["sheet"] == "References"
            ]
            self.assertEqual(1, len(missing), result["issues"])
            self.assertIn("='999999'", missing[0]["message"])

    def test_explicit_zero_references_are_validated(self):
        result = export_zero_references(1, 1.5)

        issues = [
            issue for issue in result["issues"]
            if issue["code"] == "REFERENCE_NOT_FOUND"
        ]
        self.assertFalse(result["success"], result["issues"])
        self.assertEqual({"int_id", "float_id"}, {
            issue["field"] for issue in issues
        })
        self.assertEqual({6}, {issue["row"] for issue in issues})

    def test_explicit_zero_references_pass_when_targets_define_zero(self):
        result = export_zero_references(0, 0.0)

        self.assertTrue(result["success"], result["issues"])
        self.assertEqual([], result["issues"])

    def test_split_list_reference_reports_any_missing_member(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tables = root / "tables"
            tables.mkdir()
            TypeDefinitionTemplate.ensure_exists(str(tables))
            definition = load_workbook(tables / "TypeDefinition.xlsx")
            definition["CODE"].append((
                "catalog_refs",
                "split_list(find_id(Target,DisplayOnly,target_id))",
                "",
            ))
            definition.save(tables / "TypeDefinition.xlsx")
            definition.close()
            write_workbook(
                tables / "Target.xlsx",
                fields=(
                    ("row_id", "int", "CS"),
                    ("target_id", "int", "CS"),
                ),
                data_rows=((1, 1),),
                code_rows=(("Target", "Target.json", "c"),),
                sheet="Target",
            )
            write_workbook(
                tables / "Source.xlsx",
                fields=(
                    ("row_id", "int", "CS"),
                    ("target_ids", "catalog_refs", "CS"),
                ),
                data_rows=((10, "1#99"),),
                code_rows=(("Source", "Source.json", "c"),),
                sheet="Source",
            )

            result = ExcelConverter().export_all(
                str(tables), str(root / "client"), str(root / "server"),
                "c", validation_only=True,
            )

            errors = [
                issue for issue in result["issues"]
                if issue["code"] == "REFERENCE_NOT_FOUND"
            ]
            self.assertFalse(result["success"], result["issues"])
            self.assertEqual(1, len(errors), result["issues"])
            self.assertEqual("target_ids", errors[0]["field"])
            self.assertIn("='99'", errors[0]["message"])

    def test_corrupt_workbook_is_read_and_reported_once(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tables = root / "tables"
            tables.mkdir()
            (tables / "Broken.xlsx").write_bytes(b"not an xlsx")

            result = ExcelConverter().export_all(
                str(tables), str(root / "client"), str(root / "server"), "c"
            )

            read_errors = [
                issue for issue in result["issues"]
                if issue["code"] == "WORKBOOK_READ_ERROR"
            ]
            self.assertEqual(len(read_errors), 1, result["issues"])
            self.assertNotIn("MISSING_CODE", {issue["code"] for issue in result["issues"]})


if __name__ == "__main__":
    unittest.main()
