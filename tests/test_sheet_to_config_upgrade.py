import json
import os
import string
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from openpyxl import Workbook, load_workbook
from PyQt5.QtGui import QPixmap
from PyQt5.QtWidgets import QApplication, QLabel

from sheet_to_config import i18n
from sheet_to_config.app_paths import migrate_legacy_cache, migrate_legacy_data
from sheet_to_config.dialogs import AboutDialog
from sheet_to_config.utils.exporter.template import TypeDefinitionTemplate
from sheet_to_config.utils.exporter.type_registry import TypeRegistry
from sheet_to_config.version import APP_NAME, GITHUB_URL, resource_path


ROOT = Path(__file__).resolve().parents[1]


class SheetToConfigUpgradeTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.app = QApplication.instance() or QApplication([])

    def test_catalogs_have_identical_keys_and_placeholders(self):
        catalogs = {}
        for locale_id in i18n.SUPPORTED_LOCALES:
            path = ROOT / "sheet_to_config" / "i18n" / "catalogs" / f"{locale_id}.json"
            pairs = json.loads(
                path.read_text(encoding="utf-8"),
                object_pairs_hook=lambda items: items,
            )
            keys = [key for key, _ in pairs]
            self.assertEqual(len(keys), len(set(keys)), f"duplicate keys in {locale_id}")
            catalogs[locale_id] = dict(pairs)
        self.assertTrue(all(set(value) == set(catalogs["en"]) for value in catalogs.values()))
        for locale_id, catalog in catalogs.items():
            self.assertNotIn(".exe", catalog["about.update_hint"].casefold(), locale_id)
        formatter = string.Formatter()
        for key in catalogs["en"]:
            expected = {name for _, name, _, _ in formatter.parse(catalogs["en"][key]) if name}
            for locale_id, catalog in catalogs.items():
                actual = {name for _, name, _, _ in formatter.parse(catalog[key]) if name}
                self.assertEqual(expected, actual, (locale_id, key))

    def test_locale_selection_persists_without_touching_other_settings(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            old_path = i18n._settings_path
            old_locale = i18n._current_locale
            try:
                settings = Path(temp_dir) / "config.json"
                settings.write_text('{"keep": 7}', encoding="utf-8")
                i18n._settings_path = settings
                i18n.set_locale("es")
                self.assertEqual("es", i18n.get_locale())
                self.assertEqual(7, json.loads(settings.read_text(encoding="utf-8"))["keep"])
            finally:
                i18n._settings_path = old_path
                i18n._current_locale = old_locale

    def test_legacy_state_and_cache_migrate_without_overwrite(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            old = root / "old"
            new = root / "new"
            old.mkdir()
            new.mkdir()
            (old / "projects.json").write_text("old projects", encoding="utf-8")
            (old / "theme_config.json").write_text("old theme", encoding="utf-8")
            (new / "theme_config.json").write_text("new theme", encoding="utf-8")
            copied = migrate_legacy_data(new, [old])
            self.assertTrue((new / "projects.json").is_file(), copied)
            self.assertEqual("new theme", (new / "theme_config.json").read_text(encoding="utf-8"))

            (root / "TableManager_bg_cache.png").write_bytes(b"legacy cache")
            self.assertEqual(root / "SheetToConfig_bg_cache.png", migrate_legacy_cache(root))
            self.assertEqual(b"legacy cache", (root / "SheetToConfig_bg_cache.png").read_bytes())

    def test_type_definition_is_localized_and_extended_non_destructively(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            table_dir = Path(temp_dir)
            TypeDefinitionTemplate.ensure_exists(str(table_dir), locale="ja")
            path = table_dir / "TypeDefinition.xlsx"
            workbook = load_workbook(path)
            self.assertEqual(["名前", "変換関数", "説明"], [cell.value for cell in workbook["CODE"][1]])
            workbook["CODE"].append(("custom", "string", "keep me"))
            original_rows = workbook["CODE"].max_row
            workbook.save(path)
            workbook.close()
            original_bytes = path.read_bytes()

            TypeDefinitionTemplate.ensure_exists(str(table_dir), locale="es")
            self.assertEqual(original_bytes, path.read_bytes())
            workbook = load_workbook(path)
            rows = list(workbook["CODE"].values)
            self.assertIn(("custom", "string", "keep me"), rows)
            self.assertGreaterEqual(workbook["CODE"].max_row, original_rows)
            workbook.close()
            self.assertTrue(TypeRegistry(str(table_dir)).has_type("int"))

    def test_corrupt_type_definition_is_not_replaced(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "TypeDefinition.xlsx"
            payload = b"not an xlsx"
            path.write_bytes(payload)
            with self.assertRaises(Exception):
                TypeDefinitionTemplate.ensure_exists(temp_dir, locale="en")
            self.assertEqual(payload, path.read_bytes())

    def test_product_identity_readmes_and_qr_resources(self):
        self.assertEqual("SheetToConfig", APP_NAME)
        self.assertEqual("https://github.com/liushafeiniao/SheetToConfig", GITHUB_URL)
        for name in ("alipay.png", "wechat.png"):
            path = Path(resource_path(f"assets/donate/{name}"))
            if path.is_file():
                self.assertFalse(QPixmap(str(path)).isNull(), path)
        with tempfile.TemporaryDirectory() as temp_dir, patch(
            "sheet_to_config.version.resource_path",
            side_effect=lambda relative: str(Path(temp_dir) / relative),
        ):
            dialog = AboutDialog()
            placeholders = [
                label
                for label in dialog.findChildren(QLabel)
                if label.text() == i18n.tr("donate.missing")
            ]
            self.assertEqual(2, len(placeholders))
            dialog.close()
        readmes = [ROOT / "README.md"]
        readmes.extend(sorted((ROOT / "docs" / "locales").glob("README*.md")))
        self.assertEqual(6, len(readmes))
        for readme in readmes:
            text = readme.read_text(encoding="utf-8")
            self.assertIn("SheetToConfig", text)
            self.assertIn("find_id(file_prefix, display_label, field)", text)
            self.assertNotIn("find_obj", text)
            self.assertNotIn("find_obg", text)
            self.assertNotIn("macos-preview", text)

    def test_packaged_resource_path_uses_the_package_data_directory(self):
        with tempfile.TemporaryDirectory() as temp_dir, patch.object(
            sys, "_MEIPASS", temp_dir, create=True
        ):
            self.assertEqual(
                Path(temp_dir) / "sheet_to_config" / "assets" / "donate" / "alipay.png",
                Path(resource_path("assets/donate/alipay.png")),
            )


if __name__ == "__main__":
    unittest.main()
