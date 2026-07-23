import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from sheet_to_config.utils.exporter import atomic_writer
from sheet_to_config.utils.exporter.converter import ExcelConverter
from sheet_to_config.utils.exporter.examples import (
    EXAMPLE_FILENAMES,
    create_example_workbooks,
)
from sheet_to_config.utils.exporter.template import TypeDefinitionTemplate
from sheet_to_config.utils.exporter.template_samples import sample_sheet_names


def workbook_snapshot(path: Path):
    workbook = load_workbook(path, data_only=False)
    try:
        return tuple(
            (
                ws.title,
                tuple(tuple(cell.value for cell in row) for row in ws.iter_rows()),
                str(ws.freeze_panes or ""),
                ws.auto_filter.ref or "",
            )
            for ws in workbook.worksheets
        )
    finally:
        workbook.close()


def copy_teaching_sheets_to_business_workbook(
    tables: Path, locale_id: str
) -> None:
    definition = load_workbook(tables / "TypeDefinition.xlsx", data_only=False)
    business = Workbook()
    try:
        item_sheet, reward_sheet = sample_sheet_names(locale_id)
        for index, sheet_name in enumerate((item_sheet, reward_sheet)):
            target = business.active if index == 0 else business.create_sheet()
            target.title = sheet_name
            for row in definition[sheet_name].iter_rows(values_only=True):
                target.append(list(row))
        code = business.create_sheet("CODE")
        code.append(("Sheet", "Output file", "Platform"))
        code.append((item_sheet, "Item.json", "cs"))
        code.append((reward_sheet, "Reward.json", "cs"))
        business.save(tables / "GameConfig.xlsx")
    finally:
        definition.close()
        business.close()


class ExampleWorkbookTests(unittest.TestCase):
    def test_generator_requires_a_missing_or_empty_root_by_default(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir) / "sample"
            root.mkdir()
            marker = root / "notes.txt"
            marker.write_text("keep", encoding="utf-8")

            with self.assertRaises(FileExistsError):
                create_example_workbooks(root)

            self.assertEqual("keep", marker.read_text(encoding="utf-8"))
            self.assertFalse((root / "tables").exists())

    def test_force_replaces_only_owned_files_and_preserves_unrelated_content(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir) / "sample"
            tables = root / "tables"
            tables.mkdir(parents=True)
            marker = root / "notes.txt"
            marker.write_text("keep", encoding="utf-8")
            unrelated = tables / "Other.xlsx"
            unrelated.write_bytes(b"unrelated")
            for name in EXAMPLE_FILENAMES:
                (tables / name).write_bytes(b"old")

            written = create_example_workbooks(root, force=True)

            self.assertEqual("keep", marker.read_text(encoding="utf-8"))
            self.assertEqual(b"unrelated", unrelated.read_bytes())
            self.assertEqual(set(EXAMPLE_FILENAMES), set(written))
            self.assertTrue(all(path.stat().st_size > 1000 for path in written.values()))
            self.assertFalse((root / "client").exists())
            self.assertFalse((root / "server").exists())

    def test_failed_atomic_commit_removes_only_new_empty_directories(self):
        real_commit = atomic_writer.commit_files

        def fail_during_commit(writes, deletes):
            real_replace = atomic_writer.os.replace
            call_count = 0

            def flaky_replace(source, destination):
                nonlocal call_count
                call_count += 1
                if call_count == 1:
                    raise OSError("simulated commit failure")
                return real_replace(source, destination)

            with patch.object(atomic_writer.os, "replace", side_effect=flaky_replace):
                return real_commit(writes, deletes)

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir) / "new-root"
            with patch.object(atomic_writer, "commit_files", side_effect=fail_during_commit):
                with self.assertRaises(atomic_writer.AtomicCommitError):
                    create_example_workbooks(root)
            self.assertFalse(root.exists())

    def test_embedded_sheets_are_copyable_to_one_runnable_business_workbook(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir) / "sample"
            create_example_workbooks(root, locale="zh-CN")

            tables = root / "tables"
            self.assertEqual(
                {"TypeDefinition.xlsx"},
                {path.name for path in tables.glob("*.xlsx")},
            )
            copy_teaching_sheets_to_business_workbook(tables, "zh-CN")

            result = ExcelConverter().export_all(
                str(tables),
                str(root / "client"),
                str(root / "server"),
                "cs",
            )

            self.assertTrue(result["success"], result["issues"])
            client_reward = json.loads(
                (root / "client" / "Reward.json").read_text(encoding="utf-8")
            )
            server_reward = json.loads(
                (root / "server" / "Reward.json").read_text(encoding="utf-8")
            )
            self.assertEqual([1001, 1002], client_reward["1"]["itemIds"])
            self.assertEqual([70, 30], client_reward["1"]["weights"])
            self.assertEqual(1001, client_reward["1"]["primaryItemId"])
            self.assertNotIn("rate", client_reward["1"])
            self.assertEqual(0.25, server_reward["1"]["rate"])
            self.assertNotIn("itemIds", server_reward["1"])
            self.assertFalse(result["issues"])

    def test_every_locale_teaching_sheets_validate_after_copy(self):
        for locale_id in TypeDefinitionTemplate.HEADER_LABELS:
            with self.subTest(locale=locale_id), tempfile.TemporaryDirectory() as temp_dir:
                root = Path(temp_dir) / "sample"
                create_example_workbooks(root, locale=locale_id)
                tables = root / "tables"
                copy_teaching_sheets_to_business_workbook(tables, locale_id)

                result = ExcelConverter().export_all(
                    str(tables),
                    str(root / "client"),
                    str(root / "server"),
                    "cs",
                    validation_only=True,
                )

                self.assertTrue(result["success"], result["issues"])
                self.assertFalse(result["issues"])

    def test_type_definition_teaching_sheets_are_intentionally_skipped(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir) / "sample"
            create_example_workbooks(root, locale="zh-CN")

            result = ExcelConverter().export_all(
                str(root / "tables"),
                str(root / "client"),
                str(root / "server"),
                "cs",
            )

            self.assertFalse(result["success"])
            self.assertEqual("NO_WORKBOOKS", result["issues"][0]["code"])
            self.assertFalse((root / "client").exists())
            self.assertFalse((root / "server").exists())

    def test_generation_is_logically_reproducible(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            first = Path(temp_dir) / "first"
            second = Path(temp_dir) / "second"
            create_example_workbooks(first, locale="en")
            create_example_workbooks(second, locale="en")
            for name in EXAMPLE_FILENAMES:
                with self.subTest(name=name):
                    self.assertEqual(
                        workbook_snapshot(first / "tables" / name),
                        workbook_snapshot(second / "tables" / name),
                    )


if __name__ == "__main__":
    unittest.main()
