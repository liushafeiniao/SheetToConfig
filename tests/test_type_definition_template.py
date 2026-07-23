import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from sheet_to_config.utils.exporter.template import (
    TEMPLATE_THEME,
    TypeDefinitionTemplate,
)
from sheet_to_config.utils.exporter.template_samples import (
    SAMPLE_COPY,
    item_reference_expression,
    localized_sample_type_definitions,
    localized_sample_type_names,
    sample_sheet_names,
)


class TypeDefinitionTemplateTests(unittest.TestCase):
    def test_all_supported_locales_localize_headers_aliases_and_example_rows(self):
        for locale_id, expected_headers in TypeDefinitionTemplate.HEADER_LABELS.items():
            with self.subTest(locale=locale_id), tempfile.TemporaryDirectory() as temp_dir:
                TypeDefinitionTemplate.create_template(temp_dir, locale=locale_id)
                workbook = load_workbook(Path(temp_dir) / "TypeDefinition.xlsx")
                item_sheet, reward_sheet = sample_sheet_names(locale_id)
                self.assertEqual(
                    ["CODE", "Guide", "Examples", item_sheet, reward_sheet],
                    workbook.sheetnames,
                )
                self.assertEqual(
                    list(expected_headers),
                    [cell.value for cell in workbook["CODE"][1]],
                )
                values = list(workbook["CODE"].values)[1:]
                localized_definitions = localized_sample_type_definitions(locale_id)
                self.assertEqual(
                    len(TypeDefinitionTemplate.DEFAULT_TYPES) + len(localized_definitions),
                    len(values),
                )
                self.assertTrue(all(len(row) == 4 and row[3] for row in values))
                registered = {row[0]: row[1] for row in values}
                for name, target, _description, _example in localized_definitions:
                    self.assertEqual(target, registered[name])

                type_names = localized_sample_type_names(locale_id)
                self.assertEqual(
                    [
                        "int+unique()",
                        f"{type_names['text']}+required()+unique()",
                        type_names["quality"],
                        type_names["icon"],
                    ],
                    [cell.value for cell in workbook[item_sheet][2]],
                )
                self.assertEqual(
                    [
                        f"{type_names['integer']}+unique()",
                        f"{type_names['item_reference']}+required()",
                        f"{type_names['item_reference_list']}+len(1,3)",
                        f"{type_names['integer_list']}+equalLen(itemIds)",
                        f"{type_names['number']}+range(0,1)",
                    ],
                    [cell.value for cell in workbook[reward_sheet][2]],
                )
                self.assertEqual(
                    list(SAMPLE_COPY[locale_id]["item"]),
                    [cell.value for cell in workbook[item_sheet][4]],
                )
                self.assertEqual(
                    [
                        value.format(item_sheet=item_sheet)
                        for value in SAMPLE_COPY[locale_id]["reward"]
                    ],
                    [cell.value for cell in workbook[reward_sheet][4]],
                )
                self.assertEqual(
                    ["itemId", "name", "quality", "icon"],
                    [cell.value for cell in workbook[item_sheet][1]],
                )
                self.assertEqual(
                    ["rewardId", "primaryItemId", "itemIds", "weights", "rate"],
                    [cell.value for cell in workbook[reward_sheet][1]],
                )
                self.assertIn(item_reference_expression(locale_id), {
                    row[2] for row in workbook["Examples"].iter_rows(
                        min_row=5, values_only=True
                    )
                })
                workbook.close()

    def test_unknown_locale_falls_back_to_english(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            TypeDefinitionTemplate.create_template(temp_dir, locale="xx-XX")
            workbook = load_workbook(Path(temp_dir) / "TypeDefinition.xlsx")
            self.assertEqual(
                list(TypeDefinitionTemplate.HEADER_LABELS["en"]),
                [cell.value for cell in workbook["CODE"][1]],
            )
            workbook.close()

    def test_guide_documents_constraints_references_and_unsupported_syntax(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            TypeDefinitionTemplate.create_template(temp_dir, locale="zh-CN")
            workbook = load_workbook(Path(temp_dir) / "TypeDefinition.xlsx")
            cells = [
                str(cell.value)
                for row in workbook["Guide"].iter_rows()
                for cell in row
                if cell.value is not None
            ]
            joined = "\n".join(cells)
            for token in (
                "len(1,5)", "len2(1,3)", "len3(1,2)", "equalLen",
                "coexist", "leastOne", "required", "notEmpty", "unique",
                "find_id(GameConfig,物品表,itemId)", "file_prefix", "display_label", "field",
                "nullable()", "bytes", "split_dict",
            ):
                self.assertIn(token, joined)
            self.assertIn("不要直接填写 find_id", joined)
            self.assertIn("导出时会跳过 TypeDefinition.xlsx", joined)
            self.assertIn("物品ID+required()", joined)
            self.assertIn("+ 后的约束名是固定语法", joined)
            workbook.close()

    def test_item_and_reward_are_copyable_multisheet_examples(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            TypeDefinitionTemplate.create_template(temp_dir, locale="zh-CN")
            workbook = load_workbook(Path(temp_dir) / "TypeDefinition.xlsx")
            self.assertEqual(
                ["itemId", "name", "quality", "icon"],
                [cell.value for cell in workbook["物品表"][1]],
            )
            self.assertEqual(
                [
                    "rewardId", "primaryItemId", "itemIds", "weights", "rate",
                ],
                [cell.value for cell in workbook["奖励表"][1]],
            )
            self.assertEqual(
                "物品ID+required()", workbook["奖励表"]["B2"].value
            )
            self.assertEqual(
                "物品列表+len(1,3)", workbook["奖励表"]["C2"].value
            )
            registered = {
                row[0]: row[1]
                for row in workbook["CODE"].iter_rows(min_row=2, values_only=True)
            }
            self.assertEqual(
                "find_id(GameConfig,物品表,itemId)", registered["物品ID"]
            )
            self.assertNotIn("物品引用", registered)
            self.assertNotIn("物品引用列表", registered)
            workbook.close()

    def test_new_workbook_has_readable_navigation_and_centralized_colors(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            TypeDefinitionTemplate.create_template(temp_dir, locale="en")
            workbook = load_workbook(Path(temp_dir) / "TypeDefinition.xlsx")
            expected = {
                "CODE": ("A2", TEMPLATE_THEME["tab_code"], True),
                "Guide": ("A5", TEMPLATE_THEME["tab_guide"], True),
                "Examples": ("A5", TEMPLATE_THEME["tab_examples"], True),
                "Item": ("A5", TEMPLATE_THEME["tab_item"], False),
                "Reward": ("A5", TEMPLATE_THEME["tab_reward"], False),
            }
            for name, (freeze, tab_color, has_filter) in expected.items():
                ws = workbook[name]
                self.assertEqual(freeze, ws.freeze_panes)
                self.assertFalse(ws.sheet_view.showGridLines)
                self.assertEqual(tab_color, ws.sheet_properties.tabColor.rgb[-6:])
                self.assertEqual(has_filter, bool(ws.auto_filter.ref))
                self.assertGreater(ws.column_dimensions["A"].width, 10)
            self.assertGreaterEqual(workbook["Item"].row_dimensions[4].height, 50)
            self.assertEqual(
                TEMPLATE_THEME["header_fill"],
                workbook["CODE"]["A1"].fill.fgColor.rgb[-6:],
            )
            workbook.close()

    def test_existing_two_and_three_column_code_sheets_remain_compatible(self):
        for column_count in (2, 3):
            with self.subTest(columns=column_count), tempfile.TemporaryDirectory() as temp_dir:
                path = Path(temp_dir) / "TypeDefinition.xlsx"
                workbook = Workbook()
                code = workbook.active
                code.title = "CODE"
                headers = ("Name", "Convert", "Description")[:column_count]
                custom = ("custom", "string", "keep me")[:column_count]
                code.append(headers)
                code.append(custom)
                guide = workbook.create_sheet("Guide")
                guide["A1"] = "keep guide"
                examples = workbook.create_sheet("Examples")
                examples["A1"] = "keep examples"
                workbook.save(path)
                workbook.close()

                TypeDefinitionTemplate.create_template(temp_dir, locale="en")
                workbook = load_workbook(path)
                self.assertEqual(column_count, workbook["CODE"].max_column)
                self.assertEqual(custom, tuple(
                    cell.value for cell in workbook["CODE"][2][:column_count]
                ))
                self.assertEqual("keep guide", workbook["Guide"]["A1"].value)
                self.assertEqual("keep examples", workbook["Examples"]["A1"].value)
                self.assertIn("Item", workbook.sheetnames)
                self.assertIn("Reward", workbook.sheetnames)
                self.assertIn("bytes", {
                    row[0] for row in workbook["CODE"].iter_rows(
                        min_row=2, values_only=True
                    )
                })
                workbook.close()


if __name__ == "__main__":
    unittest.main()
