import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from utils.exporter.converter import ExcelConverter
from utils.exporter.exporters.protobuf_exporter import ProtobufExporter
from utils.exporter.protobuf_schema import ProtoSchemaParser
from utils.exporter.reader import CodeSheet
from utils.exporter.template import TypeDefinitionTemplate

from tests.workbook_factory import write_workbook


def platform_proto_rows(source_for_id="id"):
    return [
        ("root", "Item", "ItemTable", "rows", 1, "ItemRow", "repeated", "", "$rows", ""),
        ("field", "Item", "ItemRow", "id", 1, "int32", "singular", "", source_for_id, ""),
        ("field", "Item", "ItemRow", "client_name", 2, "string", "singular", "", "client_name", ""),
        ("field", "Item", "ItemRow", "server_note", 3, "string", "singular", "", "server_note", ""),
        ("field", "Item", "ItemRow", "shared_values", 4, "int32", "repeated", "", "shared_values", ""),
    ]


class CodeSheetFormatTests(unittest.TestCase):
    def test_recognizes_explicit_formats_and_legacy_extensionless_json(self):
        cases = {
            "Item.PB": "pb",
            "Item.JSON": "json",
            "Item.LUA": "lua",
            "Item.proto": "proto",
            "Item": "json",
        }

        for output_name, expected_format in cases.items():
            with self.subTest(output_name=output_name):
                code = CodeSheet.from_row(["Item", output_name, "cs"])
                self.assertEqual(expected_format, code.format)
                self.assertEqual(output_name == "Item", code.implicit_format)


class ConverterProtobufIntegrationTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)
        self.root = Path(self.temp_dir.name)
        self.tables = self.root / "tables"
        self.client = self.root / "client"
        self.server = self.root / "server"
        self.tables.mkdir()

    def _write_table(self, *, proto_rows):
        return write_workbook(
            self.tables / "Item.xlsx",
            fields=(
                ("id", "int", "CS"),
                ("client_name", "string", "C"),
                ("server_note", "string", "S"),
                ("shared_values", "intList", "CS"),
            ),
            data_rows=(
                (2, "Client Two", "Server Two", "2#20"),
                (1, "Client One", "Server One", "1#10"),
            ),
            proto_rows=proto_rows,
            code_rows=(("Item", "Item.pb", "cs"),),
        )

    @staticmethod
    def _decode(schema, payload):
        exporter = ProtobufExporter(schema)
        message = exporter._message_class()
        message.ParseFromString(payload)
        return message

    def test_code_pb_exports_shared_schema_and_platform_filtered_payloads(self):
        workbook_path = self._write_table(proto_rows=platform_proto_rows())
        logs = []

        result = ExcelConverter(logs.append).export_all(
            str(self.tables), str(self.client), str(self.server), "cs"
        )

        self.assertTrue(result["success"], logs)
        self.assertEqual(1, result["success_count"])
        client_proto = self.client / "Item.proto"
        client_pb = self.client / "Item.pb"
        server_proto = self.server / "Item.proto"
        server_pb = self.server / "Item.pb"
        for path in (client_proto, client_pb, server_proto, server_pb):
            self.assertTrue(path.is_file(), path)
        self.assertFalse((self.client / "Item.pb.pb").exists())
        self.assertEqual(client_proto.read_bytes(), server_proto.read_bytes())

        schema = ProtoSchemaParser.parse_workbook(str(workbook_path), "Item")
        client_data = self._decode(schema, client_pb.read_bytes())
        server_data = self._decode(schema, server_pb.read_bytes())
        self.assertEqual([2, 1], [row.id for row in client_data.rows])
        self.assertEqual([2, 1], [row.id for row in server_data.rows])
        self.assertEqual("Client Two", client_data.rows[0].client_name)
        self.assertEqual("", client_data.rows[0].server_note)
        self.assertEqual("", server_data.rows[0].client_name)
        self.assertEqual("Server Two", server_data.rows[0].server_note)
        self.assertEqual([2, 20], list(client_data.rows[0].shared_values))
        self.assertEqual([2, 20], list(server_data.rows[0].shared_values))

    def test_proto_code_entry_fails_with_actionable_message(self):
        write_workbook(
            self.tables / "Item.xlsx",
            fields=(("id", "int", "CS"),),
            data_rows=((1,),),
            proto_rows=None,
            code_rows=(("Item", "Item.proto", "cs"),),
        )
        logs = []

        result = ExcelConverter(logs.append).export_all(
            str(self.tables), str(self.client), str(self.server), "cs"
        )

        self.assertFalse(result["success"])
        self.assertIn(".proto", "\n".join(logs))
        self.assertIn(".pb", "\n".join(logs))

    def test_client_mode_does_not_write_server_artifacts(self):
        self._write_table(proto_rows=platform_proto_rows())
        logs = []

        result = ExcelConverter(logs.append).export_all(
            str(self.tables), str(self.client), str(self.server), "c"
        )

        self.assertTrue(result["success"], logs)
        self.assertTrue((self.client / "Item.pb").is_file())
        self.assertFalse((self.server / "Item.pb").exists())
        self.assertFalse((self.server / "Item.proto").exists())

    def test_optional_empty_cell_is_absent_and_bytes_use_utf8(self):
        proto_rows = [
            ("root", "Item", "ItemTable", "rows", 1, "ItemRow", "repeated", "", "$rows", ""),
            ("field", "Item", "ItemRow", "id", 1, "int32", "singular", "", "id", ""),
            ("field", "Item", "ItemRow", "enabled", 2, "bool", "optional", "", "enabled", ""),
            ("field", "Item", "ItemRow", "payload", 3, "bytes", "singular", "", "payload", ""),
        ]
        workbook_path = write_workbook(
            self.tables / "Item.xlsx",
            fields=(("id", "int", "CS"), ("enabled", "bool", "CS"), ("payload", "bytes", "CS")),
            data_rows=((1, None, "你好"),),
            proto_rows=proto_rows,
            code_rows=(("Item", "Item.pb", "c"),),
        )
        # Simulate a pre-Protobuf TypeDefinition.xlsx without the new bytes row.
        TypeDefinitionTemplate.ensure_exists(str(self.tables))
        type_path = self.tables / "TypeDefinition.xlsx"
        type_book = load_workbook(type_path)
        type_sheet = type_book["CODE"]
        for row_index in range(type_sheet.max_row, 1, -1):
            if type_sheet.cell(row_index, 1).value == "bytes":
                type_sheet.delete_rows(row_index)
        type_book.save(type_path)
        type_book.close()
        logs = []

        result = ExcelConverter(logs.append).export_all(
            str(self.tables), str(self.client), str(self.server), "c"
        )

        self.assertTrue(result["success"], logs)
        schema = ProtoSchemaParser.parse_workbook(str(workbook_path), "Item")
        decoded = self._decode(schema, (self.client / "Item.pb").read_bytes()).rows[0]
        self.assertFalse(decoded.HasField("enabled"))
        self.assertEqual("你好".encode("utf-8"), decoded.payload)

    def test_json_and_lua_outputs_remain_functional(self):
        write_workbook(
            self.tables / "Item.xlsx",
            fields=(("id", "int", "CS"), ("name", "str", "CS")),
            data_rows=((1, "Sword"),),
            code_rows=(("Item", "Item.json", "cs"), ("Item", "Item.lua", "cs")),
        )
        logs = []

        result = ExcelConverter(logs.append).export_all(
            str(self.tables), str(self.client), str(self.server), "c"
        )

        self.assertTrue(result["success"], logs)
        self.assertEqual("Sword", json.loads((self.client / "Item.json").read_text(encoding="utf-8"))["1"]["name"])
        self.assertIn('name = "Sword"', (self.client / "Item.lua").read_text(encoding="utf-8"))
        self.assertFalse((self.client / "Item.json.json").exists())
        self.assertFalse(self.server.exists())

    def test_reference_type_is_validated_then_serialized_as_target_id_type(self):
        proto_rows = [
            ("root", "Item", "ItemTable", "rows", 1, "ItemRow", "repeated", "", "$rows", ""),
            ("field", "Item", "ItemRow", "id", 1, "int32", "singular", "", "id", ""),
            ("field", "Item", "ItemRow", "ref_id", 2, "int32", "singular", "", "ref_id", ""),
        ]
        workbook_path = write_workbook(
            self.tables / "Item.xlsx",
            fields=(("id", "int", "CS"), ("ref_id", "item_ref", "CS")),
            data_rows=((2, 1), (1, 2)),
            proto_rows=proto_rows,
            code_rows=(("Item", "Item.pb", "c"),),
        )
        TypeDefinitionTemplate.ensure_exists(str(self.tables))
        type_path = self.tables / "TypeDefinition.xlsx"
        workbook = load_workbook(type_path)
        workbook["CODE"].append(("item_ref", "find_id(Item,Item,id)", ""))
        workbook.save(type_path)
        workbook.close()

        logs = []
        result = ExcelConverter(logs.append).export_all(
            str(self.tables), str(self.client), str(self.server), "c"
        )

        self.assertTrue(result["success"], logs)
        schema = ProtoSchemaParser.parse_workbook(str(workbook_path), "Item")
        decoded = self._decode(schema, (self.client / "Item.pb").read_bytes())
        self.assertEqual([(2, 1), (1, 2)], [(row.id, row.ref_id) for row in decoded.rows])

    def test_invalid_source_fails_without_overwriting_existing_outputs(self):
        self._write_table(proto_rows=platform_proto_rows(source_for_id="missing_id"))
        self.client.mkdir()
        old_proto = self.client / "Item.proto"
        old_pb = self.client / "Item.pb"
        old_proto.write_text("old proto", encoding="utf-8")
        old_pb.write_bytes(b"old pb")
        logs = []

        result = ExcelConverter(logs.append).export_all(
            str(self.tables), str(self.client), str(self.server), "c"
        )

        self.assertFalse(result["success"])
        self.assertIn("missing_id", "\n".join(logs))
        self.assertEqual("old proto", old_proto.read_text(encoding="utf-8"))
        self.assertEqual(b"old pb", old_pb.read_bytes())


if __name__ == "__main__":
    unittest.main()
