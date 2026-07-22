"""Small workbook fixtures used by Protobuf tests."""

from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook


PROTO_HEADERS = (
    "kind",
    "sheet",
    "message",
    "field",
    "number",
    "type",
    "rule",
    "oneof",
    "source",
    "description",
)


def base_proto_rows(sheet: str = "Item") -> list[tuple[object, ...]]:
    return [
        ("root", sheet, "ItemTable", "rows", 1, "ItemRow", "repeated", "", "$rows", ""),
        ("field", sheet, "ItemRow", "id", 1, "int32", "singular", "", "id", ""),
        ("field", sheet, "ItemRow", "name", 2, "string", "singular", "", "name", ""),
    ]


def write_workbook(
    path: Path,
    *,
    fields: Sequence[tuple[str, str, str]],
    data_rows: Iterable[Sequence[object]],
    proto_rows: Iterable[Sequence[object]] | None = None,
    code_rows: Iterable[Sequence[object]] | None = None,
    sheet: str = "Item",
    package: str = "config",
    csharp_namespace: str = "Game.Config",
) -> Path:
    """Create a four-line-header workbook without touching repository data."""
    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = sheet
    data_sheet.append([field[0] for field in fields])
    data_sheet.append([field[1] for field in fields])
    data_sheet.append([field[2] for field in fields])
    data_sheet.append(["" for _ in fields])
    for row in data_rows:
        data_sheet.append(list(row))

    if proto_rows is not None:
        proto_sheet = workbook.create_sheet("PROTO")
        proto_sheet.cell(1, 1, "package")
        proto_sheet.cell(1, 2, package)
        proto_sheet.cell(2, 1, "csharp_namespace")
        proto_sheet.cell(2, 2, csharp_namespace)
        proto_sheet.append([])
        proto_sheet.append(PROTO_HEADERS)
        for row in proto_rows:
            proto_sheet.append(list(row))

    if code_rows is not None:
        code_sheet = workbook.create_sheet("CODE")
        code_sheet.append(("Sheet", "output", "platform"))
        for row in code_rows:
            code_sheet.append(list(row))

    workbook.save(path)
    workbook.close()
    return path
