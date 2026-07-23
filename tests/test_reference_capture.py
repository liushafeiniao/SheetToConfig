import json
import tempfile
import unittest
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

from openpyxl import load_workbook

from sheet_to_config.utils.exporter.converter import ExcelConverter
from sheet_to_config.utils.exporter.template import TypeDefinitionTemplate
from sheet_to_config.utils.exporter.type_registry import TypeRegistry
from tests.workbook_factory import write_workbook


class ReferenceCaptureTests(unittest.TestCase):
    def setUp(self):
        self._temp_dir = tempfile.TemporaryDirectory()
        self.tables = Path(self._temp_dir.name) / "tables"
        self.tables.mkdir()
        TypeDefinitionTemplate.ensure_exists(str(self.tables))
        definition = load_workbook(self.tables / "TypeDefinition.xlsx")
        for row in (
            ("bool_alias", "bool", ""),
            ("direct_ref", "find_id(Target,DisplayOnly,first_id)", ""),
            ("name_ref", "find_id(Target,DisplayOnly,name)", ""),
            ("typed_ref", "find_id(Target,DisplayOnly,flag)", ""),
            ("list_ref", "split_list(find(Target,DisplayOnly,first_id))", ""),
            ("alias_ref", "direct_ref", ""),
            (
                "dict_positional",
                "split_dict(find_id(Target,DisplayOnly,first_id) first;"
                "find_id(Target,DisplayOnly,second_id) second)",
                "",
            ),
            (
                "dict_named",
                "split_dict(first:find_id(Target,DisplayOnly,first_id);"
                "second:find_id(Target,DisplayOnly,second_id))",
                "",
            ),
            (
                "nested_refs",
                "split_list2(split_list(find_id(Target,DisplayOnly,first_id)))",
                "",
            ),
        ):
            definition["CODE"].append(row)
        definition.save(self.tables / "TypeDefinition.xlsx")
        definition.close()
        write_workbook(
            self.tables / "Target.xlsx",
            fields=(
                ("row_id", "int", "CS"),
                ("first_id", "int", "CS"),
                ("second_id", "int", "CS"),
                ("name", "string", "CS"),
                ("flag", "bool_alias", "CS"),
            ),
            data_rows=((1, 1, 2, "alpha", True),),
            code_rows=(("Target", "Target.json", "c"),),
            sheet="Target",
        )
        self.registry = TypeRegistry(str(self.tables))

    def tearDown(self):
        self._temp_dir.cleanup()

    def _converter(self, type_name):
        return self.registry.get_type(type_name)["convert_func"]

    def test_conversion_without_capture_keeps_legacy_output_shapes(self):
        self.assertEqual(7, self._converter("direct_ref")("7"))
        self.assertEqual([1, 2], self._converter("list_ref")("1#2"))
        self.assertEqual(
            {
                "flag": True,
                "_table": "Target",
                "_display_field": "DisplayOnly",
            },
            self._converter("typed_ref")(1),
        )

    def test_capture_records_direct_list_and_alias_references(self):
        with self.registry.capture_references() as references:
            self._converter("direct_ref")(1)
            self._converter("list_ref")("2#3")
            self._converter("alias_ref")(4)

        self.assertEqual([1, 2, 3, 4], [item["value"] for item in references])
        self.assertEqual(
            {("Target", "first_id")},
            {(item["table"], item["field"]) for item in references},
        )

    def test_both_split_dict_syntaxes_capture_both_references(self):
        for type_name in ("dict_positional", "dict_named"):
            with self.subTest(type_name=type_name):
                with self.registry.capture_references() as references:
                    result = self._converter(type_name)("1;2")

                self.assertEqual({"first": 1, "second": 2}, result)
                self.assertEqual(
                    [
                        {"table": "Target", "field": "first_id", "value": 1},
                        {"table": "Target", "field": "second_id", "value": 2},
                    ],
                    references,
                )

    def test_nested_split_list2_captures_each_leaf_reference(self):
        with self.registry.capture_references() as references:
            result = self._converter("nested_refs")("1#2|3#4")

        self.assertEqual([[1, 2], [3, 4]], result)
        self.assertEqual([1, 2, 3, 4], [item["value"] for item in references])

    def test_blank_values_are_skipped_but_zero_is_recorded(self):
        with self.registry.capture_references() as references:
            self._converter("direct_ref")(None)
            self._converter("name_ref")("   ")
            self._converter("direct_ref")(0)

        self.assertEqual([
            {"table": "Target", "field": "first_id", "value": 0}
        ], references)

    def test_typed_dictionary_reference_captures_its_converted_id(self):
        with self.registry.capture_references() as references:
            result = self._converter("typed_ref")("true")

        self.assertTrue(result["flag"])
        self.assertEqual([
            {"table": "Target", "field": "flag", "value": True}
        ], references)

    def test_nested_capture_restores_outer_capture(self):
        with self.registry.capture_references() as outer:
            self._converter("direct_ref")(1)
            with self.registry.capture_references() as inner:
                self._converter("direct_ref")(2)
            self._converter("direct_ref")(3)

        self.assertEqual([1, 3], [item["value"] for item in outer])
        self.assertEqual([2], [item["value"] for item in inner])

    def test_failed_conversion_is_not_captured_and_context_is_reset(self):
        with self.assertRaises(ValueError):
            with self.registry.capture_references() as failed:
                self._converter("direct_ref")("not-an-int")
        self.assertEqual([], failed)

        self._converter("direct_ref")(8)
        with self.registry.capture_references() as fresh:
            self._converter("direct_ref")(9)
        self.assertEqual([9], [item["value"] for item in fresh])

    def test_concurrent_captures_do_not_leak_between_contexts(self):
        def capture(value):
            with self.registry.capture_references() as references:
                self._converter("direct_ref")(value)
                return [item["value"] for item in references]

        with ThreadPoolExecutor(max_workers=2) as executor:
            results = list(executor.map(capture, (11, 22)))
        self.assertEqual([[11], [22]], results)


class ReferenceCaptureIntegrationTests(unittest.TestCase):
    def _export_split_dict(self, first_value, second_value, *, validation_only=True):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tables = root / "tables"
            client = root / "client"
            tables.mkdir()
            TypeDefinitionTemplate.ensure_exists(str(tables))
            definition = load_workbook(tables / "TypeDefinition.xlsx")
            definition["CODE"].append((
                "two_refs",
                "split_dict(find_id(Target,DisplayOnly,first_id) first;"
                "find_id(Target,DisplayOnly,second_id) second)",
                "",
            ))
            definition.save(tables / "TypeDefinition.xlsx")
            definition.close()

            write_workbook(
                tables / "Target.xlsx",
                fields=(
                    ("row_id", "int", "CS"),
                    ("first_id", "int", "CS"),
                    ("second_id", "int", "CS"),
                ),
                data_rows=((1, 1, 2),),
                code_rows=(("Target", "Target.json", "c"),),
                sheet="Target",
            )
            write_workbook(
                tables / "Source.xlsx",
                fields=(("row_id", "int", "CS"), ("refs", "two_refs", "CS")),
                data_rows=((10, f"{first_value};{second_value}"),),
                code_rows=(("Source", "Source.json", "c"),),
                sheet="Source",
            )

            result = ExcelConverter().export_all(
                str(tables), str(client), str(root / "server"),
                "c", validation_only=validation_only,
            )
            payload = None
            if not validation_only and (client / "Source.json").exists():
                payload = json.loads(
                    (client / "Source.json").read_text(encoding="utf-8")
                )
            return result, payload

    def test_integration_reports_second_missing_reference(self):
        result, _ = self._export_split_dict(1, 99)

        missing = [
            issue for issue in result["issues"]
            if issue["code"] == "REFERENCE_NOT_FOUND"
        ]
        self.assertFalse(result["success"], result["issues"])
        self.assertEqual(1, len(missing), result["issues"])
        self.assertIn("Target.second_id", missing[0]["message"])
        self.assertIn("='99'", missing[0]["message"])

    def test_integration_exports_same_json_shape_when_both_references_exist(self):
        result, payload = self._export_split_dict(1, 2, validation_only=False)

        self.assertTrue(result["success"], result["issues"])
        self.assertEqual({"first": 1, "second": 2}, payload["10"]["refs"])

    def test_conversion_failure_does_not_enqueue_partial_references(self):
        result, _ = self._export_split_dict(99, "not-an-int")

        codes = {issue["code"] for issue in result["issues"]}
        self.assertFalse(result["success"], result["issues"])
        self.assertIn("CONVERSION_ERROR", codes)
        self.assertNotIn("REFERENCE_NOT_FOUND", codes)


if __name__ == "__main__":
    unittest.main()
