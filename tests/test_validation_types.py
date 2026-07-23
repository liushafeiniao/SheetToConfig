import json
import math
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from sheet_to_config.utils.exporter.constraints import ConstraintError, ConstraintValidator
from sheet_to_config.utils.exporter.converter import ExcelConverter
from sheet_to_config.utils.exporter.core import FieldInfo, WorkSheet
from sheet_to_config.utils.exporter.expression import parse_call, parse_field_type
from sheet_to_config.utils.exporter.type_registry import (
    TypeDefinitionLoadError,
    TypeRegistry,
    UndefinedTypeError,
)
from sheet_to_config.utils.exporter.types import TypeConverter


def write_type_definition(path, definitions):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "CODE"
    worksheet.append(["类型名字", "调用函数", "说明"])
    for name, expression in definitions:
        worksheet.append([name, expression, ""])
    workbook.save(path / "TypeDefinition.xlsx")


def write_export_workbook(path, field_types, values=None):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Item"
    names = ["id"] + [f"field_{index}" for index in range(1, len(field_types))]
    worksheet.append(names)
    worksheet.append(list(field_types))
    worksheet.append(["c"] * len(field_types))
    worksheet.append([""] * len(field_types))
    worksheet.append(values or [1] + ["value"] * (len(field_types) - 1))

    code = workbook.create_sheet("CODE")
    code.append(["表名", "导出文件", "使用端"])
    code.append(["Item", "Item.json", "c"])
    workbook.save(path / "Item.xlsx")


class TypeValidationTests(unittest.TestCase):
    def test_field_type_parser_preserves_nested_constraint_arguments(self):
        parsed = parse_field_type(
            r"string+regex(^(foo|bar)\+[0-9]{1,3}$)+required()"
        )

        self.assertEqual(parsed.base_type, "string")
        self.assertEqual(parsed.constraints[0].name, "regex")
        self.assertEqual(parsed.constraints[0].args, (r"^(foo|bar)\+[0-9]{1,3}$",))
        self.assertEqual(parsed.constraints[1].name, "required")

    def test_constraints_use_raw_required_and_closed_range_and_full_regex(self):
        validator = ConstraintValidator()

        with self.assertRaisesRegex(ConstraintError, "不能为空"):
            validator.validate("required", [], "count", 0, {"count": 0}, {"count": ""})
        validator.validate("range", ["1", "3"], "count", 1, {}, {"count": 1})
        validator.validate("range", ["1", "3"], "count", 3, {}, {"count": 3})
        with self.assertRaisesRegex(ConstraintError, "范围"):
            validator.validate("range", ["1", "3"], "count", 4, {}, {"count": 4})
        validator.validate("regex", [r"A[0-9]+"], "code", "A12", {}, {"code": "A12"})
        with self.assertRaisesRegex(ConstraintError, "格式"):
            validator.validate("regex", [r"A[0-9]+"], "code", "xA12", {}, {"code": "xA12"})

    def test_equal_len_constraints_compare_converted_outer_shapes(self):
        validator = ConstraintValidator()
        row = {"left": [[1], [2]], "right": [[3], [4]]}
        validator.validate("equalLen", ["right"], "left", row["left"], row, row)
        validator.validate("equalLen2", ["right"], "left", row["left"], row, row)

        with self.assertRaisesRegex(ConstraintError, "外层"):
            validator.validate(
                "equalLen2", ["right"], "left", [[1], [2]],
                {"right": [[3]]}, {"right": [[3]]}
            )

    def test_scalar_conversions_reject_values_that_would_be_silently_changed(self):
        self.assertEqual(TypeConverter.to_int("1"), 1)
        self.assertEqual(TypeConverter.to_float("1.25"), 1.25)
        self.assertIs(TypeConverter.to_bool("off"), False)

        with self.assertRaisesRegex(ValueError, "整数"):
            TypeConverter.to_int("1.5")
        with self.assertRaisesRegex(ValueError, "布尔"):
            TypeConverter.to_bool("maybe")
        with self.assertRaisesRegex(ValueError, "有限"):
            TypeConverter.to_float(math.inf)

    def test_type_definition_supports_string_and_integer_enums(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            table_dir = Path(temp_dir)
            write_type_definition(table_dir, [
                ("quality", "enum(string,白,绿,蓝)"),
                ("stage", "enum(int,1,2,3)"),
            ])

            registry = TypeRegistry(str(table_dir))

            self.assertEqual(registry.get_type("quality")["convert_func"]("绿"), "绿")
            self.assertEqual(registry.get_type("stage")["convert_func"]("2"), 2)
            with self.assertRaisesRegex(ValueError, "枚举"):
                registry.get_type("stage")["convert_func"]("4")

    def test_type_definition_rejects_unknown_functions_when_loaded(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            table_dir = Path(temp_dir)
            write_type_definition(table_dir, [("bad", "silently_guess(int)")])

            registry = TypeRegistry(str(table_dir))
            with self.assertRaisesRegex(UndefinedTypeError, "未知"):
                registry.validate_types(["bad"])

    def test_type_definition_accepts_full_width_parentheses(self):
        expression = parse_call("split_list(find_id(mission,突破任务,id)）")

        self.assertEqual(expression.name, "split_list")
        self.assertEqual(expression.args, ("find_id(mission,突破任务,id)",))

    def test_type_definition_collects_all_invalid_rows(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            table_dir = Path(temp_dir)
            definitions = [(f"type_{row}", "string") for row in range(2, 41)]
            definitions[37 - 2] = ("bad_parenthesis", "split_list(int")
            definitions[39 - 2] = ("bad_function", "silently_guess(int)")
            write_type_definition(table_dir, definitions)

            registry = TypeRegistry(str(table_dir))
            with self.assertRaises(TypeDefinitionLoadError) as caught:
                registry.validate_types(["bad_parenthesis", "bad_function"])

            self.assertEqual(len(caught.exception.issues), 2)
            self.assertEqual(
                [(issue.row, issue.column, issue.code) for issue in caught.exception.issues],
                [
                    (37, 2, "TYPE_DEFINITION_PARENTHESIS"),
                    (39, 2, "TYPE_DEFINITION_ERROR"),
                ],
            )

    def test_string_primary_key_converts_excel_number_to_string(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            table_dir = Path(temp_dir)
            write_type_definition(table_dir, [("string", "string")])
            registry = TypeRegistry(str(table_dir))

            worksheet = WorkSheet("Items")
            worksheet.set_field_info(
                {"id": FieldInfo("id", "string", "CS")}, {"id": 0}
            )
            worksheet.add_row([1])

            converter = ExcelConverter()
            converter.type_registry = registry
            converter._current_file = "Items.xlsx"
            converter._current_sheet = "Items"
            rows = converter._worksheet_to_data(worksheet, "cs")

            self.assertEqual(rows[0]["id"], "1")
            self.assertIsInstance(rows[0]["id"], str)

            output_path = table_dir / "items.json"
            converter._write_json(rows, str(output_path))
            exported = json.loads(output_path.read_text(encoding="utf-8"))
            self.assertEqual(exported, {"1": {"id": "1"}})

    def test_converter_reports_type_definition_issues_with_source_cells(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            table_dir = Path(temp_dir)
            definitions = [(f"type_{row}", "string") for row in range(2, 41)]
            definitions[37 - 2] = ("bad_parenthesis", "split_list(int")
            definitions[39 - 2] = ("bad_function", "silently_guess(int)")
            write_type_definition(table_dir, definitions)
            write_export_workbook(
                table_dir,
                ["type_2", "bad_parenthesis", "bad_function"],
            )

            result = ExcelConverter().export_all(
                str(table_dir),
                str(table_dir / "client"),
                str(table_dir / "server"),
                validation_only=True,
            )

            self.assertFalse(result["success"])
            self.assertEqual(
                [
                    (issue["code"], issue["file"], issue["sheet"], issue["row"], issue["column"])
                    for issue in result["issues"]
                ],
                [
                    ("TYPE_DEFINITION_PARENTHESIS", "TypeDefinition.xlsx", "CODE", 37, 2),
                    ("TYPE_DEFINITION_ERROR", "TypeDefinition.xlsx", "CODE", 39, 2),
                ],
            )

    def test_unused_type_definition_errors_do_not_block_export(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            table_dir = Path(temp_dir)
            write_type_definition(table_dir, [
                ("int", "int"),
                ("string", "string"),
                ("unused_id", "find_id(Missing,Text,id)"),
                ("unused_list", "split_list(unused_id)"),
            ])
            write_export_workbook(table_dir, ["int", "string"])

            result = ExcelConverter().export_all(
                str(table_dir),
                str(table_dir / "client"),
                str(table_dir / "server"),
                validation_only=True,
            )

            self.assertTrue(result["success"])
            self.assertEqual(result["issues"], [])

    def test_complex_types_reject_empty_items_and_require_exact_dict_fields(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            table_dir = Path(temp_dir)
            write_type_definition(table_dir, [
                ("ids", "split_list(int)"),
                ("reward", "split_dict(int id;string name)"),
            ])
            registry = TypeRegistry(str(table_dir))
            ids = registry.get_type("ids")["convert_func"]
            reward = registry.get_type("reward")["convert_func"]

            self.assertEqual(ids("1#2"), [1, 2])
            with self.assertRaisesRegex(ValueError, "空"):
                ids("1##2")
            self.assertEqual(reward("1,Sword"), {"id": 1, "name": "Sword"})
            with self.assertRaisesRegex(ValueError, "数量"):
                reward("1")
            with self.assertRaisesRegex(ValueError, "未知"):
                reward("id:1;name:Sword;extra:x")


if __name__ == "__main__":
    unittest.main()
