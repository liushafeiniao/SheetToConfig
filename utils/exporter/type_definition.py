# -*- coding: utf-8 -*-
"""
类型定义管理器 - 严格从TypeDefinition.xlsx加载类型定义
所有字段类型必须在TypeDefinition.xlsx中定义，未定义则报错
"""
import os
from openpyxl import load_workbook
from typing import Dict, Optional

from .expression import parse_field_type


class TypeDefinition:
    """类型定义管理器"""
    
    def __init__(self, table_dir: str):
        """
        初始化类型定义
        
        Args:
            table_dir: 表格目录路径
        """
        self.table_dir = table_dir
        self.type_dict: Dict[str, str] = {}
        self.loaded = False
        self.load()
    
    def load(self):
        """从TypeDefinition.xlsx加载类型定义"""
        file_path = os.path.join(self.table_dir, "TypeDefinition.xlsx")
        
        if not os.path.exists(file_path):
            raise TypeDefinitionError(
                f"类型定义文件不存在: {file_path}\n"
                f"请在表格目录中创建TypeDefinition.xlsx文件，定义所有字段类型"
            )
        
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            raise TypeDefinitionError(f"无法读取TypeDefinition.xlsx: {e}")
        
        # 查找CODE表
        code_sheet = None
        for sheet_name in wb.sheetnames:
            if sheet_name.upper() == 'CODE':
                code_sheet = wb[sheet_name]
                break
        
        if not code_sheet:
            raise TypeDefinitionError(
                f"TypeDefinition.xlsx中未找到CODE工作表\n"
                f"请创建CODE工作表，格式：\n"
                f"第1列：类型名 | 第2列：转换函数"
            )
        
        # 读取类型定义
        rows = list(code_sheet.rows)
        if not rows:
            raise TypeDefinitionError("TypeDefinition.xlsx的CODE表为空")
        
        # 跳过表头（如果第一列是中文表头）
        start_row = 0
        first_cell = str(rows[0][0].value or "").strip()
        if first_cell in ('类型名', '类型', 'type', 'name'):
            start_row = 1
        
        for row in rows[start_row:]:
            if len(row) >= 2:
                type_name = str(row[0].value or "").strip()
                convert_func = str(row[1].value or "").strip()
                
                if type_name and convert_func:
                    # 类型名统一转小写存储
                    self.type_dict[type_name.lower()] = convert_func
        
        wb.close()
        self.loaded = True
        
        if not self.type_dict:
            raise TypeDefinitionError(
                "TypeDefinition.xlsx中未定义任何类型\n"
                "请至少定义以下基础类型：\n"
                "- int: int\n"
                "- float: float\n"
                "- string: string\n"
                "- intList: split_list(int)\n"
            )
    
    def get_convert_func(self, type_name: str) -> str:
        """
        获取类型的转换函数
        
        Args:
            type_name: 类型名（如'int', 'intList'）
        
        Returns:
            转换函数字符串
        
        Raises:
            UndefinedTypeError: 如果类型未定义
        """
        if not type_name:
            raise UndefinedTypeError("类型名为空")
        
        # 处理复合类型（如'int+len(1,10)'）
        base_type = parse_field_type(type_name).base_type
        
        # 处理cs/c平台标识（如'int/c'或'int/s'）
        if '/' in base_type:
            base_type = base_type.split('/')[0].strip()
        
        # 查找类型定义
        func = self.type_dict.get(base_type.lower())
        
        if func is None:
            raise UndefinedTypeError(
                f"未定义的类型: '{base_type}'\n"
                f"请在TypeDefinition.xlsx中定义该类型\n"
                f"已定义的类型: {', '.join(sorted(self.type_dict.keys()))}"
            )
        
        return func
    
    def get_all_types(self) -> Dict[str, str]:
        """获取所有类型定义"""
        return self.type_dict.copy()
    
    def is_defined(self, type_name: str) -> bool:
        """检查类型是否已定义"""
        if not type_name:
            return False
        base_type = parse_field_type(type_name).base_type.lower()
        return base_type in self.type_dict


class TypeDefinitionError(Exception):
    """类型定义错误"""
    pass


class UndefinedTypeError(Exception):
    """未定义类型错误"""
    pass
