# -*- coding: utf-8 -*-
"""
ID引用验证模块

验证填写的ID是否真实存在于被引用表中
"""
import os
from typing import Any, Dict, List, Set, Tuple, Optional
from openpyxl import load_workbook


class ReferenceError(Exception):
    """ID引用错误"""
    pass


class ReferenceValidator:
    """ID引用验证器"""
    
    def __init__(self, table_dir: str):
        self.table_dir = table_dir
        self._references: List[Dict[str, Any]] = []  # 待验证的引用
        self._id_cache: Dict[str, Set[Any]] = {}     # 表ID缓存
    
    def add_reference(self, value: Any, target_table: str, target_field: str,
                     source_file: str, source_sheet: str, row: int, col_name: str):
        """
        添加一个需要验证的ID引用
        
        Args:
            value: 引用的ID值（可能是单个值或列表）
            target_table: 目标表名（如 'item'）
            target_field: 目标字段名（如 'id'）
            source_file: 来源文件名
            source_sheet: 来源工作表名
            row: 行号
            col_name: 列名
        """
        self._references.append({
            'value': value,
            'target_table': target_table,
            'target_field': target_field,
            'source_file': source_file,
            'source_sheet': source_sheet,
            'row': row,
            'col_name': col_name
        })
    
    def validate_all(self) -> List[str]:
        """
        验证所有待验证的引用

        Returns:
            错误列表，如果没有错误返回空列表
        """
        # 按字段分组：同一来源表、工作表、字段和目标表只保留第一行。
        # 这一列的引用规则修正后，后续行会一起恢复，不需要逐条刷屏。
        error_types: Dict[Tuple[str, str, str, str], Tuple[int, Any]] = {}

        for ref in self._references:
            target_table = ref['target_table']
            target_field = ref['target_field']
            cache_key = f"{target_table}.{target_field}"

            # 获取或加载目标表的所有ID
            if cache_key not in self._id_cache:
                ids = self._load_table_ids(target_table, target_field)
                self._id_cache[cache_key] = ids

            valid_ids = self._id_cache[cache_key]

            # 验证值
            values = ref['value']
            if not isinstance(values, list):
                values = [values]

            # 展平嵌套列表
            flat_values = self._flatten_values(values)

            for val in flat_values:
                if val is None or val == "":
                    continue

                # 统一转换为字符串进行比较（避免 int/str 不匹配）
                val_str = str(val).strip()

                # 同时检查原始值和数值转换后的值
                if val_str not in valid_ids:
                    try:
                        val_num = int(float(val_str))
                        if val_num not in valid_ids:
                            self._collect_error(error_types, ref, val, target_table, target_field)
                    except (ValueError, TypeError):
                        self._collect_error(error_types, ref, val, target_table, target_field)

        # 将分组后的错误转换为日志消息（相同错误只报一次）
        return self._format_errors(error_types)

    def _collect_error(self, error_types: Dict, ref: Dict, val: Any,
                      target_table: str, target_field: str):
        """收集错误，同一来源字段只记录第一次出现的行号"""
        target = f"{target_table}.{target_field}"
        key = (ref['source_file'], ref['source_sheet'], ref['col_name'], target)

        if key not in error_types:
            error_types[key] = (ref['row'], val)

    def _format_errors(self, error_types: Dict) -> List[str]:
        """将分组后的错误格式化为日志消息"""
        errors = []
        for (file, sheet, col_name, target), (first_row, val) in error_types.items():
            errors.append(f"[{file}:{sheet}] 第{first_row}行 '{col_name}'='{val}' 在 '{target}' 中不存在")
        return errors
    
    def _flatten_values(self, values: List[Any]) -> List[Any]:
        """展平嵌套列表"""
        result = []
        for v in values:
            if isinstance(v, list):
                result.extend(self._flatten_values(v))
            else:
                result.append(v)
        return result
    
    def _load_table_ids(self, table_name: str, field_name: str) -> Set[Any]:
        """
        从目标表加载所有ID值
        
        Args:
            table_name: 表名（如 'item' 对应 item.xlsx）
            field_name: 字段名
        
        Returns:
            ID值集合
        """
        ids = set()
        
        # 查找目标Excel文件
        target_file = os.path.join(self.table_dir, f"{table_name}.xlsx")
        if not os.path.exists(target_file):
            # 尝试查找任何包含该名称的文件
            for fname in os.listdir(self.table_dir):
                if fname.lower().startswith(table_name.lower()) and fname.endswith('.xlsx'):
                    target_file = os.path.join(self.table_dir, fname)
                    break
        
        if not os.path.exists(target_file):
            raise ReferenceError(f"引用表不存在: {table_name}.xlsx")
        
        wb = None
        try:
            wb = load_workbook(target_file, read_only=True, data_only=True)
            
            # 查找包含目标字段的工作表
            found_field = False
            for sheet_name in wb.sheetnames:
                if sheet_name.upper() == 'CODE':
                    continue
                
                ws = wb[sheet_name]
                rows = list(ws.rows)
                
                if len(rows) < 4:
                    continue
                
                # 查找字段列索引
                header_row = rows[0]
                field_idx = -1
                for idx, cell in enumerate(header_row):
                    if cell.value == field_name:
                        field_idx = idx
                        break
                
                if field_idx == -1:
                    continue
                found_field = True
                
                # 读取该列的所有值（从第5行开始）
                for row in rows[4:]:
                    if field_idx < len(row):
                        val = row[field_idx].value
                        if val is not None and val != "":
                            # 同时存储字符串和数值版本
                            val_str = str(val).strip()
                            ids.add(val_str)
                            try:
                                val_num = int(float(val))
                                ids.add(val_num)
                            except (ValueError, TypeError):
                                pass
            
            if not found_field:
                raise ReferenceError(
                    f"引用字段不存在: {table_name}.{field_name}"
                )
        except ReferenceError:
            raise
        except Exception as e:
            raise ReferenceError(
                f"引用表读取失败: {target_file}: {e}"
            ) from e
        finally:
            if wb is not None:
                wb.close()
        
        return ids
    
    def clear(self):
        """清空验证队列"""
        self._references.clear()
        self._id_cache.clear()
