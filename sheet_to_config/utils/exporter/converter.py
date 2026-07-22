"""
Excel转换器模块

处理Excel到JSON/LUA的转换，包括：
- 严格类型验证（所有类型必须在TypeDefinition.xlsx中定义）
- 完整的分隔符支持（#、|、_等）
- 约束系统（len、len2、equalLen、coexist、leastOne等）
- 空值默认处理
"""

import os
import json
import re
import shutil
import tempfile
from collections import OrderedDict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Callable
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .types import TypeConverter
from .constraints import ConstraintValidator, ConstraintError
from .expression import ExpressionSyntaxError, parse_field_type
from .template import TypeDefinitionTemplate
from .type_registry import TypeRegistry, UndefinedTypeError
from .reference_validator import ReferenceValidator, ReferenceError
from .protobuf_schema import ProtoSchemaParser, ProtoSchemaError, extract_manifest
from .exporters.protobuf_exporter import ProtobufExporter
from .atomic_writer import AtomicCommitError, commit_files
from .batch_transaction import (
    IncrementalManifestRequiredError,
    prepare_batch_commit,
)
from .validation import ValidationIssue


class ConverterError(Exception):
    """转换错误"""
    pass


def _reference_id_value(value: Any, field_name: Optional[str] = None) -> Any:
    """Extract wire/reference IDs from find_id result dictionaries."""
    if isinstance(value, list):
        return [_reference_id_value(item, field_name) for item in value]
    if isinstance(value, dict):
        if field_name and field_name in value:
            return value[field_name]
        for key, item in value.items():
            if not str(key).startswith("_"):
                return item
        return None
    return value


class ExcelConverter:
    """Excel配置转换器"""

    # Row-level errors are actionable at the field level: fixing the type,
    # constraint, or column fixes every affected row. Keep the first location
    # so a large table does not drown the user in identical messages.
    _DEDUPLICATED_ISSUE_CODES = {
        'CONVERSION_ERROR',
        'CONSTRAINT_ERROR',
        'DUPLICATE_VALUE',
        'FORMULA_NO_CACHED_VALUE',
        'INVALID_PRIMARY_KEY',
        'MISSING_PRIMARY_KEY',
        'REFERENCE_NOT_FOUND',
        'REFERENCE_TABLE_ERROR',
        'UNDEFINED_TYPE',
        'WORKBOOK_READ_ERROR',
    }

    def __init__(self, log_func=None):
        self.log_func = log_func
        self.type_registry: Optional[TypeRegistry] = None
        self._constraint_validator = ConstraintValidator()
        self._reference_validator: Optional[ReferenceValidator] = None
        self._current_file = ""  # 当前处理的文件，用于错误报告
        self._current_sheet = ""  # 当前处理的工作表
        self._current_row = 0  # 当前行号
        self._client_path = ""  # 客户端路径，用于路径验证
        self._issues: List[ValidationIssue] = []
        self._issue_keys = set()
        self._asset_root = ""
        self._previous_client_path = ""
        self._previous_server_path = ""
        self._unreadable_workbooks = set()

    def _record_issue(self, code: str, message: str, *, field: str = "",
                      raw_value: Any = None, row: Optional[int] = None,
                      column: int = 0, file: str = "", sheet: str = "",
                      issue_group: str = "") -> None:
        issue_row = self._current_row if row is None else row
        issue_file = file or self._current_file
        issue_sheet = sheet or self._current_sheet
        if code in self._DEDUPLICATED_ISSUE_CODES:
            # Include the physical column when a header has no field name so
            # two unnamed cells cannot hide each other's diagnostics.
            identity = field or (f"column:{column}" if column else "")
            issue_key = (
                code, issue_file.casefold(), issue_sheet.casefold(),
                identity.casefold(), issue_group,
            )
            if issue_key in self._issue_keys:
                return
            self._issue_keys.add(issue_key)
        cell = ""
        if column > 0 and issue_row > 0:
            cell = f"{get_column_letter(column)}{issue_row}"
        location = "/".join(part for part in (issue_file, issue_sheet) if part)
        if cell:
            location = f"{location}!{cell}"
        self._issues.append(ValidationIssue(
            code=code, message=message, file=issue_file, sheet=issue_sheet,
            row=issue_row or 0, column=column, field=field,
            path=location, raw_value=raw_value,
        ))

    def _issues_as_dicts(self) -> List[dict]:
        return [issue.to_dict() for issue in self._issues]

    def _has_issue(self, code: str, file: str = "", sheet: str = "") -> bool:
        return any(
            issue.code == code
            and (not file or issue.file == file)
            and (not sheet or issue.sheet == sheet)
            for issue in self._issues
        )

    def _log(self, message: str):
        """输出日志"""
        try:
            if self.log_func:
                self.log_func(message)
            else:
                print(message)
        except UnicodeEncodeError:
            # Logging must never change commit semantics on legacy Windows
            # consoles such as GBK/cp936.
            print(message.encode('ascii', 'backslashreplace').decode('ascii'))
        except Exception:
            # A UI/log sink is observational and cannot turn a committed batch
            # into a reported export failure.
            return

    @staticmethod
    def _artifact_info(path: str, client_path: str, server_path: str,
                       csharp_path: str, workbook: str, sheet: str) -> dict:
        target = Path(path).resolve()
        for platform, root in (
            ('client', client_path), ('server', server_path), ('csharp', csharp_path)
        ):
            if not root:
                continue
            root_path = Path(root).resolve()
            try:
                target.relative_to(root_path)
                return {
                    'path': str(target), 'root': str(root_path),
                    'platform': platform, 'format': target.suffix.lstrip('.').lower(),
                    'workbook': os.path.basename(workbook), 'sheet': sheet,
                }
            except ValueError:
                continue
        raise ConverterError(f"生成产物不在已配置输出目录内: {target}")

    def _parse_field_type(self, type_str: str) -> Tuple[str, List[Tuple[str, List[str]]]]:
        """
        解析字段类型字符串
        
        Args:
            type_str: 如 "intList+len(1,5)+equalLen(other)"
        
        Returns:
            (基础类型, [(约束名, [参数]), ...])
        """
        try:
            expression = parse_field_type(type_str or "str")
            constraints = []
            for constraint in expression.constraints:
                params = list(constraint.args)
                self._constraint_validator.validate_definition(constraint.name, params)
                constraints.append((constraint.name, params))
            return expression.base_type, constraints
        except (ExpressionSyntaxError, ConstraintError) as exc:
            raise ConverterError(f"字段类型表达式无效 '{type_str}': {exc}") from exc

    def _convert_value(self, value: Any, type_name: str, field_name: str) -> Any:
        """
        转换单个值
        
        Args:
            value: 原始值
            type_name: 类型名称
            field_name: 字段名（用于错误报告）
        
        Returns:
            转换后的值
        
        Raises:
            UndefinedTypeError: 类型未定义
            ConverterError: 转换失败
        """
        if self.type_registry is None:
            raise ConverterError("类型注册表未初始化")
        
        # 获取类型定义
        type_def = self.type_registry.get_type(type_name)
        
        # 获取转换函数
        convert_func = type_def.get('convert_func')
        if not convert_func:
            raise ConverterError(f"类型 '{type_name}' 没有转换函数")
        
        # 检查是否是引用类型（用于后续验证）
        ref_info = type_def.get('reference_info')
        
        # 执行转换
        try:
            if value is None or value == "":
                # 返回默认值
                result = None
                if type_name == 'int':
                    result = 0
                elif type_name == 'float':
                    result = 0.0
                elif type_name in ('str', 'string'):
                    result = ""
                elif type_name == 'bool':
                    result = False
                elif type_name == 'bytes':
                    result = b""
                elif 'list' in type_name.lower() or type_name in ('award',):
                    result = []
                else:
                    # 尝试从convert_func获取空值处理
                    result = convert_func(value)
            else:
                result = convert_func(value)
            
            # 记录ID引用（客户端和服务端都记录，确保完整性）
            if self._reference_validator and result and hasattr(self, '_current_platform'):
                ref_info = self.type_registry.get_reference_info(type_name)
                if ref_info:
                    reference_value = _reference_id_value(result, ref_info.get('field'))
                    self._reference_validator.add_reference(
                        value=reference_value,  # 记录转换后的值（已分割）
                        target_table=ref_info['table'],
                        target_field=ref_info['field'],
                        source_file=self._current_file,
                        source_sheet=self._current_sheet,
                        row=self._current_row,
                        col_name=field_name
                    )

            # 路径类型验证（检查文件是否存在）
            if result and self._client_path:
                type_def = self.type_registry.get_type(type_name)
                func_str = type_def.get('convert_func_str', '')
                if func_str.startswith('path('):
                    self._validate_path_exists(result, field_name)

            return result
        except Exception as e:
            raise ConverterError(f"字段 '{field_name}' 转换失败: {e}")

    def _validate_path_exists(self, path_value: Any, field_name: str):
        """
        验证路径对应的文件是否存在于客户端目录

        Args:
            path_value: 路径值（可能是字符串或列表）
            field_name: 字段名（用于错误报告）

        Raises:
            ConverterError: 文件不存在时抛出
        """
        if not self._asset_root:
            return
        asset_root = Path(self._asset_root).resolve()

        # 处理列表类型的路径（如 split_list(path(...))）
        paths_to_check = []
        if isinstance(path_value, list):
            paths_to_check = [str(p) for p in path_value if p]
        else:
            paths_to_check = [str(path_value)]

        for path_str in paths_to_check:
            if not path_str:
                continue

            relative_path = Path(path_str.replace('/', os.sep).replace('\\', os.sep))
            if relative_path.is_absolute():
                raise ConverterError(
                    f"字段 '{field_name}' 路径验证失败: 资源路径必须是相对路径 '{path_str}'"
                )
            full_path = (asset_root / relative_path).resolve()
            try:
                full_path.relative_to(asset_root)
            except ValueError as exc:
                raise ConverterError(
                    f"字段 '{field_name}' 路径验证失败: 路径越过资源根目录 '{path_str}'"
                ) from exc

            # 检查文件是否存在
            if not full_path.is_file():
                raise ConverterError(
                    f"字段 '{field_name}' 路径验证失败: 文件不存在 '{path_str}' "
                    f"(查找路径: {full_path})"
                )

    def _validate_constraints(self, constraints: List[Tuple[str, List[str]]],
                             field_name: str, field_value: Any,
                             row_data: Dict, row_raw: Dict) -> None:
        """
        验证字段约束
        
        Args:
            constraints: 约束列表
            field_name: 字段名
            field_value: 字段值
            row_data: 已转换的行数据
            row_raw: 原始行数据
        
        Raises:
            ConstraintError: 验证失败
        """
        for constraint_name, params in constraints:
            try:
                self._constraint_validator.validate(
                    constraint_name, params,
                    field_name, field_value,
                    row_data, row_raw
                )
            except ConstraintError as exc:
                # Preserve the constraint identity for diagnostics without
                # changing the established user-facing error text.
                exc.constraint_name = constraint_name
                raise

    def convert_row(self, row_data: List[Any], field_info: Dict[str, Dict],
                    line: int, omit_empty_fields: Optional[set] = None) -> OrderedDict:
        """
        转换单行数据（收集所有错误）

        Args:
            row_data: 行原始数据
            field_info: 字段信息 {字段名: {type, platform, col_idx, constraints}}
            line: 行号

        Returns:
            转换后的行数据

        Raises:
            ConverterError: 收集所有错误后抛出
        """
        result = OrderedDict()
        omit_empty_fields = omit_empty_fields or set()
        raw_data = OrderedDict()  # 原始数据用于约束检查
        errors = []  # 收集所有错误

        # First pass: collect raw cells in physical column order.
        for field_name, info in field_info.items():
            col_idx = info.get('col_idx', -1)
            if 0 <= col_idx < len(row_data):
                raw_data[field_name] = row_data[col_idx]
            else:
                raw_data[field_name] = ""

        converted_fields = set()
        # Second pass: convert every field before cross-field constraints run.
        for field_name, info in field_info.items():
            field_type_full = info.get('type', 'str')
            constraints = info.get('constraints', [])

            # 获取原始值
            raw_value = raw_data.get(field_name, "")
            omit_empty = field_name in omit_empty_fields and (
                raw_value is None
                or (isinstance(raw_value, str) and raw_value.strip() == '')
            )

            try:
                # 转换值
                converted_value = self._convert_value(raw_value, field_type_full, field_name)

                result[field_name] = converted_value
                converted_fields.add(field_name)

            except UndefinedTypeError as e:
                errors.append(f"字段'{field_name}'：类型未定义，请在TypeDefinition.xlsx中添加")
                self._record_issue(
                    "UNDEFINED_TYPE", str(e), field=field_name,
                    raw_value=raw_value, column=info.get('col_idx', -1) + 1,
                )
            except ConstraintError as e:
                # 添加原始值信息帮助诊断
                raw_value_str = str(raw_value) if raw_value is not None else "(空)"
                if len(raw_value_str) > 50:
                    raw_value_str = raw_value_str[:47] + "..."
                errors.append(f"{e} (原始值: '{raw_value_str}')")
            except Exception as e:
                errors.append(f"字段'{field_name}'：{e}")
                self._record_issue(
                    "CONVERSION_ERROR", str(e), field=field_name,
                    raw_value=raw_value, column=info.get('col_idx', -1) + 1,
                )

        # Third pass: validate constraints with the fully converted row.
        for field_name, info in field_info.items():
            if field_name not in converted_fields:
                continue
            try:
                self._validate_constraints(
                    info.get('constraints', []), field_name, result[field_name],
                    result, raw_data
                )
            except ConstraintError as exc:
                errors.append(str(exc))
                self._record_issue(
                    "CONSTRAINT_ERROR", str(exc), field=field_name,
                    raw_value=raw_data.get(field_name),
                    column=info.get('col_idx', -1) + 1,
                    issue_group=getattr(exc, 'constraint_name', ''),
                )

        for field_name in omit_empty_fields:
            raw_value = raw_data.get(field_name)
            if raw_value is None or (isinstance(raw_value, str) and not raw_value.strip()):
                result.pop(field_name, None)

        # 如果有错误，汇总抛出（包含ID信息）
        if errors:
            error_msg = f"[{self._current_sheet}] 第{line}行:\n"
            for err in errors:
                error_msg += f"  - {err}\n"
            raise ConverterError(error_msg.rstrip())

        result.raw_data = raw_data
        return result

    def _build_field_info(self, worksheet) -> Dict[str, Dict]:
        """
        从工作表构建字段信息

        Returns:
            {字段名: {type, platform, desc, col_idx, constraints}}
        """
        field_info = {}
        undefined_types = []
        ref_type_errors = []

        # 需要传入 table_dir 给 type_registry 使用
        # 但这个方法没有 table_dir 参数，需要从外部传入或者存储在实例上
        # 暂时先不检查引用类型，因为缺少 table_dir 信息

        for field_name, info in worksheet.field_info.items():
            col_idx = worksheet.get_column_index(field_name)

            # 解析类型和约束
            type_full = info.field_type or 'str'
            base_type, constraints = self._parse_field_type(type_full)

            # 验证类型是否存在
            if self.type_registry and not self.type_registry.has_type(base_type):
                undefined_types.append(f"'{field_name}'(类型'{base_type}')")
                continue

            if self.type_registry and getattr(self, '_current_format', '') in ('json', 'lua'):
                convert_expression = self.type_registry.get_type(base_type).get(
                    'convert_func_str', ''
                )
                if re.search(r'(^|\W)bytes($|\W)', convert_expression):
                    message = (
                        f"字段 '{field_name}' 使用bytes类型，但bytes只允许导出为Protobuf"
                    )
                    self._record_issue(
                        'BYTES_FORMAT_ERROR', message, field=field_name,
                        row=2, column=col_idx + 1,
                    )
                    raise ConverterError(message)

            # 检查引用类型（find_id 类型）
            if self.type_registry:
                try:
                    # 获取被引用字段的类型
                    ref_type = self.type_registry.get_referenced_type(base_type)
                    if ref_type is not None:
                        # 记录引用类型，用于数据转换时的类型验证
                        pass
                        # 这里可以存储引用类型信息，供后续使用
                except Exception as e:
                    # 类型检查失败，静默处理，不阻塞导出
                    pass

            field_info[field_name] = {
                'type': base_type,
                'platform': str(info.platform or 'cs').strip().lower(),
                'desc': info.desc or '',
                'col_idx': col_idx,
                'constraints': constraints
            }

        # 如果有未定义类型，汇总抛出
        if undefined_types:
            fields = ", ".join(undefined_types)
            raise UndefinedTypeError(
                f"以下字段类型未定义，请在TypeDefinition.xlsx中添加: {fields}"
            )

        return field_info

    def _primary_key_type_message(self, field_name: str, info: Dict[str, Any]) -> str:
        """Explain why a primary-key type cannot be used in the first column."""
        type_name = str(info.get('type') or 'str').strip()
        expression = ''
        if self.type_registry is not None:
            try:
                expression = str(
                    self.type_registry.get_type(type_name).get('convert_func_str', '')
                ).strip()
            except UndefinedTypeError:
                expression = ''

        type_cell = f"{get_column_letter(info.get('col_idx', 0) + 1)}2"
        prefix = f"第一列主键 '{field_name}' 的类型单元格 {type_cell} 当前是 '{type_name}'"
        lowered = expression.lower()
        if re.search(r'\b(find_id|find)\s*\(', lowered):
            return (
                f"{prefix}，这是 find_id 跨表引用类型，不能作为本表主键。"
                f"请将 {type_cell} 改为 int（或其他标量类型）；'{type_name}' 只能用于"
                "其他字段引用本表或其他表的 ID。"
            )
        if re.search(r'\b(split_list|split_dict)\b', lowered) or 'list' in lowered:
            return (
                f"{prefix}，该类型会转换为列表或字典，不能作为主键。"
                f"请将 {type_cell} 改为 int/string 等标量类型。"
            )
        if re.search(r'\bbytes\b', lowered) or type_name.lower() == 'bytes':
            return (
                f"{prefix}，bytes 是字节串类型，不能作为主键。"
                f"请将 {type_cell} 改为 int/string 等标量类型。"
            )
        return ''

    def _worksheet_to_data(self, worksheet, cors: str,
                           omit_empty_fields: Optional[set] = None) -> List[OrderedDict]:
        """
        将工作表转换为数据列表
        
        Args:
            worksheet: WorkSheet对象
            cors: 'c'客户端 / 's'服务端 / 'cs'两者
        
        Returns:
            转换后的数据列表
        """
        # 构建字段信息
        field_info = self._build_field_info(worksheet)
        
        issue_start = len(self._issues)
        primary_field = next(
            (name for name, info in field_info.items() if info.get('col_idx') == 0),
            None,
        )
        if primary_field is None:
            self._record_issue(
                "MISSING_PRIMARY_KEY", "第一列 A1 必须填写主键字段名", row=1, column=1
            )
            raise ConverterError("第一列 A1 必须填写主键字段名")

        primary_type_message = self._primary_key_type_message(
            primary_field, field_info[primary_field]
        )
        if primary_type_message:
            primary_column = field_info[primary_field].get('col_idx', 0) + 1
            self._record_issue(
                "INVALID_PRIMARY_KEY", primary_type_message,
                field=primary_field, raw_value=field_info[primary_field].get('type'),
                row=2, column=primary_column,
            )
            raise ConverterError(primary_type_message)

        # 过滤字段（根据平台）
        filtered_field_info = {}
        for name, info in field_info.items():
            platform = info.get('platform', 'cs')
            if cors == 'c' and 'c' not in platform:
                continue
            if cors == 's' and 's' not in platform:
                continue
            filtered_field_info[name] = info

        if primary_field not in filtered_field_info:
            primary_column = field_info[primary_field].get('col_idx', 0) + 1
            self._record_issue(
                "PRIMARY_KEY_NOT_EXPORTED",
                f"第一列主键 '{primary_field}' 的平台标记必须包含 {cors.upper()}。"
                f"请将第三行 {get_column_letter(primary_column)}3 改为 CS，"
                f"或至少包含 {cors.upper()}。",
                field=primary_field, row=3,
                column=primary_column,
            )
            raise ConverterError(f"第一列主键 '{primary_field}' 未导出到 {cors.upper()} 端")
        
        # 转换所有行（Excel实际行号 = i + 5，因为前4行是表头）
        # 跳过第一列为空的行
        all_data = []
        converted_rows = []
        for i, row in enumerate(worksheet.rows):
            excel_row_num = i + 5  # Excel实际行号
            
            # 检查第一列是否为空（第一列通常是ID列）
            first_col_value = row.data[0] if row.data else None
            
            if not any(
                value is not None and (not isinstance(value, str) or value.strip())
                for value in row.data
            ):
                continue
            if first_col_value is None or (
                isinstance(first_col_value, str) and first_col_value.strip() == ''
            ):
                self._record_issue(
                    "MISSING_PRIMARY_KEY",
                    f"第一列主键不能为空；本行其他列有内容。请填写 A{excel_row_num}，"
                    "或清空整行。",
                    field=primary_field,
                    raw_value=first_col_value, row=excel_row_num, column=1,
                )
                continue
            # 其他类型（数值包括0、布尔、列表等）都是有效值
            
            self._current_row = excel_row_num  # 设置当前行号（用于ID引用验证）
            try:
                converted = self.convert_row(
                    row.data, filtered_field_info, excel_row_num, omit_empty_fields
                )
            except ConverterError:
                continue
            converted.excel_row = excel_row_num
            primary_value = converted.get(primary_field)
            if primary_value is None or isinstance(primary_value, (list, dict, bytes, bytearray)):
                self._record_issue(
                    "INVALID_PRIMARY_KEY",
                    f"第一列主键 '{primary_field}' 转换后不是标量。请检查 A2 类型，"
                    "第一列应使用 int/string 等标量类型，不能使用列表、字典或 find_id 引用类型。",
                    field=primary_field, raw_value=first_col_value,
                    row=excel_row_num, column=1,
                )
                continue
            all_data.append(converted)
            converted_rows.append((excel_row_num, converted))

        unique_fields = {primary_field}
        unique_fields.update(
            name for name, info in filtered_field_info.items()
            if any(item[0] == 'unique' for item in info.get('constraints', []))
        )
        for field_name in unique_fields:
            seen = {}
            column = filtered_field_info[field_name].get('col_idx', -1) + 1
            for excel_row_num, converted in converted_rows:
                value = converted.get(field_name)
                canonical = json.dumps(
                    value, ensure_ascii=False, sort_keys=True, default=str,
                    separators=(',', ':'),
                )
                if canonical in seen:
                    current_cell = f"{get_column_letter(column)}{excel_row_num}"
                    first_cell = f"{get_column_letter(column)}{seen[canonical]}"
                    if field_name == primary_field:
                        message = (
                            f"第一列主键 '{field_name}' 重复：{current_cell} 与 {first_cell} "
                            f"的值都是 '{value}'。请修改或删除其中一行，主键必须唯一。"
                        )
                    else:
                        message = (
                            f"唯一字段 '{field_name}' 重复：{current_cell} 与 {first_cell} "
                            f"的值都是 '{value}'。请修改其中一个值。"
                        )
                    self._record_issue(
                        "DUPLICATE_VALUE",
                        message,
                        field=field_name, raw_value=value, row=excel_row_num,
                        column=column,
                    )
                else:
                    seen[canonical] = excel_row_num

        if len(self._issues) > issue_start:
            raise ConverterError(f"工作表 {worksheet.name} 校验失败")
        
        return all_data

    def _write_json(self, data: List[OrderedDict], output_path: str):
        """Write JSON using the legacy id/ID/key priority with a safe fallback."""
        result = OrderedDict()
        key_field = next(
            (candidate for candidate in ('id', 'ID', 'key')
             if any(candidate in row for row in data)),
            next(iter(data[0]), '') if data else '',
        )
        for row in data:
            if not row:
                continue
            key_value = row.get(key_field)
            raw_key_value = getattr(row, 'raw_data', {}).get(key_field)
            excel_row = getattr(row, 'excel_row', 0)
            if (
                raw_key_value is None
                or (isinstance(raw_key_value, str) and not raw_key_value.strip())
                or key_value is None
                or isinstance(key_value, (list, dict, bytes, bytearray))
            ):
                message = (
                    f"JSON 根键字段 '{key_field}' 在第 {excel_row or '?'} 行为空或不是标量，"
                    "请填写唯一的 int/string 值。"
                )
                self._record_issue(
                    'INVALID_JSON_ROOT_KEY', message, field=key_field,
                    raw_value=key_value, row=excel_row,
                )
                raise ConverterError(message)
            rendered_key = str(key_value)
            if rendered_key in result:
                message = (
                    f"JSON 根键字段 '{key_field}' 的值 '{key_value}' 重复，"
                    f"请检查第 {excel_row or '?'} 行。"
                )
                self._record_issue(
                    'DUPLICATE_VALUE', message, field=key_field,
                    raw_value=key_value, row=excel_row,
                )
                raise ConverterError(message)
            result[rendered_key] = row

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)

    def _write_lua(self, data: List[OrderedDict], output_path: str, sheet_name: str):
        """写入LUA文件"""
        lines = [f"{sheet_name} = {{}}"]
        
        for i, row in enumerate(data):
            lines.append("")
            lines.append(f"{sheet_name}[{i+1}] = {{")
            
            for key, value in row.items():
                lua_value = self._to_lua_value(value)
                lines.append(f"    {key} = {lua_value},")
            
            lines.append("}")
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))

    def _to_lua_value(self, value: Any) -> str:
        """转换为LUA格式的值"""
        if value is None:
            return "nil"
        elif isinstance(value, bool):
            return "true" if value else "false"
        elif isinstance(value, (int, float)):
            return str(value)
        elif isinstance(value, str):
            return f'"{value}"'
        elif isinstance(value, list):
            items = [self._to_lua_value(v) for v in value]
            return "{" + ", ".join(items) + "}"
        elif isinstance(value, dict):
            items = [f'{k} = {self._to_lua_value(v)}' for k, v in value.items()]
            return "{" + ", ".join(items) + "}"
        else:
            return str(value)

    @staticmethod
    def _strip_output_extension(file_name: str, output_format: str) -> str:
        """Remove the configured output extension without changing name casing."""
        suffix = f".{output_format}"
        if file_name.lower().endswith(suffix):
            return file_name[:-len(suffix)]
        return file_name

    def export_file(self, excel_path: str, client_path: str, server_path: str,
                    mode: str = 'cs', allow_breaking_proto_change: bool = False,
                    csharp_path: str = '', export_pb: bool = True
                    ) -> Tuple[bool, List[str], List[str], List[dict]]:
        """
        导出单个Excel文件
        
        Args:
            excel_path: Excel文件路径
            client_path: 客户端输出目录
            server_path: 服务端输出目录
            mode: 导出模式 'c'/'s'/'cs'
        
        Returns:
            (是否成功, 成功列表, 失败列表)
        """
        self._current_file = os.path.basename(excel_path)
        
        # 读取CODE表配置
        codes = self._read_code_sheets(excel_path)
        if not codes:
            if (
                not self._has_issue('WORKBOOK_READ_ERROR', os.path.basename(excel_path))
                and not self._has_issue('MISSING_CODE', os.path.basename(excel_path))
            ):
                self._record_issue(
                    'MISSING_CODE', '缺少CODE工作表或CODE中没有有效导出配置',
                    file=os.path.basename(excel_path), row=0,
                )
            return False, [], ['缺少CODE工作表或有效配置'], []
        
        success_list = []
        fail_list = []
        artifacts = []
        table_dir = os.path.dirname(excel_path)
        
        for code in codes:
            self._current_sheet = code.sheet_name
            self._current_format = code.format
            if getattr(code, 'implicit_format', False):
                self._log(
                    f"[WARNING] {self._current_file}:{code.sheet_name} 的输出名 "
                    f"'{code.file_name}' 未写扩展名，本次按 JSON 兼容导出；"
                    "请在 CODE 表中补上 .json。"
                )
            
            if code.format == 'proto':
                fail_list.append(
                    f"{code.file_name} - .proto不能单独导出，请将CODE输出名改为.pb"
                )
                continue

            if code.format not in ('json', 'lua', 'pb'):
                message = f"未知或缺少输出扩展名: {code.file_name or '(空)'}"
                fail_list.append(message)
                if not self._has_issue(
                    'UNKNOWN_OUTPUT_FORMAT', self._current_file, self._current_sheet
                ):
                    self._record_issue('UNKNOWN_OUTPUT_FORMAT', message, row=0)
                continue
            
            if code.platform and code.platform not in ('c', 's', 'cs'):
                fail_list.append(f"{code.file_name} - 无效的平台: {code.platform}")
                continue
            
            # 检查平台
            export_platform = code.platform if code.platform else mode
            if mode == 'c' and export_platform not in ('c', 'cs'):
                continue
            if mode == 's' and export_platform not in ('s', 'cs'):
                continue
            if mode in ('c', 's'):
                export_platform = mode
            
            # 读取工作表
            worksheet = self._read_worksheet(excel_path, code.sheet_name)
            if not worksheet:
                fail_list.append(f"{code.file_name} - 工作表不存在: {code.sheet_name}")
                continue
            
            # 转换数据并验证引用（验证失败则跳过写入）
            generated_files = []  # 记录生成的文件，用于错误时删除
            has_error = False

            try:
                file_name = self._strip_output_extension(code.file_name, 'pb')
                proto_schema = None
                if code.format == 'pb':
                    has_proto = False
                    probe = load_workbook(excel_path, read_only=True, data_only=True)
                    has_proto = any(name.upper() == "PROTO" for name in probe.sheetnames)
                    probe.close()
                    previous_manifest = self._load_previous_manifest(
                        excel_path, code, client_path, server_path, table_dir
                    )
                    if has_proto:
                        proto_schema = ProtoSchemaParser.parse_workbook(
                            excel_path, code.sheet_name, previous_manifest,
                            message_base=file_name,
                        )
                    else:
                        proto_schema = ProtoSchemaParser.auto_from_worksheet(
                            worksheet, self.type_registry, previous_manifest,
                            message_base=file_name,
                        )
                    proto_schema.validate_excel_fields(worksheet)

                # 先转换客户端数据（如果有错误会抛出）
                client_data = None
                server_data = None
                omit_empty_fields = (
                    proto_schema.optional_empty_sources()
                    if proto_schema is not None
                    else set()
                )

                if 'c' in export_platform:
                    self._current_platform = 'c'
                    client_data = self._worksheet_to_data(
                        worksheet, 'c', omit_empty_fields
                    )
                if 's' in export_platform:
                    self._current_platform = 's'
                    server_data = self._worksheet_to_data(
                        worksheet, 's', omit_empty_fields
                    )

                # 验证ID引用（在写入文件之前验证）
                if self._reference_validator:
                    try:
                        ref_errors = self._reference_validator.validate_all()
                    except ReferenceError as exc:
                        self._record_issue(
                            'REFERENCE_TABLE_ERROR', str(exc), row=self._current_row
                        )
                        fail_list.append(f"{code.file_name} - {exc}")
                        self._reference_validator.clear()
                        continue
                    if ref_errors:
                        # 有ID引用错误，跳过文件写入
                        for error in ref_errors:
                            fail_list.append(f"{code.file_name} - {error}")
                            location_match = re.search(
                                r"第(?P<row>\d+)行 '(?P<field>[^']+)'=", error
                            )
                            issue_row = self._current_row
                            issue_field = ""
                            issue_column = 0
                            if location_match:
                                issue_row = int(location_match.group('row'))
                                issue_field = location_match.group('field')
                                issue_column = worksheet.get_column_index(issue_field) + 1
                            self._record_issue(
                                'REFERENCE_NOT_FOUND', error, row=issue_row,
                                field=issue_field, column=issue_column,
                            )
                        has_error = True
                        # 清空验证器记录，准备下一个文件
                        self._reference_validator.clear()
                        continue  # 跳过该表，处理下一个
                    self._reference_validator.clear()

                # Protobuf的schema为C/S超集，数据仍按平台分别序列化。
                if code.format == 'pb':
                    targets = []
                    target_modes = []
                    for target_mode in ['c', 's']:
                        if target_mode not in export_platform:
                            continue

                        output_path = client_path if target_mode == 'c' else server_path
                        data = client_data if target_mode == 'c' else server_data
                        rel_dir = os.path.dirname(os.path.relpath(excel_path, table_dir))
                        target_dir = os.path.join(output_path, rel_dir)
                        proto_path = os.path.join(target_dir, f"{file_name}.proto")
                        pb_path = os.path.join(target_dir, f"{file_name}.pb")
                        previous_root = (
                            self._previous_client_path
                            if target_mode == 'c' else self._previous_server_path
                        )
                        if previous_root:
                            old_proto = os.path.join(
                                previous_root, rel_dir, f"{file_name}.proto"
                            )
                            if os.path.exists(old_proto):
                                os.makedirs(target_dir, exist_ok=True)
                                shutil.copy2(old_proto, proto_path)
                        targets.append((proto_path, pb_path, data))
                        target_modes.append(target_mode)

                    csharp_target = ''
                    if csharp_path:
                        csharp_target = os.path.join(
                            csharp_path, rel_dir, f"{file_name}.cs"
                        )

                    written_files = ProtobufExporter(
                        proto_schema, source_file=excel_path
                    ).export_targets(
                        targets,
                        allow_breaking_change=allow_breaking_proto_change,
                        csharp_target=csharp_target,
                        export_pb=export_pb,
                    )
                    generated_files.extend(written_files)
                    for written_file in written_files:
                        artifacts.append(self._artifact_info(
                            written_file, client_path, server_path, csharp_path,
                            excel_path, code.sheet_name,
                        ))
                    for target_mode in target_modes:
                        platform_name = "C" if target_mode == 'c' else "S"
                        success_list.append(
                            (f"{file_name}.pb + " if export_pb else "")
                            + f"{file_name}.proto ({platform_name})"
                        )
                    if csharp_path and target_modes:
                        success_list.append(f"{file_name}.cs (C#)")
                    continue

                # JSON/Lua验证通过，按原有方式写入文件。
                for target_mode in ['c', 's']:
                    if target_mode not in export_platform:
                        continue

                    output_path = client_path if target_mode == 'c' else server_path
                    data = client_data if target_mode == 'c' else server_data

                    # 构建输出路径
                    rel_dir = os.path.dirname(os.path.relpath(excel_path, table_dir))
                    target_dir = os.path.join(output_path, rel_dir)
                    os.makedirs(target_dir, exist_ok=True)

                    # 确保文件名有正确的扩展名（避免重复）
                    file_name = self._strip_output_extension(
                        code.file_name, code.format
                    )
                    file_path = os.path.join(target_dir, f"{file_name}.{code.format}")

                    if code.format == 'json':
                        self._write_json(data, file_path)
                    else:
                        self._write_lua(data, file_path, code.sheet_name)

                    generated_files.append(file_path)
                    artifacts.append(self._artifact_info(
                        file_path, client_path, server_path, csharp_path,
                        excel_path, code.sheet_name,
                    ))

                    platform_name = "C" if target_mode == 'c' else "S"
                    display_name = file_name if not file_name.endswith(f".{code.format}") else file_name
                    success_list.append(f"{display_name}.{code.format} ({platform_name})")

            except Exception as e:
                # 转换/写入错误，删除已生成的文件
                for fpath in generated_files:
                    try:
                        os.remove(fpath)
                    except:
                        pass
                # 格式化错误消息（多行错误合并为一行）
                error_msg = str(e).replace('\n', ' ').replace('  - ', '; ')
                fail_list.append(f"{code.file_name} - {error_msg}")
                if not any(
                    issue.file == self._current_file and issue.sheet == self._current_sheet
                    for issue in self._issues
                ):
                    self._record_issue('EXPORT_ERROR', str(e), row=self._current_row)
                has_error = True
                if self._reference_validator:
                    self._reference_validator.clear()
        
        return len(fail_list) == 0, success_list, fail_list, artifacts

    def _load_previous_manifest(self, excel_path, code, client_path, server_path, table_dir):
        """Read the generated protocol manifest used for auto field-number stability."""
        rel_dir = os.path.dirname(os.path.relpath(excel_path, table_dir))
        file_name = self._strip_output_extension(code.file_name, 'pb')
        previous_client = self._previous_client_path or client_path
        previous_server = self._previous_server_path or server_path
        for output_path, platform in ((previous_client, 'c'), (previous_server, 's')):
            if platform not in (code.platform or 'cs'):
                continue
            proto_path = os.path.join(output_path, rel_dir, f"{file_name}.proto")
            if not os.path.exists(proto_path):
                continue
            try:
                with open(proto_path, 'r', encoding='utf-8') as handle:
                    manifest = extract_manifest(handle.read())
                if manifest:
                    return manifest
                # The exporter will reject unmanaged files transactionally.  Do
                # not fail here, so data/schema validation can report the real
                # error first and leave the old artifacts untouched.
                return None
            except OSError as exc:
                raise ProtoSchemaError(f"读取已有.proto失败: {proto_path}: {exc}") from exc
        return None

    def _read_code_sheets(self, file_path: str) -> List[Any]:
        """读取文件中的 CODE 表配置"""
        from .reader import CodeSheet
        
        codes = []
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            self._log(f"[ERROR] 读取文件失败: {file_path} - {e}")
            self._record_issue(
                'WORKBOOK_READ_ERROR', f"读取工作簿失败: {e}",
                file=os.path.basename(file_path),
            )
            self._unreadable_workbooks.add(os.path.basename(file_path).casefold())
            return codes

        code_sheet = None
        for sheet_name in wb.sheetnames:
            if sheet_name.upper() == 'CODE':
                code_sheet = wb[sheet_name]
                break

        if not code_sheet:
            wb.close()
            return codes

        rows = list(code_sheet.rows)
        if not rows:
            wb.close()
            return codes

        # 跳过表头行
        start_row = 0
        if rows and len(rows) > 0:
            first_cell_value = str(rows[0][0].value or "").strip()
            if first_cell_value in ('表名', '工作表名', 'Sheet', 'sheet_name', '名称'):
                start_row = 1

        for row in rows[start_row:]:
            row_data = [cell.value if cell.value else "" for cell in row]
            code = CodeSheet.from_row(row_data)
            if code and code.sheet_name:
                code.file_path = file_path
                codes.append(code)

        wb.close()
        return codes

    def _preflight_code_sheets(self, xlsx_files: List[str], mode: str) -> None:
        """Validate CODE rows and reserve every platform output path."""
        reserved = {}
        for excel_path in xlsx_files:
            workbook = os.path.basename(excel_path)
            self._current_file = workbook
            self._current_sheet = 'CODE'
            codes = self._read_code_sheets(excel_path)
            if not codes:
                if (
                    not self._has_issue('WORKBOOK_READ_ERROR', workbook)
                    and not self._has_issue('MISSING_CODE', workbook)
                ):
                    self._record_issue(
                        'MISSING_CODE', '缺少CODE工作表或CODE中没有有效导出配置',
                        file=workbook,
                    )
                continue
            for code in codes:
                self._current_sheet = code.sheet_name or 'CODE'
                if not code.sheet_name:
                    self._record_issue(
                        'INVALID_CODE', 'CODE中的工作表名不能为空', file=workbook,
                    )
                    continue
                if code.format not in ('json', 'lua', 'pb'):
                    self._record_issue(
                        'UNKNOWN_OUTPUT_FORMAT',
                        f"未知或缺少输出扩展名: {code.file_name or '(空)'}",
                        file=workbook, sheet=code.sheet_name,
                    )
                    continue
                if code.platform and code.platform not in ('c', 's', 'cs'):
                    self._record_issue(
                        'INVALID_PLATFORM', f"无效的平台: {code.platform}",
                        file=workbook, sheet=code.sheet_name,
                    )
                    continue
                target_platform = code.platform or mode
                for marker, platform in (('c', 'client'), ('s', 'server')):
                    if marker not in mode or marker not in target_platform:
                        continue
                    file_name = self._strip_output_extension(code.file_name, code.format)
                    relative = f"{file_name}.{code.format}"
                    key = (platform, os.path.normcase(relative))
                    owner = reserved.get(key)
                    if owner:
                        self._record_issue(
                            'OUTPUT_PATH_CONFLICT',
                            f"{platform}输出路径冲突: {relative}; 已由 {owner[0]}:{owner[1]} 使用",
                            file=workbook, sheet=code.sheet_name,
                        )
                    else:
                        reserved[key] = (workbook, code.sheet_name)

    def _read_worksheet(self, file_path: str, sheet_name: str):
        """读取工作表"""
        from .core import WorkSheet, FieldInfo
        
        wb = None
        formula_wb = None
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            formula_wb = load_workbook(file_path, read_only=True, data_only=False)
        except Exception as e:
            if wb is not None:
                wb.close()
            if formula_wb is not None:
                formula_wb.close()
            self._record_issue(
                'WORKBOOK_READ_ERROR', f"读取工作表失败: {e}",
                file=os.path.basename(file_path), sheet=sheet_name,
            )
            return None

        if sheet_name not in wb.sheetnames:
            wb.close()
            formula_wb.close()
            return None

        ws = wb[sheet_name]
        formula_ws = formula_wb[sheet_name]
        rows = list(ws.rows)
        formula_rows = list(formula_ws.rows)
        if len(rows) < 4:
            wb.close()
            formula_wb.close()
            return None

        worksheet = WorkSheet(sheet_name)

        # 解析4行表头格式
        names_row = rows[0]
        types_row = rows[1]
        platforms_row = rows[2]
        descs_row = rows[3]

        field_info = {}
        names = []
        
        for col_idx, name_cell in enumerate(names_row):
            name = str(name_cell.value).strip() if name_cell.value else ""
            names.append(name)
            
            if name:
                field_type = str(types_row[col_idx].value).strip() if col_idx < len(types_row) and types_row[col_idx].value else "str"
                platform = str(platforms_row[col_idx].value).strip() if col_idx < len(platforms_row) and platforms_row[col_idx].value else "cs"
                desc = str(descs_row[col_idx].value).strip() if col_idx < len(descs_row) and descs_row[col_idx].value else ""
                
                field_info[name] = FieldInfo(name, field_type, platform, desc)

        # 创建列索引映射
        column_index = {name: idx for idx, name in enumerate(names) if name}
        worksheet.set_field_info(field_info, column_index)

        # 读取数据行
        for row_offset, row in enumerate(rows[4:], start=5):
            data = [cell.value for cell in row]
            formula_row = formula_rows[row_offset - 1] if row_offset <= len(formula_rows) else ()
            for col_idx, formula_cell in enumerate(formula_row):
                cached_value = data[col_idx] if col_idx < len(data) else None
                is_formula = (
                    formula_cell.data_type == 'f'
                    or (isinstance(formula_cell.value, str) and formula_cell.value.startswith('='))
                )
                if is_formula and cached_value is None:
                    field_name = names[col_idx] if col_idx < len(names) else ""
                    self._record_issue(
                        'FORMULA_NO_CACHED_VALUE',
                        '公式没有缓存计算结果，请用Excel重新计算并保存工作簿',
                        file=os.path.basename(file_path), sheet=sheet_name,
                        row=row_offset, column=col_idx + 1, field=field_name,
                        raw_value=formula_cell.value,
                    )
            worksheet.add_row(data)

        wb.close()
        formula_wb.close()
        return worksheet

    def export_all(self, table_dir: str, client_path: str, server_path: str,
                   mode: str = 'cs', filename: str = None,
                   allow_breaking_proto_change: bool = False,
                   csharp_path: str = '', export_pb: bool = True,
                   validation_only: bool = False, asset_root: str = '',
                   locale: str = None) -> dict:
        """Validate and export a batch, returning stable structured diagnostics."""
        self._issues = []
        self._issue_keys = set()
        self._asset_root = asset_root or client_path or ''
        self._previous_client_path = client_path
        self._previous_server_path = server_path
        self._unreadable_workbooks = set()
        if not asset_root and client_path:
            self._log(
                "Warning: assetRoot is not configured; path() checks use the "
                "legacy client output directory"
            )
        elif not self._asset_root:
            self._log("Warning: assetRoot is not configured; path() existence checks are disabled")

        with tempfile.TemporaryDirectory(prefix="excel2json-stage-") as temp_dir:
            stage_client = os.path.join(temp_dir, 'client')
            stage_server = os.path.join(temp_dir, 'server')
            stage_csharp = os.path.join(temp_dir, 'csharp') if csharp_path else ''
            effective_table_dir = table_dir
            if validation_only and os.path.isdir(table_dir):
                effective_table_dir = os.path.join(temp_dir, 'tables')
                os.makedirs(effective_table_dir, exist_ok=True)
                for entry in Path(table_dir).iterdir():
                    if entry.is_file() and entry.suffix.lower() == '.xlsx':
                        shutil.copy2(entry, Path(effective_table_dir) / entry.name)
            result = self._export_all_to_paths(
                effective_table_dir, stage_client, stage_server, mode, filename,
                allow_breaking_proto_change, stage_csharp, export_pb, locale,
            )

            generated = result.pop('_generated_artifacts', [])
            workbooks = result.pop('_workbooks', [])
            success_messages = result.pop('_success_messages', [])
            result['artifacts'] = []
            result['changes'] = {}
            result['manifests'] = {}

            if result.get('success') and not self._issues:
                try:
                    prepared = prepare_batch_commit(
                        generated, workbooks, client_path, server_path, csharp_path,
                        mode, filename is not None,
                        preserve_existing_formats=('pb',) if not export_pb else (),
                    )
                    result['artifacts'] = prepared['artifacts']
                    result['changes'] = prepared['changes']
                    result['manifests'] = prepared['manifests']
                    if not validation_only:
                        commit_files(prepared['writes'], prepared['deletes'])
                    for workbook, message in success_messages:
                        self._log(f"{workbook}\n  [OK] {message}")
                except IncrementalManifestRequiredError as exc:
                    self._record_issue('INCREMENTAL_MANIFEST_REQUIRED', str(exc))
                    result['success'] = False
                    result['fail_count'] = max(1, result.get('fail_count', 0))
                except (ValueError, OSError, AtomicCommitError) as exc:
                    self._record_issue('MANIFEST_OR_COMMIT_ERROR', str(exc))
                    result['success'] = False
                    result['fail_count'] = max(1, result.get('fail_count', 0))

        if result.get('error') and not self._issues:
            self._record_issue('EXPORT_ERROR', result['error'])
        result['issues'] = self._issues_as_dicts()
        result.setdefault('artifacts', [])
        result.setdefault('changes', {})
        result.setdefault('manifests', {})
        if self._issues:
            result['success'] = False
            failed_files = {
                issue.file.lower() for issue in self._issues if issue.file
                and issue.file.lower().endswith('.xlsx')
            }
            result['fail_count'] = max(
                result.get('fail_count', 0), len(failed_files) or 1
            )
            result['success_count'] = max(
                0, result.get('count', 0) - result['fail_count']
            )
        return result

    def _export_all_to_paths(self, table_dir: str, client_path: str, server_path: str,
                             mode: str = 'cs', filename: str = None,
                             allow_breaking_proto_change: bool = False,
                             csharp_path: str = '', export_pb: bool = True,
                             locale: str = None) -> dict:
        """
        导出所有配置文件
        
        Args:
            table_dir: Excel文件目录
            client_path: 客户端输出目录
            server_path: 服务端输出目录
            mode: 导出模式 'c'/'s'/'cs'
            filename: 指定文件名（可选）
        
        Returns:
            {'success': bool, 'count': int, 'success_count': int, 'fail_count': int}
        """
        try:
            # Existing files are never overwritten; missing projects get the
            # compatible three-column template.
            TypeDefinitionTemplate.ensure_exists(table_dir, locale=locale)
            self.type_registry = TypeRegistry(table_dir)
        except Exception as e:
            return {
                'success': False,
                'count': 0,
                'success_count': 0,
                'fail_count': 1,
                'error': f"类型定义加载失败: {e}"
            }
        
        # 初始化引用验证器
        self._reference_validator = ReferenceValidator(table_dir)

        # 保存客户端路径用于路径验证
        self._client_path = client_path

        # 获取所有Excel文件
        xlsx_files = self._get_all_xlsx_files(table_dir)
        if not xlsx_files:
            code = 'SELECTED_FILE_NOT_FOUND' if filename is not None else 'NO_WORKBOOKS'
            message = (
                f"指定文件没有匹配到任何工作簿: {filename}"
                if filename is not None else "表格目录中没有可导出的.xlsx工作簿"
            )
            self._record_issue(code, message, file=filename or table_dir)
            return {
                'success': False, 'count': 0, 'success_count': 0,
                'fail_count': 1, '_generated_artifacts': [], '_workbooks': [],
                '_success_messages': [],
            }
        
        if filename is not None:
            # 支持逗号分隔的多个文件名
            filename_list = [
                os.path.splitext(f.strip())[0].lower()
                for f in filename.split(',') if f.strip()
            ]
            if not filename_list:
                self._record_issue(
                    'SELECTED_FILE_NOT_FOUND', '指定文件名不能为空', file=filename,
                )
                return {
                    'success': False, 'count': 0, 'success_count': 0,
                    'fail_count': 1, '_generated_artifacts': [], '_workbooks': [],
                    '_success_messages': [],
                }
            xlsx_files = [f for f in xlsx_files 
                         if os.path.splitext(os.path.basename(f))[0].lower() in filename_list]
            if not xlsx_files:
                self._record_issue(
                    'SELECTED_FILE_NOT_FOUND',
                    f"指定文件没有匹配到任何工作簿: {filename}", file=filename,
                )
                return {
                    'success': False, 'count': 0, 'success_count': 0,
                    'fail_count': 1, '_generated_artifacts': [], '_workbooks': [],
                    '_success_messages': [],
                }

        self._preflight_code_sheets(xlsx_files, mode)

        total_files = len(xlsx_files)
        total_success = 0
        total_fail = 0
        generated_artifacts = []
        success_messages = []
        
        for excel_path in xlsx_files:
            xlsx_name = os.path.basename(excel_path)

            if xlsx_name.casefold() in self._unreadable_workbooks:
                total_fail += 1
                continue

            success, success_list, fail_list, file_artifacts = self.export_file(
                excel_path, client_path, server_path, mode,
                allow_breaking_proto_change,
                csharp_path,
                export_pb,
            )
            generated_artifacts.extend(file_artifacts)

            # 只在完全成功时输出成功日志
            if success and not fail_list:
                for item in success_list:
                    success_messages.append((xlsx_name, item))
                total_success += 1
            else:
                # 有失败时，只输出失败日志
                for item in fail_list:
                    self._log(f"{xlsx_name}\n  [ERROR] {item}")
                total_fail += 1
        
        return {
            'success': total_fail == 0,
            'count': total_files,
            'success_count': total_success,
            'fail_count': total_fail,
            '_generated_artifacts': generated_artifacts,
            '_workbooks': [os.path.basename(path) for path in xlsx_files],
            '_success_messages': success_messages,
        }

    def _get_all_xlsx_files(self, directory: str) -> List[str]:
        """获取所有Excel文件（排除TypeDefinition.xlsx）"""
        files = []
        if not os.path.exists(directory):
            return files
        
        for filename in os.listdir(directory):
            if filename.endswith('.xlsx') and not filename.startswith('~$'):
                # 排除TypeDefinition.xlsx
                if filename.lower() != 'typedefinition.xlsx':
                    files.append(os.path.join(directory, filename))
        
        return sorted(files)
