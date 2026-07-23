"""Localized teaching worksheets embedded in TypeDefinition.xlsx."""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


SAMPLE_SHEET_NAMES = {
    "en": ("Item", "Reward"),
    "zh-CN": ("物品表", "奖励表"),
    "ja": ("アイテム", "報酬"),
    "ko": ("아이템", "보상"),
    "es": ("Objetos", "Recompensas"),
    "zh-TW": ("物品表", "獎勵表"),
}


SAMPLE_COPY = {
    "en": {
        "note": (
            "{item_sheet} and {reward_sheet} are teaching worksheets only: "
            "TypeDefinition.xlsx is skipped during export. Copy both sheets into "
            "GameConfig.xlsx and add a business CODE sheet to run them. Row 1 uses "
            "English camelCase code fields; row 2 uses localized type names except the "
            "referenced itemId source keeps canonical int. Constraint names after + are "
            "fixed syntax. The second find_id argument is a label, not a sheet selector."
        ),
        "alias_description": "English sample alias for {target}",
        "reference_descriptions": (
            "Item ID in GameConfig.xlsx",
            "One-dimensional item ID list separated by #",
            "Two-dimensional item ID list separated by | and #",
            "Three-dimensional item ID list separated by _, |, and #",
        ),
        "item": (
            "Unique item ID; canonical int keeps references as scalar IDs",
            "Required display name", "Quality: white / green / blue",
            "Client-only icon path",
        ),
        "reward": (
            "Unique reward ID", "Direct {item_sheet}.itemId reference", "Item ID list",
            "Matches itemIds length", "Server-only closed range",
        ),
    },
    "zh-CN": {
        "note": (
            "{item_sheet}、{reward_sheet}仅为格式示例：导出时会跳过 TypeDefinition.xlsx。实际使用时请把两个分表复制到 "
            "GameConfig.xlsx，并添加业务 CODE 表。第 1 行固定使用英文驼峰代码字段；第 2 行使用中文类型名，"
            "但作为引用源的 itemId 保留标准 int。+ 后的约束名是固定语法。find_id 第 2 个参数只是显示标签，"
            "不用于选择分表。"
        ),
        "alias_description": "中文示例别名，等同于 {target}",
        "reference_descriptions": (
            "引用 GameConfig.xlsx 中物品表的 itemId",
            "物品 ID 一维列表，使用 # 分隔",
            "物品 ID 二维列表，使用 | 和 # 分隔",
            "物品 ID 三维列表，使用 _、| 和 # 分隔",
        ),
        "item": (
            "唯一物品 ID；保留固定 int 以让引用导出为纯 ID",
            "必填且唯一的显示名", "品质：white / green / blue", "仅客户端图标路径",
        ),
        "reward": (
            "唯一奖励 ID", "直接引用{item_sheet}.itemId", "物品 ID 列表",
            "长度必须与 itemIds 相同", "仅服务端闭区间概率",
        ),
    },
    "ja": {
        "note": (
            "{item_sheet} と {reward_sheet} は形式説明用です。TypeDefinition.xlsx はエクスポート対象外です。実行するには両シートを "
            "GameConfig.xlsx にコピーし、業務用 CODE シートを追加してください。1 行目は英語の camelCase コードフィールド、"
            "2 行目は日本語の型名を使い、参照元 itemId のみ標準 int を維持します。+ 以降の制約名は固定構文です。"
            "find_id の第2引数は表示ラベルです。"
        ),
        "alias_description": "日本語のサンプル別名（{target} と同等）",
        "reference_descriptions": (
            "GameConfig.xlsx のアイテム itemId を参照",
            "# 区切りのアイテム ID 1次元リスト",
            "| と # 区切りのアイテム ID 2次元リスト",
            "_、|、# 区切りのアイテム ID 3次元リスト",
        ),
        "item": (
            "一意の ID。参照をスカラー ID で出力するため int は固定",
            "必須かつ一意の表示名", "品質：white / green / blue", "クライアント専用アイコンパス",
        ),
        "reward": (
            "一意の報酬 ID", "{item_sheet}.itemId を直接参照", "アイテム ID リスト",
            "itemIds と同じ長さ", "サーバー専用の閉区間確率",
        ),
    },
    "ko": {
        "note": (
            "{item_sheet}과 {reward_sheet}은 형식 예시이며 TypeDefinition.xlsx는 내보내기에서 제외됩니다. 실행하려면 두 시트를 "
            "GameConfig.xlsx로 복사하고 업무 CODE 시트를 추가하세요. 1행은 영문 camelCase 코드 필드, 2행은 한국어 타입 이름을 "
            "사용하되 참조 원본 itemId만 표준 int를 유지합니다. + 뒤의 제약 이름은 고정 문법이며 find_id의 두 번째 인수는 "
            "표시용 라벨입니다."
        ),
        "alias_description": "{target}에 대응하는 한국어 예시 별칭",
        "reference_descriptions": (
            "GameConfig.xlsx의 아이템 itemId 참조",
            "# 구분 아이템 ID 1차원 목록",
            "|와 # 구분 아이템 ID 2차원 목록",
            "_, |, # 구분 아이템 ID 3차원 목록",
        ),
        "item": (
            "고유 ID. 참조를 스칼라 ID로 내보내기 위해 int는 고정",
            "필수이자 고유한 표시 이름", "품질: white / green / blue", "클라이언트 전용 아이콘 경로",
        ),
        "reward": (
            "고유 보상 ID", "{item_sheet}.itemId 직접 참조", "아이템 ID 목록",
            "itemIds와 같은 길이", "서버 전용 닫힌 구간 확률",
        ),
    },
    "es": {
        "note": (
            "{item_sheet} y {reward_sheet} son hojas didácticas: TypeDefinition.xlsx se omite al exportar. "
            "Cópielas a GameConfig.xlsx y añada una hoja CODE de negocio para ejecutarlas. "
            "La fila 1 usa campos de código en inglés con camelCase; la fila 2 usa tipos en español, "
            "salvo que el itemId referenciado conserva el int canónico. Los nombres de restricción "
            "después de + son sintaxis fija. El segundo argumento de find_id es una etiqueta."
        ),
        "alias_description": "Alias de ejemplo en español para {target}",
        "reference_descriptions": (
            "ID de objeto en GameConfig.xlsx",
            "Lista unidimensional de IDs separada por #",
            "Lista bidimensional de IDs separada por | y #",
            "Lista tridimensional de IDs separada por _, | y #",
        ),
        "item": (
            "ID único; int se mantiene para exportar referencias escalares",
            "Nombre visible obligatorio y único", "Calidad: white / green / blue",
            "Ruta de icono solo cliente",
        ),
        "reward": (
            "ID de recompensa único", "Referencia directa a {item_sheet}.itemId", "Lista de IDs de objetos",
            "Misma longitud que itemIds", "Rango cerrado solo servidor",
        ),
    },
    "zh-TW": {
        "note": (
            "{item_sheet}、{reward_sheet}僅為格式範例：匯出時會略過 TypeDefinition.xlsx。實際使用時請把兩個分頁複製到 "
            "GameConfig.xlsx，並加入業務 CODE 表。第 1 列固定使用英文駝峰程式欄位；第 2 列使用繁中型別名稱，"
            "但作為引用來源的 itemId 保留標準 int。+ 後的約束名稱是固定語法。find_id 第 2 個參數只是顯示標籤，"
            "不用來選擇分頁。"
        ),
        "alias_description": "繁體中文範例別名，等同於 {target}",
        "reference_descriptions": (
            "引用 GameConfig.xlsx 中物品表的 itemId",
            "物品 ID 一維清單，使用 # 分隔",
            "物品 ID 二維清單，使用 | 與 # 分隔",
            "物品 ID 三維清單，使用 _、| 與 # 分隔",
        ),
        "item": (
            "唯一物品 ID；保留固定 int 讓引用匯出為純 ID",
            "必填且唯一的顯示名稱", "品質：white / green / blue", "僅用戶端圖示路徑",
        ),
        "reward": (
            "唯一獎勵 ID", "直接引用{item_sheet}.itemId", "物品 ID 清單",
            "長度必須與 itemIds 相同", "僅伺服器閉區間機率",
        ),
    },
}


_SAMPLE_ALIAS_SPECS = (
    ("integer", "int", "42"),
    ("text", "string", "hello"),
    ("quality", "qualityEnum", "green"),
    ("icon", "iconPath", "sword"),
    ("integer_list", "intList", "70#30"),
    ("number", "float", "0.25"),
)

_REFERENCE_ROLES = (
    ("item_reference", 0, "1001"),
    ("item_reference_list", 1, "1001#1002"),
    ("item_reference_list2", 2, "1001#1002|1003"),
    ("item_reference_list3", 3, "1001#1002|1003_1004"),
)


LOCALIZED_SAMPLE_TYPE_NAMES = {
    "en": {
        "integer": "integer",
        "text": "text",
        "quality": "quality",
        "icon": "iconAsset",
        "item_reference": "itemId",
        "item_reference_list": "itemList",
        "item_reference_list2": "itemList2",
        "item_reference_list3": "itemList3",
        "integer_list": "integerList",
        "number": "number",
    },
    "zh-CN": {
        "integer": "整数",
        "text": "字符串",
        "quality": "品质",
        "icon": "图标路径",
        "item_reference": "物品ID",
        "item_reference_list": "物品列表",
        "item_reference_list2": "物品列表2",
        "item_reference_list3": "物品列表3",
        "integer_list": "整数列表",
        "number": "浮点数",
    },
    "ja": {
        "integer": "整数",
        "text": "文字列",
        "quality": "品質",
        "icon": "アイコンパス",
        "item_reference": "アイテムID",
        "item_reference_list": "アイテムリスト",
        "item_reference_list2": "アイテムリスト2",
        "item_reference_list3": "アイテムリスト3",
        "integer_list": "整数リスト",
        "number": "浮動小数",
    },
    "ko": {
        "integer": "정수",
        "text": "문자열",
        "quality": "품질",
        "icon": "아이콘경로",
        "item_reference": "아이템ID",
        "item_reference_list": "아이템목록",
        "item_reference_list2": "아이템목록2",
        "item_reference_list3": "아이템목록3",
        "integer_list": "정수목록",
        "number": "실수",
    },
    "es": {
        "integer": "entero",
        "text": "texto",
        "quality": "calidad",
        "icon": "rutaIcono",
        "item_reference": "idObjeto",
        "item_reference_list": "listaObjetos",
        "item_reference_list2": "listaObjetos2",
        "item_reference_list3": "listaObjetos3",
        "integer_list": "listaEnteros",
        "number": "numeroDecimal",
    },
    "zh-TW": {
        "integer": "整數",
        "text": "字串",
        "quality": "品質",
        "icon": "圖示路徑",
        "item_reference": "物品ID",
        "item_reference_list": "物品清單",
        "item_reference_list2": "物品清單2",
        "item_reference_list3": "物品清單3",
        "integer_list": "整數清單",
        "number": "浮點數",
    },
}


def sample_note(locale_id: str) -> str:
    item_sheet, reward_sheet = sample_sheet_names(locale_id)
    return SAMPLE_COPY.get(locale_id, SAMPLE_COPY["en"])["note"].format(
        item_sheet=item_sheet,
        reward_sheet=reward_sheet,
    )


def sample_sheet_names(locale_id: str) -> tuple[str, str]:
    return SAMPLE_SHEET_NAMES.get(locale_id, SAMPLE_SHEET_NAMES["en"])


def localized_sample_type_names(locale_id: str) -> dict[str, str]:
    names = LOCALIZED_SAMPLE_TYPE_NAMES.get(
        locale_id, LOCALIZED_SAMPLE_TYPE_NAMES["en"]
    )
    return dict(names)


def item_reference_expression(locale_id: str, list_depth: int = 0) -> str:
    item_sheet, _reward_sheet = sample_sheet_names(locale_id)
    expression = f"find_id(GameConfig,{item_sheet},itemId)"
    if list_depth:
        expression = f"split_list{list_depth if list_depth > 1 else ''}({expression})"
    return expression


def localized_sample_type_definitions(locale_id: str):
    copy = SAMPLE_COPY.get(locale_id, SAMPLE_COPY["en"])
    names = localized_sample_type_names(locale_id)
    aliases = tuple(
        (
            names[role],
            target,
            copy["alias_description"].format(target=target),
            example,
        )
        for role, target, example in _SAMPLE_ALIAS_SPECS
    )
    references = tuple(
        (
            names[role],
            item_reference_expression(locale_id, list_depth),
            copy["reference_descriptions"][index],
            example,
        )
        for index, (role, list_depth, example) in enumerate(_REFERENCE_ROLES)
    )
    return aliases + references


def _style_row(ws, row: int, max_col: int, fill: str, theme, *, header=False) -> None:
    border = Border(bottom=Side(style="thin", color=theme["border"]))
    for cell in ws[row][:max_col]:
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(
            color=theme["header_font"] if header else theme["text"],
            bold=header,
        )
        cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="left")
        cell.border = border
    ws.row_dimensions[row].height = 28


def _create_sample_sheet(
    workbook,
    sheet_name: str,
    fields,
    rows,
    widths,
    tab_color: str,
    theme,
    *,
    description_row_height: int = 36,
) -> bool:
    if sheet_name in workbook.sheetnames:
        return False
    ws = workbook.create_sheet(sheet_name)
    for field_index in range(4):
        ws.append([field[field_index] for field in fields])
    for row in rows:
        ws.append(list(row))

    styles = (
        (1, theme["header_fill"], True),
        (2, theme["section_fill"], False),
        (3, theme["note_fill"], False),
        (4, theme["compat_fill"], False),
    )
    for row, fill, header in styles:
        _style_row(ws, row, ws.max_column, fill, theme, header=header)
    ws.row_dimensions[4].height = description_row_height
    for row in range(5, ws.max_row + 1):
        _style_row(ws, row, ws.max_column, theme["body_fill"], theme)

    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = tab_color
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    return True


def add_example_sheets(workbook, locale_id: str, theme) -> bool:
    """Add localized copyable teaching sheets without replacing existing sheets."""
    copy = SAMPLE_COPY.get(locale_id, SAMPLE_COPY["en"])
    type_names = localized_sample_type_names(locale_id)
    item_sheet, reward_sheet = sample_sheet_names(locale_id)
    item_descriptions = copy["item"]
    reward_descriptions = tuple(
        description.format(item_sheet=item_sheet)
        for description in copy["reward"]
    )
    changed = _create_sample_sheet(
        workbook,
        item_sheet,
        (
            # Keep referenced ID sources on the canonical scalar name so legacy
            # JSON/Lua reference output remains a plain ID instead of metadata.
            ("itemId", "int+unique()", "CS", item_descriptions[0]),
            ("name", f"{type_names['text']}+required()+unique()", "CS", item_descriptions[1]),
            ("quality", type_names["quality"], "CS", item_descriptions[2]),
            ("icon", type_names["icon"], "C", item_descriptions[3]),
        ),
        ((1001, "Bronze Sword", "green", "sword"), (1002, "Iron Shield", "blue", "shield")),
        {"A": 34, "B": 32, "C": 24, "D": 28},
        theme["tab_item"],
        theme,
        description_row_height=52,
    )
    changed = _create_sample_sheet(
        workbook,
        reward_sheet,
        (
            ("rewardId", f"{type_names['integer']}+unique()", "CS", reward_descriptions[0]),
            ("primaryItemId", f"{type_names['item_reference']}+required()", "CS", reward_descriptions[1]),
            ("itemIds", f"{type_names['item_reference_list']}+len(1,3)", "C", reward_descriptions[2]),
            ("weights", f"{type_names['integer_list']}+equalLen(itemIds)", "C", reward_descriptions[3]),
            ("rate", f"{type_names['number']}+range(0,1)", "S", reward_descriptions[4]),
        ),
        ((1, 1001, "1001#1002", "70#30", 0.25), (2, 1002, "1002", "100", 1.0)),
        {"A": 18, "B": 28, "C": 30, "D": 34, "E": 26},
        theme["tab_reward"],
        theme,
    ) or changed
    return changed
