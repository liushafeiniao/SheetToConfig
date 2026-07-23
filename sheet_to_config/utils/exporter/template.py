# -*- coding: utf-8 -*-
"""
类型定义模板生成器 - 包含完整类型定义和约束说明

【分隔符规则说明】
为了方便使用，系统默认以下分隔符规则：
- 一维列表 (List):  使用 # 分隔，如 1#2#3
- 二维列表 (List2): 使用 | 分隔外层，# 分隔内层，如 1#2|3#4
- 三维列表 (List3): 使用 _ 分隔最外层，| 分隔中层，# 分隔内层，如 1#2|3#4_5#6|7#8

特殊情况可使用 split_list[符号] 自定义分隔符。
"""
import os
import locale as _locale
import tempfile
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from typing import List, Tuple

from .template_samples import (
    add_example_sheets,
    item_reference_expression,
    localized_sample_type_definitions,
    localized_sample_type_names,
    sample_note,
    sample_sheet_names,
)


TEMPLATE_THEME = {
    'title_fill': '173F3A',
    'section_fill': 'DCECE8',
    'header_fill': '27665E',
    'header_font': 'FFFFFF',
    'body_fill': 'FFFFFF',
    'note_fill': 'FFF4CC',
    'compat_fill': 'F3F0FA',
    'border': 'C9D8D4',
    'text': '18302C',
    'muted': '526963',
    'tab_code': '27665E',
    'tab_guide': 'C7922D',
    'tab_examples': '6E5AA8',
    'tab_item': '3B82A0',
    'tab_reward': 'B76E79',
}



def _default_locale() -> str:
    """Choose a template language without importing the Qt GUI i18n layer."""
    try:
        name = (_locale.getlocale()[0] or '').replace('_', '-').lower()
    except (ValueError, TypeError):
        name = ''
    if name.startswith(('zh-tw', 'zh-hk', 'zh-hant')):
        return 'zh-TW'
    if name.startswith('zh'):
        return 'zh-CN'
    for locale_id in ('ja', 'ko', 'es', 'en'):
        if name.startswith(locale_id):
            return locale_id
    return 'en'


class TypeDefinitionTemplate:
    """类型定义模板生成器"""
    
    # 完整的类型定义
    DEFAULT_TYPES: List[Tuple[str, str, str]] = [
        # (类型名, 转换函数, 说明)
        ("int", "int", "整数，空值=0"),
        ("float", "float", "浮点数，空值=0.0"),
        ("string", "string", "字符串，空值=''"),
        ("str", "string", "string的别名"),
        ("bool", "bool", "布尔值，空值=false"),
        ("bytes", "bytes", "UTF-8字节串，空值=b''（用于Protobuf）"),
        ("text_key", "text_key", "文本键，数字转int，其他转string"),
        ("qualityEnum", "enum(string,white,green,blue)", "字符串枚举，校验后仍导出原字符串"),
        ("stageEnum", "enum(int,1,2,3)", "整数枚举，校验后仍导出原整数"),

        # === 列表类型（使用默认分隔符）===
        # 一维列表：使用 # 分隔
        ("intList", "split_list(int)", "一维整数列表，使用#分隔，如：1#2#3"),
        ("strList", "split_list(string)", "一维字符串列表，使用#分隔，如：a#b#c"),
        ("floatList", "split_list(float)", "一维浮点列表，使用#分隔，如：1.5#2.5#3.5"),

        # 二维列表：使用 | 分隔外层，# 分隔内层
        ("intList2", "split_list2(int)", "二维整数列表，外层用|分隔，内层用#分隔，如：1#2|3#4|5#6"),
        ("strList2", "split_list2(string)", "二维字符串列表，如：a#b|c#d"),
        ("floatList2", "split_list2(float)", "二维浮点列表，如：1.5#2.5|3.5#4.5"),

        # 三维列表：使用 _ 分隔最外层，| 分隔中层，# 分隔内层
        ("intList3", "split_list3(int)", "三维整数列表，_分隔最外层，|分隔中层，#分隔内层，如：1#2|3#4_5#6|7#8"),
        ("strList3", "split_list3(string)", "三维字符串列表"),
        ("floatList3", "split_list3(float)", "三维浮点列表"),

        # === 自定义分隔符列表（特殊情况使用）===
        ("strList_xhx", "split_list[_](string)", "下划线分隔的字符串列表，如：a_b_c"),
        ("strList_fh", "split_list[;](string)", "分号分隔的字符串列表，如：a;b;c"),

        # === 奖励类型 ===
        ("award", "split_list2(split_dict(find_id(item,道具,id) id;int minNum;int maxNum;int chance))",
         "奖励结构，如：1001,1,10,50|1002,2,20,30"),

        # === 路径类型 ===
        ("path", "path(res/,.json)", "路径处理，自动拼接前后缀"),
        ("iconPath", "path(res/icon/,.png)", "图标路径"),
    ]

    TYPE_EXAMPLES = {
        'int': '42',
        'float': '0.75',
        'string': 'hello',
        'str': 'hello',
        'bool': 'true',
        'bytes': 'hello',
        'text_key': '1001 / quest_intro',
        'qualityEnum': 'green',
        'stageEnum': '2',
        'intList': '1#2#3',
        'strList': 'a#b#c',
        'floatList': '1.5#2.5#3.5',
        'intList2': '1#2|3#4',
        'strList2': 'a#b|c#d',
        'floatList2': '1.5#2.5|3.5#4.5',
        'intList3': '1#2|3#4_5#6|7#8',
        'strList3': 'a#b|c#d_e#f',
        'floatList3': '1#2|3#4_5#6',
        'strList_xhx': 'a_b_c',
        'strList_fh': 'a;b;c',
        'award': '1001,1,10,50|1002,2,20,30',
        'path': 'config/item',
        'iconPath': 'sword',
    }
    
    # 约束方法说明
    CONSTRAINT_DOCS: List[Tuple[str, str, str]] = [
        ("约束方法", "参数", "说明和示例"),
        ("len", "min,max", "长度限制：intList+len(1,5) - 列表长度1-5"),
        ("len2", "min,max", "二维列表元素长度：intList2+len2(1,3) - 每个内层列表长度1-3"),
        ("len3", "min,max", "三维列表元素长度：intList3+len3(1,2)"),
        ("equalLen", "字段名", "长度相同：物品列表2+equalLen(权重) - 与权重字段长度相同"),
        ("equalLen2", "字段名", "二维列表长度相同"),
        ("coexist", "字段名", "同时存在或同时为空：字段A+coexist(字段B)"),
        ("leastOne", "字段名1,字段名2", "至少一个不为空"),
        ("required / notEmpty", "无", "检查原始单元格必填：int+required()"),
        ("range", "min,max", "闭区间数值范围：float+range(0,1)"),
        ("regex", "pattern", "整值正则匹配：string+regex(^item_[0-9]+$)"),
        ("unique", "无", "按转换后的值检查整列唯一：string+unique()"),
    ]

    HEADER_LABELS = {
        'en': ('Name', 'Convert', 'Description', 'Cell example'),
        'zh-CN': ('类型名', '转换函数', '说明', '单元格示例'),
        'ja': ('名前', '変換関数', '説明', 'セル例'),
        'ko': ('이름', '변환 함수', '설명', '셀 예시'),
        'es': ('Nombre', 'Conversión', 'Descripción', 'Ejemplo de celda'),
        'zh-TW': ('型別名稱', '轉換函式', '說明', '儲存格範例'),
    }

    GUIDE_TEXT = {
        'en': ('TypeDefinition guide', 'List separators', 'Constraints', 'Examples'),
        'zh-CN': ('类型定义说明', '列表分隔符', '约束说明', '示例'),
        'ja': ('型定義ガイド', 'リスト区切り', '制約', '例'),
        'ko': ('타입 정의 안내', '목록 구분자', '제약 조건', '예시'),
        'es': ('Guía de tipos', 'Separadores de listas', 'Restricciones', 'Ejemplos'),
        'zh-TW': ('型別定義說明', '清單分隔符', '約束說明', '範例'),
    }

    GUIDE_COLUMNS = {
        'en': ('Constraint', 'Arguments', 'Description'),
        'zh-CN': ('约束方法', '参数', '说明和示例'),
        'ja': ('制約', '引数', '説明と例'),
        'ko': ('제약 조건', '인수', '설명 및 예시'),
        'es': ('Restricción', 'Argumentos', 'Descripción y ejemplo'),
        'zh-TW': ('約束方法', '參數', '說明與範例'),
    }

    SEPARATOR_ROWS = {
        'en': ('1D: # separates items', '2D: | separates rows, # separates items', '3D: _ separates layers, | rows, # items'),
        'zh-CN': ('一维：# 分隔元素', '二维：| 分隔外层，# 分隔内层', '三维：_ 分隔最外层，| 分隔中层，# 分隔内层'),
        'ja': ('1次元：# で要素を区切る', '2次元：| が行、# が要素を区切る', '3次元：_ が層、| が行、# が要素を区切る'),
        'ko': ('1차원: #로 요소 구분', '2차원: |는 행, #는 요소 구분', '3차원: _는 층, |는 행, #는 요소 구분'),
        'es': ('1D: # separa elementos', '2D: | separa filas y # elementos', '3D: _ separa capas, | filas y # elementos'),
        'zh-TW': ('一維：# 分隔元素', '二維：| 分隔外層，# 分隔內層', '三維：_ 分隔最外層，| 分隔中層，# 分隔內層'),
    }

    GUIDE_COPY = {
        'en': {
            'title': 'TypeDefinition guide',
            'summary': 'Put conversion expressions in CODE. In a data sheet row 2, use only a registered type name plus constraints.',
            'main_path': 'Main path', 'step': 'Step', 'where': 'Where',
            'syntax': 'Copyable syntax', 'valid': 'Valid value',
            'invalid': 'Invalid value', 'meaning': 'Key behavior',
            'code_step': 'Register a reusable type',
            'data_step': 'Use the registered name in data-sheet row 2',
            'value_step': 'Enter cell data from row 5 onward',
            'separators': 'List separators', 'constraints': 'Field constraints',
            'references': 'Cross-workbook references', 'parameter': 'Parameter',
            'limits': 'Output limits and unsupported syntax',
            'file_prefix': 'Target .xlsx exact name or filename prefix.',
            'display_label': 'Display metadata only; it does not select a worksheet.',
            'field': 'Exact target field name in non-CODE/PROTO sheets; it must have an independent scalar definition.',
            'ref_note': 'Register find_id(...) in CODE, then use that registered name in a data sheet. Do not write find_id(...) directly in row 2.',
            'json_note': 'JSON keeps reference metadata/dictionary shape when the converter produces it; Protobuf stores the referenced scalar ID.',
            'bytes_note': 'bytes is Protobuf-only. JSON and Lua exports reject bytes fields.',
            'dict_note': 'split_dict needs an explicit PROTO schema for Protobuf. Dictionary keys are field names, not conversion expressions.',
            'unsupported': 'Not supported: nullable(), default(...), date/time, decimal/long, union types, or string-length constraints. len* applies to lists.',
            'blank': '(blank)', 'same_column': 'same value again in the column',
            'examples_title': 'Copyable examples',
            'examples_summary': 'Copy the Name and Convert cells into CODE, then use the Name in a business workbook data-type row.',
            'category': 'Category', 'name': 'Name', 'convert': 'CODE Convert',
            'data_type': 'Data-sheet row 2', 'cell_input': 'Cell input',
            'note': 'Result / note', 'recommended': 'Recommended',
            'compatibility': 'Compatibility only; prefer the recommended syntax above.',
        },
        'zh-CN': {
            'title': '类型定义说明',
            'summary': '转换表达式写在 TypeDefinition/CODE；业务数据表第 2 行只写“已注册类型名 + 约束”。',
            'main_path': '主路径', 'step': '步骤', 'where': '填写位置',
            'syntax': '可复制语法', 'valid': '有效值', 'invalid': '无效值',
            'meaning': '关键语义', 'code_step': '注册可复用类型',
            'data_step': '在业务数据表第 2 行使用注册名',
            'value_step': '从第 5 行开始填写单元格数据',
            'separators': '列表分隔符', 'constraints': '字段约束',
            'references': '跨工作簿引用', 'parameter': '参数',
            'limits': '输出限制与不支持语法',
            'file_prefix': '目标 .xlsx 的精确文件名或文件名前缀。',
            'display_label': '仅作显示元数据，不用于选择工作表。',
            'field': '非 CODE/PROTO 工作表中的精确字段名；目标列必须有独立标量定义。',
            'ref_note': '先在 CODE 注册 find_id(...)，业务数据表第 2 行再使用注册名；不要直接填写 find_id(...)。',
            'json_note': '转换结果为字典时，JSON 保留引用字典/元数据；Protobuf 只写目标标量 ID。',
            'bytes_note': 'bytes 仅允许 Protobuf；JSON 和 Lua 会拒绝 bytes 字段。',
            'dict_note': 'split_dict 导出 Protobuf 时需要显式 PROTO；字典 key 只是字段名，不是转换表达式。',
            'unsupported': '暂不支持 nullable()、default(...)、日期时间、decimal/long、union 或字符串长度约束；len* 只检查列表。',
            'blank': '（空白）', 'same_column': '同列再次出现相同值',
            'examples_title': '可复制示例',
            'examples_summary': '把 Name 与 Convert 复制到 CODE，再在业务表的类型行使用 Name。',
            'category': '类别', 'name': '名称', 'convert': 'CODE 转换函数',
            'data_type': '数据表第 2 行', 'cell_input': '单元格输入',
            'note': '结果 / 说明', 'recommended': '推荐',
            'compatibility': '仅兼容旧项目；新项目优先使用上方推荐语法。',
        },
        'ja': {
            'title': '型定義ガイド',
            'summary': '変換式は TypeDefinition/CODE に記述し、データシート2行目には登録済み型名と制約だけを記述します。',
            'main_path': '基本手順', 'step': '手順', 'where': '記入場所',
            'syntax': 'コピー可能な構文', 'valid': '有効値', 'invalid': '無効値',
            'meaning': '重要な動作', 'code_step': '再利用型を登録',
            'data_step': 'データシート2行目で登録名を使用',
            'value_step': '5行目以降にセル値を入力',
            'separators': 'リスト区切り', 'constraints': 'フィールド制約',
            'references': 'ブック間参照', 'parameter': '引数',
            'limits': '出力制限と未対応構文',
            'file_prefix': '対象 .xlsx の完全名またはファイル名プレフィックス。',
            'display_label': '表示用メタデータのみ。シート選択には使いません。',
            'field': 'CODE/PROTO 以外の正確なフィールド名。独立したスカラー定義が必要です。',
            'ref_note': 'find_id(...) は CODE に登録し、データシート2行目では登録名を使います。',
            'json_note': 'JSON は変換結果の参照辞書を保持し、Protobuf は参照先スカラーIDを保存します。',
            'bytes_note': 'bytes は Protobuf 専用です。JSON/Lua は拒否します。',
            'dict_note': 'split_dict の Protobuf 出力には明示的な PROTO が必要です。key は名前であり変換式ではありません。',
            'unsupported': 'nullable/default、日時、decimal/long、union、文字列長制約は未対応です。len* はリスト専用です。',
            'blank': '（空）', 'same_column': '同じ列で同じ値を再使用',
            'examples_title': 'コピー可能な例',
            'examples_summary': 'Name と Convert を CODE にコピーし、業務ブックの型行では Name を使います。',
            'category': '分類', 'name': '名前', 'convert': 'CODE 変換',
            'data_type': 'データシート2行目', 'cell_input': 'セル入力',
            'note': '結果 / 注記', 'recommended': '推奨',
            'compatibility': '互換性専用。新規では上の推奨構文を使用してください。',
        },
        'ko': {
            'title': '타입 정의 안내',
            'summary': '변환식은 TypeDefinition/CODE에 쓰고, 데이터 시트 2행에는 등록된 타입 이름과 제약만 씁니다.',
            'main_path': '기본 경로', 'step': '단계', 'where': '작성 위치',
            'syntax': '복사 가능한 구문', 'valid': '유효 값', 'invalid': '잘못된 값',
            'meaning': '핵심 동작', 'code_step': '재사용 타입 등록',
            'data_step': '데이터 시트 2행에서 등록 이름 사용',
            'value_step': '5행부터 셀 데이터 입력',
            'separators': '목록 구분자', 'constraints': '필드 제약',
            'references': '통합 문서 간 참조', 'parameter': '인수',
            'limits': '출력 제한 및 미지원 구문',
            'file_prefix': '대상 .xlsx의 정확한 이름 또는 파일명 접두사입니다.',
            'display_label': '표시 메타데이터일 뿐 워크시트를 선택하지 않습니다.',
            'field': 'CODE/PROTO가 아닌 시트의 정확한 필드명이며 독립 스칼라 정의가 필요합니다.',
            'ref_note': 'find_id(...)를 CODE에 등록한 뒤 데이터 시트 2행에서는 등록 이름을 사용합니다.',
            'json_note': 'JSON은 변환된 참조 사전 모양을 유지하고 Protobuf는 참조 스칼라 ID를 저장합니다.',
            'bytes_note': 'bytes는 Protobuf 전용이며 JSON/Lua에서는 거부됩니다.',
            'dict_note': 'split_dict를 Protobuf로 내보내려면 명시적 PROTO가 필요합니다. key는 이름이지 변환식이 아닙니다.',
            'unsupported': 'nullable/default, 날짜/시간, decimal/long, union, 문자열 길이 제약은 지원하지 않습니다. len*은 목록 전용입니다.',
            'blank': '(빈 값)', 'same_column': '같은 열에서 같은 값 반복',
            'examples_title': '복사 가능한 예시',
            'examples_summary': 'Name과 Convert를 CODE에 복사하고 업무 통합 문서의 타입 행에서는 Name을 사용합니다.',
            'category': '범주', 'name': '이름', 'convert': 'CODE 변환',
            'data_type': '데이터 시트 2행', 'cell_input': '셀 입력',
            'note': '결과 / 설명', 'recommended': '권장',
            'compatibility': '호환성 전용입니다. 새 프로젝트는 위 권장 구문을 사용하세요.',
        },
        'es': {
            'title': 'Guía de TypeDefinition',
            'summary': 'Escriba las conversiones en TypeDefinition/CODE; en la fila 2 de datos use solo el tipo registrado y sus restricciones.',
            'main_path': 'Ruta principal', 'step': 'Paso', 'where': 'Ubicación',
            'syntax': 'Sintaxis copiable', 'valid': 'Valor válido', 'invalid': 'Valor inválido',
            'meaning': 'Comportamiento clave', 'code_step': 'Registrar un tipo reutilizable',
            'data_step': 'Usar el nombre registrado en la fila 2',
            'value_step': 'Introducir datos desde la fila 5',
            'separators': 'Separadores de listas', 'constraints': 'Restricciones de campo',
            'references': 'Referencias entre libros', 'parameter': 'Argumento',
            'limits': 'Límites de salida y sintaxis no admitida',
            'file_prefix': 'Nombre exacto o prefijo del archivo .xlsx de destino.',
            'display_label': 'Solo metadatos visibles; no selecciona la hoja.',
            'field': 'Nombre exacto en hojas no CODE/PROTO; requiere una definición escalar independiente.',
            'ref_note': 'Registre find_id(...) en CODE y use el nombre registrado en la fila 2; no escriba find_id(...) directamente.',
            'json_note': 'JSON conserva la forma de diccionario de la referencia; Protobuf guarda el ID escalar.',
            'bytes_note': 'bytes es exclusivo de Protobuf; JSON y Lua lo rechazan.',
            'dict_note': 'split_dict requiere un PROTO explícito para Protobuf. Las claves son nombres, no expresiones.',
            'unsupported': 'No se admiten nullable/default, fecha/hora, decimal/long, union ni longitud de cadenas. len* es solo para listas.',
            'blank': '(vacío)', 'same_column': 'repetir el mismo valor en la columna',
            'examples_title': 'Ejemplos copiables',
            'examples_summary': 'Copie Name y Convert a CODE y use Name en la fila de tipos del libro de negocio.',
            'category': 'Categoría', 'name': 'Nombre', 'convert': 'Conversión CODE',
            'data_type': 'Fila 2 de datos', 'cell_input': 'Entrada de celda',
            'note': 'Resultado / nota', 'recommended': 'Recomendado',
            'compatibility': 'Solo compatibilidad; para proyectos nuevos use la sintaxis recomendada.',
        },
        'zh-TW': {
            'title': '型別定義說明',
            'summary': '轉換運算式寫在 TypeDefinition/CODE；業務資料表第 2 列只寫「已註冊型別名稱 + 約束」。',
            'main_path': '主要流程', 'step': '步驟', 'where': '填寫位置',
            'syntax': '可複製語法', 'valid': '有效值', 'invalid': '無效值',
            'meaning': '關鍵語意', 'code_step': '註冊可重用型別',
            'data_step': '在業務資料表第 2 列使用註冊名稱',
            'value_step': '從第 5 列開始填寫儲存格資料',
            'separators': '清單分隔符', 'constraints': '欄位約束',
            'references': '跨活頁簿引用', 'parameter': '參數',
            'limits': '輸出限制與不支援語法',
            'file_prefix': '目標 .xlsx 的精確檔名或檔名前綴。',
            'display_label': '僅供顯示，不用來選擇工作表。',
            'field': '非 CODE/PROTO 工作表中的精確欄位名稱；必須有獨立純量定義。',
            'ref_note': '先在 CODE 註冊 find_id(...)，資料表第 2 列再使用註冊名稱；不要直接填 find_id(...)。',
            'json_note': 'JSON 保留轉換後的引用字典形狀；Protobuf 只寫入目標純量 ID。',
            'bytes_note': 'bytes 僅支援 Protobuf；JSON 與 Lua 會拒絕 bytes 欄位。',
            'dict_note': 'split_dict 匯出 Protobuf 時需要明確 PROTO；字典 key 是欄位名稱，不是轉換運算式。',
            'unsupported': '暫不支援 nullable/default、日期時間、decimal/long、union 或字串長度約束；len* 只檢查清單。',
            'blank': '（空白）', 'same_column': '同欄再次出現相同值',
            'examples_title': '可複製範例',
            'examples_summary': '把 Name 與 Convert 複製到 CODE，再在業務活頁簿的型別列使用 Name。',
            'category': '類別', 'name': '名稱', 'convert': 'CODE 轉換函式',
            'data_type': '資料表第 2 列', 'cell_input': '儲存格輸入',
            'note': '結果 / 說明', 'recommended': '推薦',
            'compatibility': '僅相容舊專案；新專案優先使用上方推薦語法。',
        },
    }

    CONSTRAINT_EXAMPLES = (
        ('len', 'intList+len(1,5)', '1#2', '', 'list length 1-5; len(2) means at least 2, not exactly 2'),
        ('len2', 'intList2+len2(1,3)', '1#2|3', '1#2#3#4', 'checks every inner list'),
        ('len3', 'intList3+len3(1,2)', '1#2|3_4', '1#2#3', 'checks every innermost list'),
        ('equalLen', 'intList+equalLen(Weights)', 'IDs=1#2; Weights=10#20', 'IDs=1#2; Weights=10', 'compares converted outer length; skips when current field is empty'),
        ('equalLen2', 'intList2+equalLen2(Weights)', 'IDs=1#2|3; Weights=5#6|7', 'IDs=1#2|3; Weights=5|7', 'outer and corresponding inner lengths must match'),
        ('coexist', 'string+coexist(EndAt)', 'StartAt=1; EndAt=2', 'StartAt=1; EndAt=', 'both fields are present or both blank'),
        ('leastOne', 'string+leastOne(Email,Phone)', 'Email=a@b.com; Phone=', 'Email=; Phone=', 'list every field explicitly; the current field is not added automatically'),
        ('required / notEmpty', 'string+required()', 'hello', '', 'checks the original cell before default conversion; notEmpty() is an alias'),
        ('range', 'float+range(0,1)', '0.25', '1.5', 'closed interval: both endpoints are valid'),
        ('regex', 'string+regex(^item_[0-9]+$)', 'item_12', 'xitem_12', 'full-value regular-expression match'),
        ('unique', 'string+unique()', 'alpha', 'alpha', 'checks converted values across the whole column'),
    )

    EXAMPLE_DEFINITIONS = (
        ('alias', 'positiveInt', 'int', 'positiveInt+range(1,999)', '10', 'recommended'),
        ('forward alias', 'earlyAlias', 'laterType', 'earlyAlias', 'hello', 'laterType may appear on a later CODE row'),
        ('enum', 'statusEnum', 'enum(string,"in progress",done)', 'statusEnum+required()', 'in progress', 'quoted values may contain spaces'),
        ('enum list', 'qualityList', 'split_list(enum(string,white,green,blue))', 'qualityList+len(1,3)', 'white#green', 'recommended'),
        ('custom separator', 'tagList', 'split_list[;](string)', 'tagList', 'red;green;blue', 'recommended'),
        ('dictionary', 'reward', 'split_dict(int id;int count)', 'reward', '1001,2', 'positional input'),
        ('dictionary', 'rewardNamed', 'split_dict(id:int;count:int)', 'rewardNamed', 'id:1001;count:2', 'named input; keys are names, not expressions'),
        ('path', 'rawPath', 'path()', 'rawPath', 'icons/sword.png', 'pass through'),
        ('path', 'resourcePath', 'path(res/)', 'resourcePath', 'item.json', 'prefix only'),
        ('path', 'iconPath2', 'path(res/icon/,.png)', 'iconPath2', 'sword', 'prefix and suffix'),
        ('reference', 'itemId', 'find_id(GameConfig,Item,itemId)', 'itemId+required()', '1001', 'register in CODE; validates GameConfig.xlsx field itemId'),
        ('reference list', 'itemList', 'split_list(find_id(GameConfig,Item,itemId))', 'itemList+len(1,5)', '1001#1002', 'each ID is validated'),
        ('constraint chain', 'safeCode', 'string', 'safeCode+required()+unique()+regex(^item_[0-9]+$)', 'item_12', 'recommended'),
        ('compatibility', 'legacyList', 'split_list_ex(;,string)', 'legacyList', 'a;b;c', 'compatibility'),
        ('compatibility', 'legacyCommon', 'commonStringParamForSplit', 'legacyCommon', 'a#b#c', 'compatibility'),
        ('forward alias target', 'laterType', 'string', 'laterType', 'hello', 'recommended'),
    )

    CRITICAL_CONSTRAINT_NOTES = {
        'en': {
            'len': 'With one argument it means a minimum, not an exact length.',
            'equalLen': 'Validation is skipped when the current field is empty.',
            'leastOne': 'List every participating field explicitly; the current field is not added automatically.',
            'required / notEmpty': 'Both names check the original cell before empty defaults are applied.',
        },
        'zh-CN': {
            'len': '单参数表示最小长度，不表示固定长度。',
            'equalLen': '当前字段为空时跳过校验。',
            'leastOne': '必须显式列出所有参与字段，当前字段不会自动加入。',
            'required / notEmpty': '两者同义，均在空值默认转换之前检查原始单元格。',
        },
        'ja': {
            'len': '引数が1つの場合は最小長であり、固定長ではありません。',
            'equalLen': '現在のフィールドが空なら検証をスキップします。',
            'leastOne': '対象フィールドをすべて明示します。現在のフィールドは自動追加されません。',
            'required / notEmpty': 'どちらも空値の既定変換前に元セルを検査します。',
        },
        'ko': {
            'len': '인수가 하나면 고정 길이가 아니라 최소 길이입니다.',
            'equalLen': '현재 필드가 비어 있으면 검사를 건너뜁니다.',
            'leastOne': '참여 필드를 모두 명시해야 하며 현재 필드는 자동 추가되지 않습니다.',
            'required / notEmpty': '둘 다 빈 값 기본 변환 전에 원본 셀을 검사합니다.',
        },
        'es': {
            'len': 'Con un argumento indica un mínimo, no una longitud exacta.',
            'equalLen': 'Se omite la validación si el campo actual está vacío.',
            'leastOne': 'Enumere todos los campos; el campo actual no se añade automáticamente.',
            'required / notEmpty': 'Ambos comprueban la celda original antes de aplicar valores vacíos.',
        },
        'zh-TW': {
            'len': '單參數表示最小長度，不表示固定長度。',
            'equalLen': '目前欄位為空時會略過驗證。',
            'leastOne': '必須明確列出所有參與欄位，目前欄位不會自動加入。',
            'required / notEmpty': '兩者同義，皆在空值預設轉換前檢查原始儲存格。',
        },
    }

    EXAMPLE_CATEGORY_LABELS = {
        'en': {},
        'zh-CN': {
            'alias': '别名', 'forward alias': '向后引用别名', 'enum': '枚举',
            'enum list': '枚举列表', 'custom separator': '自定义分隔符',
            'dictionary': '字典', 'path': '路径', 'reference': '引用',
            'reference list': '引用列表', 'constraint chain': '约束链',
            'compatibility': '兼容语法', 'forward alias target': '向后别名目标',
        },
        'ja': {
            'alias': '別名', 'forward alias': '後方参照別名', 'enum': '列挙',
            'enum list': '列挙リスト', 'custom separator': 'カスタム区切り',
            'dictionary': '辞書', 'path': 'パス', 'reference': '参照',
            'reference list': '参照リスト', 'constraint chain': '制約チェーン',
            'compatibility': '互換構文', 'forward alias target': '後方別名の対象',
        },
        'ko': {
            'alias': '별칭', 'forward alias': '후방 참조 별칭', 'enum': '열거형',
            'enum list': '열거형 목록', 'custom separator': '사용자 구분자',
            'dictionary': '사전', 'path': '경로', 'reference': '참조',
            'reference list': '참조 목록', 'constraint chain': '제약 체인',
            'compatibility': '호환 구문', 'forward alias target': '후방 별칭 대상',
        },
        'es': {
            'alias': 'Alias', 'forward alias': 'Alias adelantado', 'enum': 'Enum',
            'enum list': 'Lista enum', 'custom separator': 'Separador personalizado',
            'dictionary': 'Diccionario', 'path': 'Ruta', 'reference': 'Referencia',
            'reference list': 'Lista de referencias', 'constraint chain': 'Cadena de restricciones',
            'compatibility': 'Compatibilidad', 'forward alias target': 'Destino del alias',
        },
        'zh-TW': {
            'alias': '別名', 'forward alias': '向後引用別名', 'enum': '列舉',
            'enum list': '列舉清單', 'custom separator': '自訂分隔符',
            'dictionary': '字典', 'path': '路徑', 'reference': '引用',
            'reference list': '引用清單', 'constraint chain': '約束鏈',
            'compatibility': '相容語法', 'forward alias target': '向後別名目標',
        },
    }

    EXAMPLE_NOTE_TRANSLATIONS = {
        'zh-CN': {
            'laterType may appear on a later CODE row': 'laterType 可以定义在 CODE 的后续行。',
            'quoted values may contain spaces': '带引号的枚举值可以包含空格。',
            'positional input': '按字段定义顺序填写。',
            'named input; keys are names, not expressions': '命名输入；key 是字段名，不是转换表达式。',
            'pass through': '原样保留路径。', 'prefix only': '只添加前缀。',
            'prefix and suffix': '同时添加前缀与后缀。',
            'register in CODE; validates GameConfig.xlsx field itemId': '先在 CODE 注册；校验 GameConfig.xlsx 的 itemId 字段。',
            'each ID is validated': '列表中的每个 ID 都会校验。',
        },
        'ja': {
            'laterType may appear on a later CODE row': 'laterType は CODE の後の行で定義できます。',
            'quoted values may contain spaces': '引用符付き列挙値には空白を含められます。',
            'positional input': 'フィールド定義順に入力します。',
            'named input; keys are names, not expressions': '名前付き入力。key は名前であり変換式ではありません。',
            'pass through': 'パスをそのまま保持します。', 'prefix only': '接頭辞のみ追加します。',
            'prefix and suffix': '接頭辞と接尾辞を追加します。',
            'register in CODE; validates GameConfig.xlsx field itemId': 'CODE に登録し、GameConfig.xlsx の itemId を検証します。',
            'each ID is validated': 'リスト内の各 ID を検証します。',
        },
        'ko': {
            'laterType may appear on a later CODE row': 'laterType은 CODE의 뒤 행에 정의할 수 있습니다.',
            'quoted values may contain spaces': '따옴표로 감싼 열거 값에는 공백을 넣을 수 있습니다.',
            'positional input': '필드 정의 순서대로 입력합니다.',
            'named input; keys are names, not expressions': '이름 입력이며 key는 이름이지 변환식이 아닙니다.',
            'pass through': '경로를 그대로 유지합니다.', 'prefix only': '접두사만 추가합니다.',
            'prefix and suffix': '접두사와 접미사를 모두 추가합니다.',
            'register in CODE; validates GameConfig.xlsx field itemId': 'CODE에 등록하고 GameConfig.xlsx의 itemId를 검사합니다.',
            'each ID is validated': '목록의 각 ID를 검사합니다.',
        },
        'es': {
            'laterType may appear on a later CODE row': 'laterType puede definirse en una fila posterior de CODE.',
            'quoted values may contain spaces': 'Los valores entre comillas pueden contener espacios.',
            'positional input': 'Introduzca los valores en el orden definido.',
            'named input; keys are names, not expressions': 'Entrada con nombres; las claves son nombres, no expresiones.',
            'pass through': 'Conserva la ruta sin cambios.', 'prefix only': 'Añade solo el prefijo.',
            'prefix and suffix': 'Añade prefijo y sufijo.',
            'register in CODE; validates GameConfig.xlsx field itemId': 'Regístrelo en CODE; valida el campo itemId de GameConfig.xlsx.',
            'each ID is validated': 'Se valida cada ID de la lista.',
        },
        'zh-TW': {
            'laterType may appear on a later CODE row': 'laterType 可以定義在 CODE 的後續列。',
            'quoted values may contain spaces': '加引號的列舉值可以包含空格。',
            'positional input': '依欄位定義順序填寫。',
            'named input; keys are names, not expressions': '命名輸入；key 是欄位名稱，不是轉換運算式。',
            'pass through': '原樣保留路徑。', 'prefix only': '只加入前綴。',
            'prefix and suffix': '同時加入前綴與後綴。',
            'register in CODE; validates GameConfig.xlsx field itemId': '先在 CODE 註冊；驗證 GameConfig.xlsx 的 itemId 欄位。',
            'each ID is validated': '清單中的每個 ID 都會驗證。',
        },
    }

    _LOCALIZED_GROUPS = {
        'en': {
            'int': 'Integer; empty values become 0', 'float': 'Float; empty values become 0.0',
            'string': 'String; empty values become an empty string', 'str': 'Alias of string',
            'bool': 'Boolean; empty values become false', 'bytes': 'UTF-8 bytes for Protobuf',
            'text_key': 'Text key: numeric values become int, others string',
            'enum': 'Enum values are validated and exported unchanged', 'list': 'List using the configured separator',
            'path': 'Path with the configured prefix and suffix', 'ref': 'Reference ID from another workbook',
            'dict': 'Structured dictionary value',
        },
        'zh-CN': {
            'int': '整数，空值=0', 'float': '浮点数，空值=0.0', 'string': '字符串，空值为空字符串',
            'str': 'string 的别名', 'bool': '布尔值，空值=false', 'bytes': 'UTF-8 字节串（用于 Protobuf）',
            'text_key': '文本键：数字转整数，其他转字符串', 'enum': '枚举值校验后原样导出',
            'list': '按配置分隔符解析列表', 'path': '按前后缀处理路径', 'ref': '引用其他工作簿的 ID',
            'dict': '结构化字典值',
        },
        'ja': {
            'int': '整数、空値は0', 'float': '浮動小数、空値は0.0', 'string': '文字列、空値は空文字列',
            'str': 'string の別名', 'bool': '真偽値、空値はfalse', 'bytes': 'Protobuf 用 UTF-8 バイト列',
            'text_key': 'テキストキー：数値は整数、それ以外は文字列', 'enum': '列挙値を検証してそのまま出力',
            'list': '指定した区切りでリストを解析', 'path': '前後の接頭辞・接尾辞でパスを処理',
            'ref': '別のワークブックの ID を参照', 'dict': '構造化された辞書値',
        },
        'ko': {
            'int': '정수, 빈 값은 0', 'float': '실수, 빈 값은 0.0', 'string': '문자열, 빈 값은 빈 문자열',
            'str': 'string의 별칭', 'bool': '불리언, 빈 값은 false', 'bytes': 'Protobuf용 UTF-8 바이트',
            'text_key': '텍스트 키: 숫자는 정수, 그 외는 문자열', 'enum': '열거형 값을 검증하고 그대로 출력',
            'list': '설정된 구분자로 목록을 변환', 'path': '지정한 접두사와 접미사로 경로 처리',
            'ref': '다른 워크북의 ID 참조', 'dict': '구조화된 딕셔너리 값',
        },
        'es': {
            'int': 'Entero; los valores vacíos son 0', 'float': 'Decimal; los valores vacíos son 0.0',
            'string': 'Cadena; los valores vacíos son cadena vacía', 'str': 'Alias de string',
            'bool': 'Booleano; los valores vacíos son false', 'bytes': 'Bytes UTF-8 para Protobuf',
            'text_key': 'Clave de texto: números a entero, otros a cadena', 'enum': 'Enum validado y exportado sin cambios',
            'list': 'Lista separada según el delimitador configurado', 'path': 'Ruta con prefijo y sufijo configurados',
            'ref': 'ID de referencia de otro libro', 'dict': 'Valor de diccionario estructurado',
        },
        'zh-TW': {
            'int': '整數，空值=0', 'float': '浮點數，空值=0.0', 'string': '字串，空值為空字串',
            'str': 'string 的別名', 'bool': '布林值，空值=false', 'bytes': 'Protobuf 使用的 UTF-8 位元組',
            'text_key': '文字鍵：數字轉整數，其餘轉字串', 'enum': '列舉值驗證後原樣輸出',
            'list': '依設定分隔符解析清單', 'path': '依前後綴處理路徑', 'ref': '引用其他活頁簿的 ID',
            'dict': '結構化字典值',
        },
    }

    @classmethod
    def _localized_description(cls, type_name: str, default: str, locale_id: str) -> str:
        """Return a localized explanation while keeping the machine expression stable."""
        groups = cls._LOCALIZED_GROUPS.get(locale_id, cls._LOCALIZED_GROUPS['en'])
        if type_name in groups:
            return groups[type_name]
        if type_name in {'qualityEnum', 'stageEnum'}:
            return groups['enum']
        if type_name == 'award':
            return groups['dict']
        if type_name == 'iconPath' or type_name == 'path':
            return groups['path']
        if type_name.startswith(('intList', 'strList', 'floatList')):
            return groups['list']
        return default if locale_id == 'zh-CN' else groups['string']

    @classmethod
    def _localized_constraint_rows(cls, locale_id: str):
        if locale_id == 'zh-CN':
            return cls.CONSTRAINT_DOCS[1:]
        descriptions = {
            'en': {
                'len': 'List length range', 'len2': 'Inner list length range', 'len3': 'Nested list length range',
                'equalLen': 'Same outer length as another field', 'equalLen2': 'Same two-dimensional length',
                'coexist': 'Both fields exist or are both empty', 'leastOne': 'At least one field is present',
                'required / notEmpty': 'Original cell is required', 'range': 'Closed numeric range',
                'regex': 'Full-value regular expression', 'unique': 'Converted values must be unique',
            },
            'ja': {
                'len': 'リスト長の範囲', 'len2': '内側リスト長の範囲', 'len3': '入れ子リスト長の範囲',
                'equalLen': '他フィールドと外側の長さを一致', 'equalLen2': '2次元の長さを一致',
                'coexist': '両方存在または両方空', 'leastOne': '少なくとも1つを入力',
                'required / notEmpty': '元セルを必須にする', 'range': '数値の閉区間',
                'regex': '値全体の正規表現', 'unique': '変換後の値を一意にする',
            },
            'ko': {
                'len': '목록 길이 범위', 'len2': '내부 목록 길이 범위', 'len3': '중첩 목록 길이 범위',
                'equalLen': '다른 필드와 외부 길이 일치', 'equalLen2': '2차원 길이 일치',
                'coexist': '두 필드가 함께 존재하거나 비어 있음', 'leastOne': '하나 이상 입력',
                'required / notEmpty': '원본 셀 필수', 'range': '닫힌 숫자 범위',
                'regex': '전체 값 정규식', 'unique': '변환된 값의 유일성',
            },
            'es': {
                'len': 'Rango de longitud de lista', 'len2': 'Rango de longitud interna', 'len3': 'Rango de lista anidada',
                'equalLen': 'Misma longitud externa que otro campo', 'equalLen2': 'Misma longitud bidimensional',
                'coexist': 'Ambos campos presentes o vacíos', 'leastOne': 'Al menos un campo presente',
                'required / notEmpty': 'Celda original obligatoria', 'range': 'Rango numérico cerrado',
                'regex': 'Expresión regular del valor completo', 'unique': 'Valores convertidos únicos',
            },
            'zh-TW': {
                'len': '清單長度範圍', 'len2': '內層清單長度範圍', 'len3': '巢狀清單長度範圍',
                'equalLen': '與其他欄位外層長度相同', 'equalLen2': '二維長度相同',
                'coexist': '兩欄同時存在或同時為空', 'leastOne': '至少填寫一欄',
                'required / notEmpty': '原始儲存格必填', 'range': '閉區間數值範圍',
                'regex': '完整值正規表示式', 'unique': '轉換後的值必須唯一',
            },
        }.get(locale_id, {})
        return [
            (name, args, descriptions.get(name, description))
            for name, args, description in cls.CONSTRAINT_DOCS[1:]
        ]

    @classmethod
    def _copy(cls, locale_id: str):
        return cls.GUIDE_COPY.get(locale_id, cls.GUIDE_COPY['en'])

    @classmethod
    def _style_row(cls, ws, row: int, max_col: int, kind: str = 'body') -> None:
        theme = TEMPLATE_THEME
        styles = {
            'title': (theme['title_fill'], theme['header_font'], True, 16),
            'section': (theme['section_fill'], theme['text'], True, 11),
            'header': (theme['header_fill'], theme['header_font'], True, 10),
            'body': (theme['body_fill'], theme['text'], False, 10),
            'note': (theme['note_fill'], theme['text'], False, 10),
            'compat': (theme['compat_fill'], theme['text'], False, 10),
        }
        fill_color, font_color, bold, size = styles[kind]
        border = Border(bottom=Side(style='thin', color=theme['border']))
        for cell in ws[row][:max_col]:
            cell.fill = PatternFill('solid', fgColor=fill_color)
            cell.font = Font(color=font_color, bold=bold, size=size)
            cell.alignment = Alignment(
                vertical='center',
                wrap_text=True,
                horizontal='left',
            )
            cell.border = border
        ws.row_dimensions[row].height = 28 if kind == 'title' else 32

    @classmethod
    def _merged_row(cls, ws, text: str, max_col: int, kind: str) -> int:
        ws.append([text] + [''] * (max_col - 1))
        row = ws.max_row
        cls._style_row(ws, row, max_col, kind)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
        return row

    @classmethod
    def _style_body_range(
        cls, ws, start_row: int, end_row: int, max_col: int, kind: str = 'body'
    ) -> None:
        for row in range(start_row, end_row + 1):
            cls._style_row(ws, row, max_col, kind)

    @classmethod
    def _finish_sheet(
        cls,
        ws,
        widths,
        freeze_panes: str,
        tab_color: str,
        auto_filter: str = None,
    ) -> None:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = freeze_panes
        ws.sheet_properties.tabColor = tab_color
        for column, width in widths.items():
            ws.column_dimensions[column].width = width
        if auto_filter:
            ws.auto_filter.ref = auto_filter

    @classmethod
    def _style_code_sheet(cls, ws) -> None:
        cls._style_row(ws, 1, 4, 'header')
        if ws.max_row >= 2:
            cls._style_body_range(ws, 2, ws.max_row, 4)
        cls._finish_sheet(
            ws,
            {'A': 22, 'B': 78, 'C': 48, 'D': 36},
            'A2',
            TEMPLATE_THEME['tab_code'],
            f'A1:D{ws.max_row}',
        )

    @classmethod
    def _definition_rows(cls, locale_id: str):
        rows = [
            (
                type_name,
                convert_func,
                cls._localized_description(type_name, desc, locale_id),
                cls.TYPE_EXAMPLES[type_name],
            )
            for type_name, convert_func, desc in cls.DEFAULT_TYPES
        ]
        rows.extend(localized_sample_type_definitions(locale_id))
        return rows
    
    @classmethod
    def _write_atomic(cls, workbook, target_path: str) -> None:
        parent = os.path.dirname(os.path.abspath(target_path))
        fd, temp_path = tempfile.mkstemp(prefix='.typedefinition-', suffix='.xlsx', dir=parent)
        os.close(fd)
        try:
            workbook.save(temp_path)
            # Close before replace; Windows keeps workbook handles locked.
            workbook.close()
            os.replace(temp_path, target_path)
        finally:
            if os.path.exists(temp_path):
                os.unlink(temp_path)

    @classmethod
    def _guide_sheet(cls, workbook, locale_id: str):
        if 'Guide' in workbook.sheetnames:
            return False
        copy = cls._copy(locale_id)
        ws = workbook.create_sheet('Guide')

        cls._merged_row(ws, copy['title'], 5, 'title')
        cls._merged_row(ws, copy['summary'], 5, 'note')
        cls._merged_row(ws, sample_note(locale_id), 5, 'note')

        cls._merged_row(ws, copy['main_path'], 5, 'section')
        ws.append([copy['step'], copy['where'], copy['syntax'], copy['valid'], copy['meaning']])
        cls._style_row(ws, ws.max_row, 5, 'header')
        type_names = localized_sample_type_names(locale_id)
        item_reference = type_names['item_reference']
        item_sheet, reward_sheet = sample_sheet_names(locale_id)
        reference_expression = item_reference_expression(locale_id)
        main_rows = (
            (copy['code_step'], 'TypeDefinition.xlsx / CODE', f'{item_reference} | {reference_expression}', item_reference, copy['ref_note']),
            (copy['data_step'], f'GameConfig.xlsx / {reward_sheet} / 2', f'{item_reference}+required()', f'{item_reference}+required()', copy['summary']),
            (copy['value_step'], f'GameConfig.xlsx / {reward_sheet} / 5+', '1001', '1001', copy['field']),
        )
        start = ws.max_row + 1
        for row in main_rows:
            ws.append(row)
        cls._style_body_range(ws, start, ws.max_row, 5)
        ws.append([])

        cls._merged_row(ws, copy['separators'], 5, 'section')
        ws.append([copy['parameter'], copy['syntax'], copy['valid'], copy['invalid'], copy['meaning']])
        cls._style_row(ws, ws.max_row, 5, 'header')
        separator_rows = (
            ('1D / #', 'split_list(int)', '1#2#3', '1,2,3', cls.SEPARATOR_ROWS.get(locale_id, cls.SEPARATOR_ROWS['en'])[0]),
            ('2D / | + #', 'split_list2(int)', '1#2|3#4', '1,2|3,4', cls.SEPARATOR_ROWS.get(locale_id, cls.SEPARATOR_ROWS['en'])[1]),
            ('3D / _ + | + #', 'split_list3(int)', '1#2|3#4_5#6', '1,2|3,4', cls.SEPARATOR_ROWS.get(locale_id, cls.SEPARATOR_ROWS['en'])[2]),
        )
        start = ws.max_row + 1
        for row in separator_rows:
            ws.append(row)
        cls._style_body_range(ws, start, ws.max_row, 5)
        ws.append([])

        cls._merged_row(ws, copy['constraints'], 5, 'section')
        constraint_header_row = ws.max_row + 1
        constraint_label = cls.GUIDE_COLUMNS.get(
            locale_id, cls.GUIDE_COLUMNS['en']
        )[0]
        ws.append([constraint_label, copy['syntax'], copy['valid'], copy['invalid'], copy['meaning']])
        cls._style_row(ws, ws.max_row, 5, 'header')
        localized_descriptions = {
            name: description
            for name, _arguments, description in cls._localized_constraint_rows(locale_id)
        }
        critical = cls.CRITICAL_CONSTRAINT_NOTES.get(
            locale_id, cls.CRITICAL_CONSTRAINT_NOTES['en']
        )
        start = ws.max_row + 1
        for name, syntax, valid, invalid, fallback_note in cls.CONSTRAINT_EXAMPLES:
            note = localized_descriptions.get(name, fallback_note)
            if name in critical:
                note = f'{note} {critical[name]}'
            if not invalid:
                invalid = copy['blank']
            if name == 'unique':
                invalid = f'alpha / {copy["same_column"]}'
            ws.append([name, syntax, valid, invalid, note])
        cls._style_body_range(ws, start, ws.max_row, 5)
        constraint_end_row = ws.max_row
        ws.append([])

        cls._merged_row(ws, copy['references'], 5, 'section')
        ws.append([copy['parameter'], copy['syntax'], copy['meaning'], '', ''])
        cls._style_row(ws, ws.max_row, 5, 'header')
        reference_rows = (
            ('file_prefix', 'GameConfig', copy['file_prefix'], '', ''),
            ('display_label', item_sheet, copy['display_label'], '', ''),
            ('field', 'itemId', copy['field'], '', ''),
        )
        start = ws.max_row + 1
        for row in reference_rows:
            ws.append(row)
        cls._style_body_range(ws, start, ws.max_row, 5)
        short_reference = reference_expression.replace('find_id(', 'find(', 1)
        cls._merged_row(ws, f'{reference_expression} / {short_reference}', 5, 'note')
        cls._merged_row(ws, copy['ref_note'], 5, 'note')
        cls._merged_row(ws, copy['json_note'], 5, 'note')
        ws.append([])

        cls._merged_row(ws, copy['limits'], 5, 'section')
        cls._merged_row(ws, copy['bytes_note'], 5, 'note')
        cls._merged_row(ws, copy['dict_note'], 5, 'note')
        cls._merged_row(ws, copy['unsupported'], 5, 'note')

        cls._finish_sheet(
            ws,
            {'A': 25, 'B': 34, 'C': 48, 'D': 32, 'E': 62},
            'A5',
            TEMPLATE_THEME['tab_guide'],
            f'A{constraint_header_row}:E{constraint_end_row}',
        )
        return True

    @classmethod
    def _examples_sheet(cls, workbook, locale_id: str):
        if 'Examples' in workbook.sheetnames:
            return False
        copy = cls._copy(locale_id)
        category_labels = cls.EXAMPLE_CATEGORY_LABELS.get(locale_id, {})
        ws = workbook.create_sheet('Examples')
        cls._merged_row(ws, copy['examples_title'], 6, 'title')
        cls._merged_row(ws, copy['examples_summary'], 6, 'note')
        ws.append([])
        ws.append([
            copy['category'], copy['name'], copy['convert'], copy['data_type'],
            copy['cell_input'], copy['note'],
        ])
        header_row = ws.max_row
        cls._style_row(ws, header_row, 6, 'header')
        type_names = localized_sample_type_names(locale_id)
        for category, name, convert, data_type, cell_input, note in cls.EXAMPLE_DEFINITIONS:
            if category == 'reference':
                name = type_names['item_reference']
                convert = item_reference_expression(locale_id)
                data_type = f'{name}+required()'
            elif category == 'reference list':
                name = type_names['item_reference_list']
                convert = item_reference_expression(locale_id, 1)
                data_type = f'{name}+len(1,5)'
            if note == 'recommended':
                localized_note = copy['recommended']
            elif note == 'compatibility':
                localized_note = copy['compatibility']
            else:
                localized_note = cls.EXAMPLE_NOTE_TRANSLATIONS.get(
                    locale_id, {}
                ).get(note, note)
            ws.append([
                category_labels.get(category, category),
                name,
                convert,
                data_type,
                cell_input,
                localized_note,
            ])
            kind = 'compat' if category == 'compatibility' else 'body'
            cls._style_row(ws, ws.max_row, 6, kind)
        cls._finish_sheet(
            ws,
            {'A': 24, 'B': 23, 'C': 78, 'D': 48, 'E': 34, 'F': 58},
            'A5',
            TEMPLATE_THEME['tab_examples'],
            f'A{header_row}:F{ws.max_row}',
        )
        return True

    @classmethod
    def create_template(cls, table_dir: str, locale: str = None):
        """Create or non-destructively extend a TypeDefinition workbook."""
        type_def_path = os.path.join(table_dir, "TypeDefinition.xlsx")

        os.makedirs(table_dir, exist_ok=True)
        locale_id = locale or _default_locale()
        if locale_id not in cls.HEADER_LABELS:
            locale_id = 'en'

        if os.path.exists(type_def_path):
            workbook = load_workbook(type_def_path)
            changed = False
            if 'CODE' not in workbook.sheetnames:
                workbook.close()
                raise ValueError('TypeDefinition.xlsx is missing the CODE worksheet')
            ws = workbook['CODE']
            existing = {
                str(row[0].value).strip().casefold()
                for row in ws.iter_rows(min_row=2)
                if row and row[0].value
            }
            for type_name, convert_func, desc, example in cls._definition_rows(locale_id):
                if type_name.casefold() not in existing:
                    header_width = max(
                        2,
                        min(4, max(
                            (cell.column for cell in ws[1] if cell.value is not None),
                            default=2,
                        )),
                    )
                    row = [
                        type_name,
                        convert_func,
                        desc,
                        example,
                    ]
                    ws.append(row[:header_width])
                    existing.add(type_name.casefold())
                    changed = True
            changed = cls._guide_sheet(workbook, locale_id) or changed
            changed = cls._examples_sheet(workbook, locale_id) or changed
            changed = add_example_sheets(workbook, locale_id, TEMPLATE_THEME) or changed
            if changed:
                try:
                    cls._write_atomic(workbook, type_def_path)
                finally:
                    workbook.close()
            else:
                workbook.close()
            return

        workbook = Workbook()
        ws = workbook.active
        ws.title = "CODE"
        ws.append(list(cls.HEADER_LABELS[locale_id]))
        for row in cls._definition_rows(locale_id):
            ws.append(list(row))
        cls._style_code_sheet(ws)
        cls._guide_sheet(workbook, locale_id)
        cls._examples_sheet(workbook, locale_id)
        add_example_sheets(workbook, locale_id, TEMPLATE_THEME)
        cls._write_atomic(workbook, type_def_path)

    @classmethod
    def ensure_exists(cls, table_dir: str, locale: str = None):
        """Create a missing workbook without modifying an existing one."""
        type_def_path = os.path.join(table_dir, "TypeDefinition.xlsx")
        if os.path.exists(type_def_path):
            workbook = load_workbook(type_def_path, read_only=True, data_only=True)
            try:
                if 'CODE' not in workbook.sheetnames:
                    raise ValueError('TypeDefinition.xlsx is missing the CODE worksheet')
            finally:
                workbook.close()
            return
        cls.create_template(table_dir, locale=locale)
