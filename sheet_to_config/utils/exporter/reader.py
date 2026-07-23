# -*- coding: utf-8 -*-
"""
Excel 读取器
使用 openpyxl 读取 .xlsx 文件
"""
import os
from openpyxl import load_workbook
from typing import List, Dict, Any, Optional
from .core import WorkSheet


def _tr(key: str, **params: Any) -> str:
    """Load UI translations only when a reader log is actually emitted."""
    from sheet_to_config.i18n import tr

    return tr(key, **params)


class CodeSheet:
    """CODE 表配置"""

    def __init__(self):
        self.sheet_name = ""      # 工作表名称
        self.file_name = ""       # 输出文件名
        self.format = ""          # 输出格式 json/lua/pb
        self.platform = ""        # 目标平台 c/s/cs
        self.converter = ""       # 转换函数名
        self.param = ""           # 参数
        self.full_path = ""       # 完整路径（兼容旧格式）
        self.implicit_format = False

    @classmethod
    def from_row(cls, row: List[Any]) -> Optional['CodeSheet']:
        """从行数据创建 CODE 配置

        用户格式：
        第一列：工作表名
        第二列：输出文件名（可能包含路径，如 ../client/Item.json）
        第三列：平台（c=客户端，s=服务端）
        第四列：（备用）
        """
        if len(row) < 3:
            return None

        code = cls()
        code.sheet_name = str(row[0] or "").strip()
        code.file_name = str(row[1] or "").strip()
        code.platform = str(row[2] or "").strip().lower()  # c/s/cs
        code.converter = str(row[3] or "").strip() if len(row) > 3 else ""
        code.param = str(row[4] or "").strip() if len(row) > 4 else ""

        # 格式从文件名扩展名判断
        lower_file_name = code.file_name.lower()
        if lower_file_name.endswith('.json'):
            code.format = 'json'
        elif lower_file_name.endswith('.lua'):
            code.format = 'lua'
        elif lower_file_name.endswith('.pb'):
            code.format = 'pb'
        elif lower_file_name.endswith('.proto'):
            # .proto 只作为 .pb 的配套产物，不能作为独立导出格式。
            code.format = 'proto'
        elif os.path.splitext(code.file_name)[1] == '':
            # One-version compatibility path for legacy CODE rows.
            code.format = 'json'
            code.implicit_format = True
        else:
            code.format = ''

        # 保存完整路径（用于导出）
        code.full_path = code.file_name
        # 提取纯文件名（用于显示）
        if '/' in code.file_name or '\\' in code.file_name:
            code.file_name = os.path.basename(code.file_name)

        return code


class ExcelReader:
    """Excel 读取器"""

    def __init__(self, log_callback=None):
        self.log_callback = log_callback
        self._cache = {}

    def _log(self, message: str):
        if self.log_callback:
            self.log_callback(message)

    def _is_four_line_header(self, rows: List) -> bool:
        """
        检测是否为四行表头格式
        通过第二行是否有类型定义来判断
        """
        if len(rows) < 4:
            return False
        # 检查第二行是否包含类型标识 (int, str, string, float, bool 等)
        for cell in rows[1]:
            val = str(cell.value or "").lower().strip()
            if val in ('int', 'str', 'string', 'float', 'bool', 'c', 's', 'cs', 'x'):
                return True
        return False

    def read_excel(self, file_path: str, sheet_name: str = None) -> Optional[WorkSheet]:
        """读取 Excel 文件，支持四行表头格式"""
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            self._log(_tr('log.workbook_read_failed', path=file_path, detail=e))
            return None

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                self._log(_tr('log.worksheet_missing', sheet=sheet_name))
                wb.close()
                return None
            ws = wb[sheet_name]
        else:
            ws = wb.active
            sheet_name = ws.title

        worksheet = WorkSheet(sheet_name)
        rows = list(ws.rows)
        if not rows:
            wb.close()
            return worksheet

        # 检测是否为四行表头格式
        if self._is_four_line_header(rows):
            # 四行表头格式
            # 第一行：字段名
            # 第二行：字段类型
            # 第三行：导出端 (C/S/X/CS)
            # 第四行：字段说明
            from .core import FieldInfo

            field_info = {}
            col_count = len(rows[0])
            for col_idx in range(col_count):
                name = str(rows[0][col_idx].value or "").strip()
                field_type = str(rows[1][col_idx].value or "").strip()
                platform = str(rows[2][col_idx].value or "").strip().upper()
                desc = str(rows[3][col_idx].value or "").strip()

                if name:  # 有字段名的列才添加
                    field_info[name] = FieldInfo(name, field_type, platform, desc)

            # 检查是否所有字段都是 X（不导出），如果是则返回空工作表
            if field_info and all(info.platform == 'X' for info in field_info.values()):
                wb.close()
                return worksheet  # 返回空工作表（没有字段和数据）
            
            # 过滤掉 platform 为 X 的字段
            field_info = {name: info for name, info in field_info.items() if info.platform != 'X'}
            
            # 创建字段名到列索引的映射
            column_index = {name: idx for idx, name in enumerate(names) if name and name in field_info}
            worksheet.set_field_info(field_info, column_index)
            data_start_row = 4
        else:
            # 简单单行表头格式（兼容旧版）
            headers = []
            for cell in rows[0]:
                headers.append(cell.value if cell.value else "")
            worksheet.set_headers(headers)
            data_start_row = 0

        # 读取数据行
        for row in rows[data_start_row + 1:]:
            data = [cell.value for cell in row]
            worksheet.add_row(data)

        wb.close()
        return worksheet

    def read_code_sheets(self, file_path: str) -> List[CodeSheet]:
        """读取文件中的 CODE 表配置"""
        codes = []
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            self._log(_tr('log.workbook_read_failed', path=file_path, detail=e))
            return codes

        code_sheet = None
        for sheet_name in wb.sheetnames:
            if sheet_name.upper() == 'CODE':
                code_sheet = wb[sheet_name]
                break

        if not code_sheet:
            wb.close()
            return codes

        rows = list(code_sheet.rows)
        if not rows:
            wb.close()
            return codes

        # 跳过表头行（如果第一列是"表名"等中文，说明是表头）
        start_row = 0
        if rows and len(rows) > 0:
            first_cell_value = str(rows[0][0].value or "").strip()
            if first_cell_value in ('表名', '工作表名', 'Sheet', 'sheet_name', '名称'):
                start_row = 1  # 跳过表头行

        for row in rows[start_row:]:
            row_data = [cell.value if cell.value else "" for cell in row]
            code = CodeSheet.from_row(row_data)
            if code and code.sheet_name:
                codes.append(code)

        wb.close()
        return codes

    def get_all_xlsx_files(self, directory: str) -> List[str]:
        """获取目录下所有 xlsx 文件（排除 TypeDefinition.xlsx）"""
        files = []
        if not os.path.exists(directory):
            return files

        for filename in os.listdir(directory):
            if filename.endswith('.xlsx') and not filename.startswith('~$'):
                # 排除 TypeDefinition.xlsx（类型定义文件）
                if filename.lower() != 'typedefinition.xlsx':
                    files.append(os.path.join(directory, filename))
        return sorted(files)
