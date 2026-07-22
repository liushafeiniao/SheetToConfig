"""
类型注册表模块

管理类型定义，提供严格类型验证：
- 所有类型必须在TypeDefinition.xlsx中定义
- 无默认类型，未定义类型将导致错误
- 支持从类型定义解析转换函数
"""

import os
from typing import Any, Callable, Dict, List, Optional, Set, Tuple
from openpyxl import load_workbook

from .expression import (
    ExpressionSyntaxError, parse_call, parse_field_type, split_top_level, unquote,
)


class UndefinedTypeError(Exception):
    """未定义类型错误"""
    pass


class _ReferenceTypeCycle(Exception):
    """Internal marker for a cycle made only of find/find_id edges."""

    pass


class TypeRegistry:
    """类型注册表"""
    
    def __init__(self, table_dir: str):
        """
        初始化类型注册表
        
        Args:
            table_dir: Excel文件目录
        
        Raises:
            FileNotFoundError: TypeDefinition.xlsx不存在
            UndefinedTypeError: 类型定义解析失败
        """
        self.table_dir = table_dir
        self._types: Dict[str, Dict[str, Any]] = {}
        self._converter_states: Dict[str, str] = {}
        self._reference_field_cache: Dict[
            Tuple[str, str], Tuple[Optional[str], Tuple[Tuple[str, str], ...]]
        ] = {}
        self._load_type_definition()
    
    def _load_type_definition(self):
        """加载类型定义文件"""
        type_def_path = os.path.join(self.table_dir, "TypeDefinition.xlsx")
        
        if not os.path.exists(type_def_path):
            raise FileNotFoundError(
                f"类型定义文件不存在: {type_def_path}\n"
                f"请运行工具自动生成或手动创建TypeDefinition.xlsx"
            )
        
        wb = load_workbook(type_def_path, read_only=True, data_only=True)
        
        if 'CODE' not in wb.sheetnames:
            wb.close()
            raise UndefinedTypeError("TypeDefinition.xlsx中缺少CODE工作表")
        
        ws = wb['CODE']
        rows = list(ws.rows)
        
        if len(rows) < 2:
            wb.close()
            raise UndefinedTypeError("TypeDefinition.xlsx数据不完整")
        
        # 解析表头
        header = [cell.value for cell in rows[0]]
        
        # 查找列索引
        col_indices = {}
        for i, col_name in enumerate(header):
            if col_name:
                col_name_str = str(col_name).strip().lower()
                if col_name_str in {
                    'name', 'type name', 'typename', 'type',
                    '类型名', '类型名称', '名前', '이름', 'nombre', '型別名稱',
                }:
                    col_indices['name'] = i
                elif col_name_str in {
                    'convert', 'convert func', 'converter', 'conversion',
                    '转换函数', '変換関数', '변환 함수', 'conversión', '轉換函式',
                }:
                    col_indices['convert'] = i
        
        # 如果没有找到标准列，假设第一列是名称，第二列是转换函数
        if 'name' not in col_indices:
            col_indices['name'] = 0
        if 'convert' not in col_indices:
            col_indices['convert'] = 1
        
        try:
            # Register raw definitions first so aliases may safely reference a
            # definition that appears later in the workbook.
            for row in rows[1:]:
                values = [cell.value for cell in row]

                if len(values) <= col_indices['name']:
                    continue

                type_name = (
                    str(values[col_indices['name']]).strip()
                    if values[col_indices['name']] else None
                )
                if not type_name:
                    continue

                convert_func_str = ""
                if 'convert' in col_indices and len(values) > col_indices['convert']:
                    convert_func_str = (
                        str(values[col_indices['convert']]).strip()
                        if values[col_indices['convert']] else ""
                    )

                self._types[type_name] = {
                    'name': type_name,
                    'convert_func_str': convert_func_str,
                    'convert_func': None,
                }

            if not self._types:
                raise UndefinedTypeError("TypeDefinition.xlsx中没有找到有效的类型定义")

            # bytes was added with Protobuf support. Keep existing projects
            # working when TypeDefinition.xlsx predates this type.
            if 'bytes' not in self._types:
                self._types['bytes'] = {
                    'name': 'bytes',
                    'convert_func_str': 'bytes',
                    'convert_func': None,
                }

            for type_name in list(self._types):
                self._ensure_converter(type_name, [])
        finally:
            wb.close()
    
    def _ensure_converter(self, type_name: str, chain: List[str]) -> Callable:
        """Build one converter with DFS cycle detection."""
        if type_name not in self._types:
            raise UndefinedTypeError(f"类型别名引用了未定义类型: {type_name}")

        state = self._converter_states.get(type_name, 'UNSEEN')
        if state == 'DONE':
            return self._types[type_name]['convert_func']
        if state == 'BUILDING':
            cycle_start = chain.index(type_name) if type_name in chain else 0
            cycle = chain[cycle_start:] + [type_name]
            raise UndefinedTypeError("类型别名存在循环: " + " -> ".join(cycle))

        self._converter_states[type_name] = 'BUILDING'
        current_chain = chain + [type_name]
        try:
            converter = self._parse_convert_func(
                self._types[type_name].get('convert_func_str', ''),
                current_chain,
            )
        except Exception:
            self._types[type_name]['convert_func'] = None
            self._converter_states[type_name] = 'UNSEEN'
            raise

        self._types[type_name]['convert_func'] = converter
        self._converter_states[type_name] = 'DONE'
        return converter

    def _parse_convert_func(self, func_str: str,
                            build_chain: Optional[List[str]] = None) -> Callable:
        """
        解析转换函数字符串
        
        Args:
            func_str: 转换函数字符串，如 "int", "split_list(int)", "split_list2(split_list(int))"
        
        Returns:
            转换函数
        """
        from .types import TypeConverter
        
        if not func_str:
            # 默认返回字符串转换
            return TypeConverter.to_string
        
        func_str = func_str.strip()
        try:
            expression = parse_call(func_str)
        except ExpressionSyntaxError as exc:
            raise UndefinedTypeError(f"转换函数语法错误 '{func_str}': {exc}") from exc
        if expression.is_call:
            return self._create_convert_func(
                expression.name, list(expression.args), func_str, build_chain
            )
        return self._create_simple_convert_func(expression.name, build_chain)
    
    def _parse_args(self, args_str: str) -> List[Any]:
        """解析函数参数"""
        return [part for part in split_top_level(args_str, (',', ';')) if part]

    @staticmethod
    def _require_arg_count(func_name: str, args: List[str], minimum: int,
                           maximum: Optional[int] = None) -> None:
        maximum = minimum if maximum is None else maximum
        if not minimum <= len(args) <= maximum:
            expected = str(minimum) if minimum == maximum else f"{minimum}-{maximum}"
            raise UndefinedTypeError(
                f"转换函数 {func_name} 参数数量错误: 需要{expected}个，实际{len(args)}个"
            )
    
    def _create_convert_func(self, func_name: str, args: List[str],
                             full_func_str: str = '',
                             build_chain: Optional[List[str]] = None) -> Callable:
        """创建转换函数"""
        from .types import TypeConverter
        
        # 处理带方括号的自定义分隔符语法，如 split_list[_] 或 split_list[|]
        separator = None
        if '[' in func_name and ']' in func_name:
            base_name = func_name[:func_name.find('[')]
            sep_part = func_name[func_name.find('[')+1:func_name.find(']')]
            separator = TypeConverter.SEPARATOR_MAP.get(sep_part, sep_part)
            func_name = base_name
        
        if func_name == 'int':
            self._require_arg_count(func_name, args, 0)
            return TypeConverter.to_int
        elif func_name == 'float':
            self._require_arg_count(func_name, args, 0)
            return TypeConverter.to_float
        elif func_name == 'str' or func_name == 'string':
            self._require_arg_count(func_name, args, 0)
            return TypeConverter.to_string
        elif func_name == 'bool':
            self._require_arg_count(func_name, args, 0)
            return TypeConverter.to_bool
        elif func_name == 'bytes':
            self._require_arg_count(func_name, args, 0)
            return TypeConverter.to_bytes
        elif func_name == 'enum':
            if len(args) < 2:
                raise UndefinedTypeError("转换函数 enum 至少需要基础类型和一个允许值")
            base_type = args[0].strip().lower()
            if base_type not in ('int', 'string', 'str'):
                raise UndefinedTypeError("enum 仅支持 int 或 string 基础类型")
            base_converter = (
                TypeConverter.to_int if base_type == 'int' else TypeConverter.to_string
            )
            try:
                allowed_values = tuple(base_converter(unquote(item)) for item in args[1:])
            except Exception as exc:
                raise UndefinedTypeError(f"enum 允许值转换失败: {exc}") from exc

            def convert_enum(value):
                converted = base_converter(value)
                if converted not in allowed_values:
                    allowed = ", ".join(str(item) for item in allowed_values)
                    raise ValueError(f"值 '{value}' 不在枚举允许值中: {allowed}")
                return converted

            return convert_enum
        elif func_name == 'split_list':
            self._require_arg_count(func_name, args, 1)
            if args:
                inner_type = args[0]
                inner_func = self._parse_convert_func(inner_type, build_chain)
                sep = separator if separator else '#'
                return lambda v: TypeConverter.split_list(v, inner_func, sep)
            return TypeConverter.split_list
        elif func_name == 'split_list_ex':
            self._require_arg_count(func_name, args, 2)
            # 自定义分隔符版本
            if len(args) >= 2:
                sep = args[0]
                inner_type = args[1]
                return lambda v: TypeConverter.split_list(v, inner_type, sep)
            return TypeConverter.split_list
        elif func_name == 'split_list2':
            self._require_arg_count(func_name, args, 1)
            if args:
                inner_func = self._parse_convert_func(args[0], build_chain)
                return lambda v: TypeConverter.split_list2(v, inner_func)
            return TypeConverter.split_list2
        elif func_name == 'split_list3':
            self._require_arg_count(func_name, args, 1)
            if args:
                inner_func = self._parse_convert_func(args[0], build_chain)
                return lambda v: TypeConverter.split_list3(v, inner_func)
            return TypeConverter.split_list3
        elif func_name == 'split_dict':
            if not args:
                raise UndefinedTypeError("转换函数 split_dict 至少需要一个字段定义")
            fields = []
            for definition in args:
                if ':' in definition and ' ' not in definition:
                    field_name, field_expression = definition.split(':', 1)
                else:
                    try:
                        field_expression, field_name = definition.rsplit(None, 1)
                    except ValueError as exc:
                        raise UndefinedTypeError(
                            f"split_dict 字段定义错误: {definition}"
                        ) from exc
                field_name = field_name.strip()
                if not field_name or any(name == field_name for name, _ in fields):
                    raise UndefinedTypeError(f"split_dict 字段名无效或重复: {field_name}")
                fields.append((
                    field_name,
                    self._parse_convert_func(field_expression, build_chain),
                ))
            return lambda v: TypeConverter.split_dict_fields(v, fields)
        elif func_name == 'find_id':
            self._require_arg_count(func_name, args, 3)
            if len(args) >= 3:
                return self._create_reference_converter(*args[:3])
            return lambda v: v
        elif func_name == 'find':
            self._require_arg_count(func_name, args, 3)
            if len(args) >= 3:
                return self._create_reference_converter(*args[:3])
            return lambda v: v
        elif func_name == 'path':
            self._require_arg_count(func_name, args, 0, 2)
            if args:
                return lambda v: TypeConverter.path(v, *args)
            return TypeConverter.path
        elif func_name == 'text_key':
            self._require_arg_count(func_name, args, 0)
            return TypeConverter.to_text_key
        elif func_name == 'commonStringParamForSplit':
            self._require_arg_count(func_name, args, 0)
            return lambda v: TypeConverter.split_list(v, 'string', '#')
        else:
            raise UndefinedTypeError(f"未知的转换函数: {func_name}")

    def _create_reference_converter(self, table_name: str, display_label: str,
                                    id_field: str) -> Callable:
        """Preserve legacy scalar/dict shape while typing the referenced ID."""
        from .types import TypeConverter

        legacy_type = self.get_legacy_referenced_field_type(table_name, id_field)
        scalar_type = self.resolve_referenced_scalar_type(table_name, id_field)
        if legacy_type == 'int':
            return TypeConverter._convert_find_id_int
        if legacy_type in ('str', 'string'):
            return TypeConverter._convert_find_id_str
        if legacy_type == 'float':
            return TypeConverter._convert_find_id_float

        scalar_converter = self.scalar_converter(scalar_type)
        return lambda value: TypeConverter.find_id_typed(
            value, table_name, display_label, id_field, scalar_converter
        )

    def _create_simple_convert_func(self, type_name: str,
                                    build_chain: Optional[List[str]] = None) -> Callable:
        """创建简单类型转换函数"""
        from .types import TypeConverter
        
        type_map = {
            'int': TypeConverter.to_int,
            'float': TypeConverter.to_float,
            'str': TypeConverter.to_string,
            'string': TypeConverter.to_string,
            'bool': TypeConverter.to_bool,
            'bytes': TypeConverter.to_bytes,
            'text_key': TypeConverter.to_text_key,
            'commonStringParamForSplit': lambda value: TypeConverter.split_list(
                value, 'string', '#'
            ),
        }
        
        converter = type_map.get(type_name)
        if converter is None and type_name in self._types:
            return self._ensure_converter(type_name, build_chain or [])
        if converter is None:
            raise UndefinedTypeError(f"未知的转换函数: {type_name}")
        return converter
    
    def has_type(self, type_name: str) -> bool:
        """检查类型是否已定义"""
        return type_name in self._types
    
    def get_reference_info(self, type_name: str) -> Optional[Dict[str, str]]:
        """
        获取类型的引用信息（用于ID引用验证）
        
        Args:
            type_name: 类型名称
        
        Returns:
            {'table': 表名, 'field': 字段名} 或 None
        """
        try:
            base_type = parse_field_type(str(type_name)).base_type
        except ExpressionSyntaxError:
            return None
        return self._reference_info_from_type(base_type, set())

    def _reference_info_from_type(
        self, type_name: str, visited: Set[str]
    ) -> Optional[Dict[str, str]]:
        """Follow aliases and nested converters to the first find/find_id."""
        if type_name in visited or type_name not in self._types:
            return None
        definition = self._types[type_name].get('convert_func_str', '')
        if not definition:
            return None
        return self._reference_info_from_conversion(
            definition, visited | {type_name}
        )

    def _reference_info_from_conversion(
        self, conversion: str, visited: Set[str]
    ) -> Optional[Dict[str, str]]:
        try:
            expression = parse_call(conversion)
        except ExpressionSyntaxError:
            return None
        if not expression.is_call:
            return self._reference_info_from_type(expression.name, visited)

        name = expression.name.split('[', 1)[0].strip().lower()
        args = list(expression.args)
        if name in ('find', 'find_id'):
            if len(args) < 3:
                return None
            return {'table': args[0], 'field': args[2]}

        nested_expressions: List[str] = []
        if name in ('split_list', 'split_list2', 'split_list3') and args:
            nested_expressions.append(args[0])
        elif name == 'split_list_ex' and len(args) >= 2:
            nested_expressions.append(args[1])
        elif name == 'split_dict':
            for definition in args:
                if ':' in definition and ' ' not in definition:
                    _, nested = definition.split(':', 1)
                else:
                    try:
                        nested, _ = definition.rsplit(None, 1)
                    except ValueError:
                        continue
                nested_expressions.append(nested)

        for nested in nested_expressions:
            reference = self._reference_info_from_conversion(nested, visited)
            if reference:
                return reference
        return None
    
    def get_type(self, type_name: str) -> Dict[str, Any]:
        """
        获取类型定义
        
        Args:
            type_name: 类型名称
        
        Returns:
            类型定义字典
        
        Raises:
            UndefinedTypeError: 类型未定义
        """
        if type_name not in self._types:
            available = list(self._types.keys())
            raise UndefinedTypeError(
                f"类型 '{type_name}' 未定义。\n"
                f"请在TypeDefinition.xlsx中添加此类型定义。\n"
                f"可用类型: {', '.join(available)}"
            )
        
        return self._types[type_name]
    
    def get_all_types(self) -> List[str]:
        """获取所有类型名称"""
        return list(self._types.keys())

    def get_referenced_type(self, type_name: str) -> Optional[str]:
        """Return the canonical scalar for a direct find/find_id definition."""
        definition = self._types.get(type_name)
        if not definition:
            return None
        try:
            expression = parse_call(definition.get('convert_func_str', ''))
        except ExpressionSyntaxError:
            return None
        if not expression.is_call or expression.name.lower() not in ('find', 'find_id'):
            return None
        if len(expression.args) < 3:
            return None
        return self.resolve_referenced_scalar_type(
            expression.args[0], expression.args[2]
        )

    def find_reference_workbook(self, table_name: str) -> Optional[str]:
        """Resolve a reference workbook exactly, then by deterministic prefix."""
        exact = os.path.join(self.table_dir, f"{table_name}.xlsx")
        if os.path.isfile(exact):
            return exact
        lowered = str(table_name).casefold()
        for filename in sorted(os.listdir(self.table_dir), key=str.casefold):
            if filename.casefold().startswith(lowered) and filename.casefold().endswith('.xlsx'):
                candidate = os.path.join(self.table_dir, filename)
                if os.path.isfile(candidate):
                    return candidate
        return None

    def _referenced_field_candidates(
        self, table_name: str, field_name: str
    ) -> Tuple[Optional[str], List[Tuple[str, str]]]:
        """Return every matching (sheet, raw type) in the target workbook."""
        cache_key = (
            str(table_name).strip().casefold(),
            str(field_name).strip().casefold(),
        )
        cached = self._reference_field_cache.get(cache_key)
        if cached is not None:
            workbook_path, candidates = cached
            return workbook_path, list(candidates)

        workbook_path = self.find_reference_workbook(table_name)
        if not workbook_path:
            self._reference_field_cache[cache_key] = (None, ())
            return None, []

        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        candidates: List[Tuple[str, str]] = []
        wanted = str(field_name).strip()
        try:
            for sheet_name in workbook.sheetnames:
                if sheet_name.upper() in ('CODE', 'PROTO'):
                    continue
                worksheet = workbook[sheet_name]
                rows = list(worksheet.iter_rows(
                    min_row=1, max_row=2, values_only=True
                ))
                if len(rows) < 2:
                    continue
                names, types = rows
                for index, value in enumerate(names):
                    if str(value or '').strip() != wanted:
                        continue
                    raw_type = types[index] if index < len(types) else None
                    if raw_type is not None and str(raw_type).strip():
                        candidates.append((sheet_name, str(raw_type).strip()))
        finally:
            workbook.close()
        self._reference_field_cache[cache_key] = (
            workbook_path, tuple(candidates)
        )
        return workbook_path, candidates

    def get_legacy_referenced_field_type(self, table_name: str,
                                         field_name: str) -> str:
        """Return the old first-match type used only to preserve output shape."""
        _, candidates = self._referenced_field_candidates(table_name, field_name)
        if not candidates:
            return 'str'
        try:
            return parse_field_type(candidates[0][1]).base_type
        except ExpressionSyntaxError:
            return 'str'

    @staticmethod
    def scalar_converter(scalar_type: str) -> Callable:
        """Return the identity conversion for one canonical scalar type."""
        from .types import TypeConverter

        converters = {
            'int': TypeConverter.to_int,
            'str': TypeConverter.to_string,
            'string': TypeConverter.to_string,
            'float': TypeConverter.to_float,
            'bool': TypeConverter.to_bool,
            'bytes': TypeConverter.to_bytes,
        }
        converter = converters.get(str(scalar_type).lower())
        if converter is None:
            raise UndefinedTypeError(
                f"find_id引用目标不是支持的标量类型: {scalar_type}"
            )
        return converter

    def resolve_scalar_type(self, type_name: str) -> str:
        """Resolve a TypeDefinition name to a canonical scalar type."""
        return self._resolve_type_scalar(type_name, (), ())

    def _resolve_type_scalar(
        self, type_name: str, type_stack: Tuple[str, ...],
        reference_stack: Tuple[Tuple[str, str], ...],
    ) -> str:
        try:
            base_type = parse_field_type(str(type_name)).base_type
        except ExpressionSyntaxError as exc:
            raise UndefinedTypeError(f"字段类型表达式无效 '{type_name}': {exc}") from exc

        normalized = {
            'int': 'int', 'str': 'str', 'string': 'str',
            'float': 'float', 'bool': 'bool', 'bytes': 'bytes',
        }.get(base_type.lower())
        if normalized:
            return normalized
        if base_type in type_stack:
            start = type_stack.index(base_type)
            cycle = type_stack[start:] + (base_type,)
            raise UndefinedTypeError("类型别名存在循环: " + " -> ".join(cycle))

        definition = self._types.get(base_type)
        if not definition:
            raise UndefinedTypeError(f"find_id引用目标使用了未定义类型: {base_type}")
        return self._resolve_conversion_scalar(
            definition.get('convert_func_str', ''),
            type_stack + (base_type,), reference_stack,
        )

    def _resolve_conversion_scalar(
        self, expression_text: str, type_stack: Tuple[str, ...],
        reference_stack: Tuple[Tuple[str, str], ...],
    ) -> str:
        try:
            expression = parse_call(expression_text)
        except ExpressionSyntaxError as exc:
            raise UndefinedTypeError(
                f"转换函数语法错误 '{expression_text}': {exc}"
            ) from exc

        if not expression.is_call:
            return self._resolve_type_scalar(
                expression.name, type_stack, reference_stack
            )

        name = expression.name.split('[', 1)[0].lower()
        args = list(expression.args)
        if name in ('int', 'str', 'string', 'float', 'bool', 'bytes'):
            self._require_arg_count(name, args, 0)
            return 'str' if name in ('str', 'string') else name
        if name == 'enum':
            if not args:
                raise UndefinedTypeError("enum缺少基础标量类型")
            return self._resolve_type_scalar(args[0], type_stack, reference_stack)
        if name == 'path':
            self._require_arg_count(name, args, 0, 2)
            return 'str'
        if name in ('find', 'find_id'):
            self._require_arg_count(name, args, 3)
            return self._resolve_referenced_scalar_type(
                args[0], args[2], type_stack, reference_stack
            )
        if name in ('split_list', 'split_list2', 'split_list3', 'split_list_ex',
                    'split_dict', 'dict', 'award', 'commonstringparamforsplit'):
            raise UndefinedTypeError(
                f"find_id引用目标必须是标量，不能使用 {expression.name}"
            )
        if name == 'text_key':
            raise UndefinedTypeError(
                "find_id引用目标不能使用text_key：它没有唯一的静态标量类型"
            )
        raise UndefinedTypeError(f"find_id引用目标使用了未知转换函数: {expression.name}")

    def resolve_referenced_scalar_type(self, table_name: str,
                                       field_name: str) -> str:
        """Resolve a find/find_id target field to its final scalar type."""
        return self._resolve_referenced_scalar_type(
            table_name, field_name, (), ()
        )

    def _resolve_referenced_scalar_type(
        self, table_name: str, field_name: str,
        type_stack: Tuple[str, ...],
        reference_stack: Tuple[Tuple[str, str], ...],
    ) -> str:
        workbook_path, candidates = self._referenced_field_candidates(
            table_name, field_name
        )
        if not workbook_path or not candidates:
            # ReferenceValidator owns the actionable missing-table/field error.
            return 'str'

        key = (
            os.path.normcase(os.path.abspath(workbook_path)),
            str(field_name).casefold(),
        )
        if key in reference_stack:
            raise _ReferenceTypeCycle(
                f"find_id引用形成循环: {table_name}.{field_name}"
            )
        next_references = reference_stack + (key,)

        resolved: Set[str] = set()
        cycles: List[str] = []
        errors: List[str] = []
        reference_candidates: List[Tuple[str, str]] = []
        for sheet_name, raw_type in candidates:
            if self.get_reference_info(raw_type):
                reference_candidates.append((sheet_name, raw_type))
                continue
            try:
                resolved.add(self._resolve_type_scalar(
                    raw_type, type_stack, next_references
                ))
            except _ReferenceTypeCycle as exc:
                cycles.append(f"{sheet_name}: {exc}")
            except UndefinedTypeError as exc:
                errors.append(f"{sheet_name}: {exc}")

        target = f"{os.path.basename(workbook_path)}.{field_name}"
        if len(resolved) > 1:
            raise UndefinedTypeError(
                f"引用目标 {target} 在多个工作表中的类型冲突: "
                + ", ".join(sorted(resolved))
            )
        if resolved:
            return next(iter(resolved))
        if errors:
            raise UndefinedTypeError(
                f"无法确定引用目标 {target} 的标量类型: " + "; ".join(errors)
            )
        if reference_candidates:
            # Preserve the wire/conversion scalar through a reference-only
            # chain. ReferenceValidator still rejects the target because these
            # columns never contribute legal IDs.
            fallback_resolved: Set[str] = set()
            fallback_cycles: List[str] = []
            fallback_errors: List[str] = []
            for sheet_name, raw_type in reference_candidates:
                try:
                    fallback_resolved.add(self._resolve_type_scalar(
                        raw_type, type_stack, next_references
                    ))
                except _ReferenceTypeCycle as exc:
                    fallback_cycles.append(f"{sheet_name}: {exc}")
                except UndefinedTypeError as exc:
                    fallback_errors.append(f"{sheet_name}: {exc}")
            if fallback_errors:
                raise UndefinedTypeError(
                    f"无法确定引用目标 {target} 的标量类型: "
                    + "; ".join(fallback_errors)
                )
            if len(fallback_resolved) > 1:
                raise UndefinedTypeError(
                    f"引用目标 {target} 在多个工作表中的类型冲突: "
                    + ", ".join(sorted(fallback_resolved))
                )
            if fallback_resolved:
                return next(iter(fallback_resolved))
            if fallback_cycles and len(fallback_cycles) == len(reference_candidates):
                # Keep the historic schema fallback for a pure reference cycle;
                # validation will still report the missing independent column.
                return 'int'
        if cycles and len(cycles) == len(candidates):
            return 'int'
        raise UndefinedTypeError(
            f"无法确定引用目标 {target} 的标量类型"
        )

    def get_referenced_field_type(self, table_name: str, field_name: str) -> str:
        """Backward-compatible public alias for final scalar inference."""
        return self.resolve_referenced_scalar_type(table_name, field_name)
