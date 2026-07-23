# -*- coding: utf-8 -*-
"""Protobuf schema parsing, rendering, and compatibility validation."""

from __future__ import annotations

import json
import os
import re
from collections import OrderedDict
from dataclasses import dataclass, field
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import load_workbook


MANIFEST_VERSION = 1
MANIFEST_BEGIN = "// excel2json-schema-manifest-begin"
MANIFEST_END = "// excel2json-schema-manifest-end"

PROTO_SCALAR_TYPES = {
    "double",
    "float",
    "int32",
    "int64",
    "uint32",
    "uint64",
    "sint32",
    "sint64",
    "fixed32",
    "fixed64",
    "sfixed32",
    "sfixed64",
    "bool",
    "string",
    "bytes",
}

PROTO_IDENTIFIER_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")
PROTO_PACKAGE_RE = re.compile(
    r"^[A-Za-z_][A-Za-z0-9_]*(?:\.[A-Za-z_][A-Za-z0-9_]*)*$"
)
PROTO_FIELD_NUMBER_MAX = 536_870_911
PROTO_RESERVED_NUMBER_START = 19_000
PROTO_RESERVED_NUMBER_END = 19_999
PROTO_KEYWORDS = {
    "syntax",
    "import",
    "weak",
    "public",
    "package",
    "option",
    "optional",
    "required",
    "repeated",
    "oneof",
    "map",
    "reserved",
    "message",
    "enum",
    "service",
    "rpc",
    "returns",
    "stream",
    "extend",
    "extensions",
    "to",
    "max",
    "group",
    "true",
    "false",
}


class ProtoSchemaError(Exception):
    """Raised when a PROTO worksheet or schema transition is invalid."""


@dataclass(frozen=True)
class ProtoField:
    kind: str
    sheet: str
    message: str
    name: str
    number: int
    type_name: str
    rule: str
    oneof: str
    source: str
    description: str = ""

    def manifest_dict(self) -> Dict[str, Any]:
        return {
            "name": self.name,
            "number": self.number,
            "type": self.type_name,
            "rule": self.rule,
            "oneof": self.oneof,
            "source": self.source,
        }


@dataclass
class ProtoMessage:
    name: str
    fields: List[ProtoField] = field(default_factory=list)
    reserved_numbers: set[int] = field(default_factory=set)
    reserved_names: set[str] = field(default_factory=set)

    def manifest_dict(self) -> Dict[str, Any]:
        return {
            "name": self.name,
            "fields": [
                item.manifest_dict()
                for item in sorted(self.fields, key=lambda value: value.number)
            ],
            "reserved_numbers": sorted(self.reserved_numbers),
            "reserved_names": sorted(self.reserved_names),
        }


@dataclass
class ProtoSchema:
    package: str
    csharp_namespace: str
    sheet: str
    root_message: str
    root_field: ProtoField
    messages: "OrderedDict[str, ProtoMessage]"

    @property
    def row_message(self) -> ProtoMessage:
        return self.messages[self.root_field.type_name]

    def manifest(self) -> Dict[str, Any]:
        return {
            "format_version": MANIFEST_VERSION,
            "package": self.package,
            "csharp_namespace": self.csharp_namespace,
            "sheet": self.sheet,
            "root_message": self.root_message,
            "messages": [
                self.messages[name].manifest_dict()
                for name in sorted(self.messages)
            ],
        }

    def validate_excel_fields(self, worksheet: Any) -> None:
        """Require every exported Excel column to be mapped by the row message."""
        expected = {
            name
            for name, info in worksheet.field_info.items()
            if str(info.platform or "cs").strip().lower() != "x"
        }

        mapped = set()
        for proto_field in self.row_message.fields:
            source = proto_field.source
            if not source.startswith("$"):
                mapped.add(source.split(".", 1)[0])

        missing = sorted(expected - mapped)
        unknown = sorted(mapped - expected)
        errors = []
        if missing:
            errors.append(f"未映射的Excel字段: {', '.join(missing)}")
        if unknown:
            errors.append(f"PROTO source引用了不存在的Excel字段: {', '.join(unknown)}")
        if errors:
            raise ProtoSchemaError(f"PROTO[{self.sheet}] " + "; ".join(errors))

    def optional_empty_sources(self) -> set[str]:
        """Return direct Excel fields whose empty cells mean proto absence."""
        return {
            item.source
            for item in self.row_message.fields
            if item.rule == "optional"
            and not item.source.startswith("$")
            and "." not in item.source
        }

    def render_proto(self) -> str:
        manifest_json = json.dumps(
            self.manifest(), ensure_ascii=False, indent=2, sort_keys=True
        )
        lines = [MANIFEST_BEGIN]
        lines.extend(f"// {line}" for line in manifest_json.splitlines())
        lines.extend(
            [
                MANIFEST_END,
                "",
                'syntax = "proto3";',
                "",
                f"package {self.package};",
                f'option csharp_namespace = "{_escape_proto_string(self.csharp_namespace)}";',
                "",
            ]
        )

        for message_name in self._ordered_message_names():
            message = self.messages[message_name]
            lines.extend(_render_message(message))
            lines.append("")

        return "\n".join(lines).rstrip() + "\n"

    def _ordered_message_names(self) -> List[str]:
        names = [self.root_message]
        names.extend(sorted(name for name in self.messages if name != self.root_message))
        return names


class ProtoSchemaParser:
    """Parse the PROTO worksheet for one exported data sheet."""

    REQUIRED_COLUMNS = (
        "kind",
        "sheet",
        "message",
        "field",
        "number",
        "type",
        "rule",
        "oneof",
        "source",
        "description",
    )

    @classmethod
    def parse_workbook(
        cls, file_path: str, sheet_name: str,
        previous_manifest: Optional[Dict[str, Any]] = None,
        message_base: Optional[str] = None,
    ) -> ProtoSchema:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            proto_sheet_name = next(
                (name for name in workbook.sheetnames if name.upper() == "PROTO"),
                None,
            )
            if not proto_sheet_name:
                # Convention mode is also exposed through this public parser API.
                from .type_registry import TypeRegistry
                from .core import WorkSheet, FieldInfo
                if sheet_name not in workbook.sheetnames:
                    raise ProtoSchemaError(f"工作表不存在: {sheet_name}")
                data_ws = workbook[sheet_name]
                rows = list(data_ws.iter_rows(min_row=1, max_row=4, values_only=True))
                if len(rows) < 4:
                    raise ProtoSchemaError(f"工作表 {sheet_name} 表头不完整")
                inferred = WorkSheet(sheet_name)
                field_info = OrderedDict()
                names = []
                for index, raw_name in enumerate(rows[0]):
                    name = str(raw_name or "").strip()
                    names.append(name)
                    if name:
                        field_info[name] = FieldInfo(
                            name, str(rows[1][index] or "str"),
                            str(rows[2][index] or "cs"), str(rows[3][index] or "")
                        )
                inferred.set_field_info(field_info, {n: i for i, n in enumerate(names) if n})
                try:
                    registry = TypeRegistry(os.path.dirname(file_path))
                except Exception as exc:
                    raise ProtoSchemaError(
                        f"{file_path} 缺少PROTO工作表，且无法自动推导 {sheet_name}.pb: {exc}"
                    ) from exc
                return cls.auto_from_worksheet(
                    inferred, registry, previous_manifest,
                    message_base=message_base,
                )

            worksheet = workbook[proto_sheet_name]
            if _cell_text(worksheet.cell(1, 1)).lower() != "package":
                raise ProtoSchemaError("PROTO!A1必须为 package")
            if _cell_text(worksheet.cell(2, 1)).lower() != "csharp_namespace":
                raise ProtoSchemaError("PROTO!A2必须为 csharp_namespace")
            package = _cell_text(worksheet.cell(1, 2)) or "config"
            csharp_namespace = _cell_text(worksheet.cell(2, 2)) or "Game.Config"
            cls._validate_package(package, csharp_namespace)

            proto_max_row = worksheet.max_row or 0
            row4 = tuple(next(worksheet.iter_rows(min_row=4, max_row=4), ())) \
                if proto_max_row >= 4 else ()
            headers = {
                _cell_text(cell).lower(): index
                for index, cell in enumerate(row4)
                if _cell_text(cell)
            }
            missing_headers = [
                column for column in cls.REQUIRED_COLUMNS if column not in headers
            ]
            if missing_headers:
                # A metadata-only PROTO sheet intentionally omits row 4 and
                # delegates field/message generation to convention mode.
                if not headers or not any(_cell_text(cell) for cell in row4):
                    from .type_registry import TypeRegistry
                    inferred = cls._infer_data_worksheet(workbook, sheet_name)
                    try:
                        registry = TypeRegistry(os.path.dirname(file_path))
                    except Exception as exc:
                        raise ProtoSchemaError(
                            f"PROTO只有命名空间但无法自动推导 {sheet_name}: {exc}"
                        ) from exc
                    return cls.auto_from_worksheet(
                        inferred, registry, previous_manifest,
                        package=package, csharp_namespace=csharp_namespace,
                        message_base=message_base,
                    )
                raise ProtoSchemaError(
                    "PROTO第4行缺少列: " + ", ".join(missing_headers)
                )

            raw_rows = []
            for excel_row, row in enumerate(
                worksheet.iter_rows(min_row=5, values_only=True), start=5
            ):
                values = {
                    column: _value_text(row[index]) if index < len(row) else ""
                    for column, index in headers.items()
                }
                if not any(values.values()):
                    continue
                if values["sheet"] != sheet_name:
                    continue
                values["excel_row"] = excel_row
                raw_rows.append(values)

            if not raw_rows:
                from .type_registry import TypeRegistry
                inferred = cls._infer_data_worksheet(workbook, sheet_name)
                try:
                    registry = TypeRegistry(os.path.dirname(file_path))
                except Exception as exc:
                    raise ProtoSchemaError(
                        f"PROTO只有命名空间但无法自动推导 {sheet_name}: {exc}"
                    ) from exc
                return cls.auto_from_worksheet(
                    inferred, registry, previous_manifest,
                    package=package, csharp_namespace=csharp_namespace,
                    message_base=message_base,
                )
            return cls._build_schema(
                package, csharp_namespace, sheet_name, raw_rows
            )
        except ProtoSchemaError as exc:
            prefix = f"[{os.path.basename(file_path)}:{sheet_name}]"
            if str(exc).startswith(prefix):
                raise
            raise ProtoSchemaError(f"{prefix} {exc}") from exc
        finally:
            workbook.close()

    @staticmethod
    def _infer_data_worksheet(workbook: Any, sheet_name: str) -> Any:
        from .core import WorkSheet, FieldInfo
        if sheet_name not in workbook.sheetnames:
            raise ProtoSchemaError(f"工作表不存在: {sheet_name}")
        data_ws = workbook[sheet_name]
        rows = list(data_ws.iter_rows(min_row=1, max_row=4, values_only=True))
        if len(rows) < 4:
            raise ProtoSchemaError(f"工作表 {sheet_name} 表头不完整")
        inferred = WorkSheet(sheet_name)
        field_info = OrderedDict()
        names = []
        for index, raw_name in enumerate(rows[0]):
            name = str(raw_name or "").strip()
            names.append(name)
            if name:
                field_info[name] = FieldInfo(
                    name, str(rows[1][index] or "str"),
                    str(rows[2][index] or "cs"), str(rows[3][index] or "")
                )
        inferred.set_field_info(field_info, {n: i for i, n in enumerate(names) if n})
        return inferred

    @classmethod
    def auto_from_worksheet(
        cls,
        worksheet: Any,
        type_registry: Any,
        previous_manifest: Optional[Dict[str, Any]] = None,
        package: Optional[str] = None,
        csharp_namespace: Optional[str] = None,
        message_base: Optional[str] = None,
    ) -> ProtoSchema:
        """Build a convention-based schema for ordinary Excel fields.

        The generated shape is
        ``OutputNameTable { repeated OutputNameRow rows = 1; }``.  The
        output name is the configured CODE file name without its extension;
        the worksheet name remains the data source name.
        TypeDefinition conversion expressions determine scalar/list dimensions;
        explicit PROTO remains available for structures that cannot be inferred.
        """
        sheet = str(worksheet.name)
        message_prefix = _auto_identifier(message_base or sheet)
        table_name = f"{message_prefix}Table"
        row_name = f"{message_prefix}Row"
        package = package or (previous_manifest or {}).get("package", "config")
        namespace = csharp_namespace or (previous_manifest or {}).get(
            "csharp_namespace", "Game.Config"
        )
        cls._validate_package(package, namespace)

        messages: "OrderedDict[str, ProtoMessage]" = OrderedDict()
        messages[table_name] = ProtoMessage(table_name)
        messages[row_name] = ProtoMessage(row_name)
        root = ProtoField("root", sheet, table_name, "rows", 1, row_name, "repeated", "", "$rows", "")
        messages[table_name].fields.append(root)

        old_row = {}
        old_reserved_numbers: set[int] = set()
        old_reserved_names: set[str] = set()
        if previous_manifest:
            for msg in previous_manifest.get("messages", []):
                if msg.get("name") == row_name:
                    old_row = {f.get("name"): f for f in msg.get("fields", [])}
                    old_reserved_numbers.update(msg.get("reserved_numbers", []))
                    old_reserved_names.update(msg.get("reserved_names", []))

        used_names: set[str] = set()
        used_numbers: set[int] = set()
        used_numbers.update(int(v) for v in old_reserved_numbers if str(v).isdigit())
        old_active_numbers = {
            int(v.get("number")) for v in old_row.values()
            if str(v.get("number", "")).isdigit()
        }
        fields_by_source: Dict[str, ProtoField] = {}
        for index, (source, info) in enumerate(worksheet.field_info.items(), start=1):
            if str(getattr(info, "platform", "cs") or "cs").strip().lower() == "x":
                continue
            field_name = _auto_field_name(source, index, used_names)
            configured_type = str(getattr(info, "field_type", "str") or "str")
            # Constraints belong to the controlled TypeDefinition syntax and
            # do not change the underlying protobuf shape.
            configured_type = configured_type.split("+", 1)[0].strip()
            type_name, rule, extra_messages = cls._auto_type(
                sheet, field_name, configured_type, type_registry,
                message_prefix=message_prefix,
            )
            for message in extra_messages:
                messages.setdefault(message.name, message)
            old = old_row.get(field_name) or old_row.get(source)
            number = int(old["number"]) if old and str(old.get("number", "")).isdigit() else index
            if old is None:
                number = max(index, 2)
            while number in used_numbers or (old is None and number in old_active_numbers) or 19000 <= number <= 19999:
                number += 1
            used_numbers.add(number)
            field = ProtoField("field", sheet, row_name, field_name, number, type_name, rule, "", source, getattr(info, "desc", "") or "")
            messages[row_name].fields.append(field)
            fields_by_source[source] = field

        # Removed auto fields are reserved automatically to preserve wire compatibility.
        current_names = {f.name for f in messages[row_name].fields}
        for name, old in old_row.items():
            if name not in current_names:
                old_reserved_names.add(name)
                try:
                    old_reserved_numbers.add(int(old.get("number")))
                except (TypeError, ValueError):
                    pass
        messages[row_name].reserved_names.update(old_reserved_names)
        messages[row_name].reserved_numbers.update(old_reserved_numbers)
        schema = ProtoSchema(package, namespace, sheet, table_name, root, messages)
        schema.validate_excel_fields(worksheet)
        cls._validate_messages(messages)
        return schema

    @classmethod
    def _auto_type(
        cls,
        sheet: str,
        field_name: str,
        type_name: str,
        registry: Any,
        message_prefix: Optional[str] = None,
    ):
        try:
            definition = registry.get_type(type_name)
        except Exception as exc:
            raise ProtoSchemaError(
                f"字段 {sheet}.{field_name} 类型 {type_name} 无法推导Protobuf类型: {exc}"
            ) from exc
        func = definition.get("convert_func_str", "") or type_name
        try:
            base, dims = _auto_parse_conversion(func, registry)
        except ProtoSchemaError:
            raise
        except Exception as exc:
            raise ProtoSchemaError(
                f"字段 {sheet}.{field_name} 类型 {type_name} 无法推导Protobuf类型: {exc}"
            ) from exc
        if base == "dict":
            raise ProtoSchemaError(
                f"字段 {sheet}.{field_name} 类型 {type_name} 无法自动推导Protobuf结构，请增加PROTO工作表"
            )
        scalar = {"int": "int32", "float": "double", "str": "string", "string": "string", "bool": "bool", "bytes": "bytes", "path": "string", "text_key": "text_key"}.get(base)
        if not scalar:
            raise ProtoSchemaError(
                f"字段 {sheet}.{field_name} 类型 {type_name} 无法自动推导Protobuf类型，请增加PROTO工作表"
            )
        messages: List[ProtoMessage] = []
        prefix = _auto_identifier(message_prefix or sheet)
        if scalar == "text_key":
            msg_name = f"{prefix}{_auto_identifier(field_name)}TextKey"
            messages.append(ProtoMessage(msg_name, [
                ProtoField("field", sheet, msg_name, "int_value", 1, "int64", "singular", "value", "$self", ""),
                ProtoField("field", sheet, msg_name, "string_value", 2, "string", "singular", "value", "$self", ""),
            ]))
            current = msg_name
        else:
            current = scalar
        if dims == 1 and scalar != "text_key":
            return current, "repeated", messages
        for depth in range(max(0, dims - 1)):
            msg_name = f"{prefix}{_auto_identifier(field_name)}Level{dims-depth-1}"
            messages.append(ProtoMessage(msg_name, [ProtoField("field", sheet, msg_name, "values", 1, current, "repeated", "", "$self", "")]))
            current = msg_name
        return current, ("repeated" if dims == 1 else "repeated" if dims > 1 else "singular"), messages

    @classmethod
    def _build_schema(
        cls,
        package: str,
        csharp_namespace: str,
        sheet_name: str,
        rows: Iterable[Dict[str, Any]],
    ) -> ProtoSchema:
        messages: "OrderedDict[str, ProtoMessage]" = OrderedDict()
        root_fields: List[ProtoField] = []

        for row in rows:
            kind = row["kind"].lower()
            excel_row = row["excel_row"]
            if kind not in ("root", "field", "reserved"):
                raise ProtoSchemaError(
                    f"PROTO第{excel_row}行 kind无效: {row['kind']}"
                )

            message_name = row["message"]
            cls._validate_identifier(message_name, "message", excel_row)
            message = messages.setdefault(message_name, ProtoMessage(message_name))

            if kind == "reserved":
                cls._add_reserved(message, row)
                continue

            proto_field = cls._parse_field(kind, sheet_name, row)
            message.fields.append(proto_field)
            if kind == "root":
                root_fields.append(proto_field)

        if len(root_fields) != 1:
            raise ProtoSchemaError(
                f"PROTO[{sheet_name}] 必须且只能定义一个root行，当前为{len(root_fields)}个"
            )

        root_field = root_fields[0]
        if root_field.rule != "repeated" or root_field.source != "$rows":
            raise ProtoSchemaError(
                f"PROTO[{sheet_name}] root必须使用 rule=repeated 且 source=$rows"
            )
        if root_field.name != "rows" or root_field.number != 1:
            raise ProtoSchemaError(
                f"PROTO[{sheet_name}] root字段固定为 rows = 1"
            )

        cls._validate_messages(messages)
        if root_field.type_name in PROTO_SCALAR_TYPES:
            raise ProtoSchemaError("root字段type必须引用Row message，不能使用标量")
        if root_field.type_name not in messages:
            raise ProtoSchemaError(
                f"root引用了未定义的message: {root_field.type_name}"
            )
        if root_field.message == root_field.type_name:
            raise ProtoSchemaError("root message不能将自身作为Row message")
        if len(messages[root_field.message].fields) != 1:
            raise ProtoSchemaError(
                f"root message {root_field.message} 只能包含 repeated rows = 1"
            )
        cls._validate_optional_sources(messages[root_field.type_name])

        return ProtoSchema(
            package=package,
            csharp_namespace=csharp_namespace,
            sheet=sheet_name,
            root_message=root_field.message,
            root_field=root_field,
            messages=messages,
        )

    @classmethod
    def _parse_field(
        cls, kind: str, sheet_name: str, row: Dict[str, Any]
    ) -> ProtoField:
        excel_row = row["excel_row"]
        field_name = row["field"]
        type_name = row["type"]
        rule = (row["rule"] or "singular").lower()
        oneof = row["oneof"]
        source = row["source"] or field_name

        cls._validate_identifier(field_name, "field", excel_row)
        if type_name not in PROTO_SCALAR_TYPES:
            cls._validate_identifier(type_name, "type", excel_row)
        if rule not in ("singular", "optional", "repeated"):
            raise ProtoSchemaError(
                f"PROTO第{excel_row}行 rule无效: {row['rule']}"
            )
        if oneof:
            cls._validate_identifier(oneof, "oneof", excel_row)
            if rule != "singular":
                raise ProtoSchemaError(
                    f"PROTO第{excel_row}行 oneof字段只能使用singular"
                )
        if kind == "root" and oneof:
            raise ProtoSchemaError(f"PROTO第{excel_row}行 root不能属于oneof")
        if kind != "root" and source == "$rows":
            raise ProtoSchemaError(f"PROTO第{excel_row}行 只有root可以使用source=$rows")
        if source.startswith("$") and source not in ("$self", "$rows"):
            raise ProtoSchemaError(
                f"PROTO第{excel_row}行 source特殊值无效: {source}"
            )

        number = cls._parse_field_number(row["number"], excel_row)
        return ProtoField(
            kind=kind,
            sheet=sheet_name,
            message=row["message"],
            name=field_name,
            number=number,
            type_name=type_name,
            rule=rule,
            oneof=oneof,
            source=source,
            description=row["description"],
        )

    @classmethod
    def _add_reserved(cls, message: ProtoMessage, row: Dict[str, Any]) -> None:
        excel_row = row["excel_row"]
        if not row["field"]:
            raise ProtoSchemaError(
                f"PROTO第{excel_row}行 reserved必须填写原字段名"
            )
        cls._validate_identifier(row["field"], "reserved field", excel_row)
        number = cls._parse_field_number(row["number"], excel_row)
        message.reserved_numbers.add(number)
        message.reserved_names.add(row["field"])

    @classmethod
    def _validate_messages(
        cls, messages: "OrderedDict[str, ProtoMessage]"
    ) -> None:
        for message in messages.values():
            names: Dict[str, ProtoField] = {}
            numbers: Dict[int, ProtoField] = {}
            for proto_field in message.fields:
                if proto_field.name in names:
                    raise ProtoSchemaError(
                        f"message {message.name} 字段名重复: {proto_field.name}"
                    )
                if proto_field.number in numbers:
                    raise ProtoSchemaError(
                        f"message {message.name} 字段号重复: {proto_field.number}"
                    )
                if proto_field.name in message.reserved_names:
                    raise ProtoSchemaError(
                        f"message {message.name} 字段名已reserved: {proto_field.name}"
                    )
                if proto_field.number in message.reserved_numbers:
                    raise ProtoSchemaError(
                        f"message {message.name} 字段号已reserved: {proto_field.number}"
                    )
                names[proto_field.name] = proto_field
                numbers[proto_field.number] = proto_field

            oneof_names = {item.oneof for item in message.fields if item.oneof}
            oneof_conflicts = sorted(oneof_names.intersection(names))
            if oneof_conflicts:
                raise ProtoSchemaError(
                    f"message {message.name} oneof名称与字段名冲突: "
                    + ", ".join(oneof_conflicts)
                )

            for proto_field in message.fields:
                if (
                    proto_field.type_name not in PROTO_SCALAR_TYPES
                    and proto_field.type_name not in messages
                ):
                    raise ProtoSchemaError(
                        f"字段 {message.name}.{proto_field.name} 引用了未定义的message: "
                        f"{proto_field.type_name}"
                    )

    @staticmethod
    def _validate_optional_sources(row_message: ProtoMessage) -> None:
        direct_sources: Dict[str, List[ProtoField]] = {}
        for proto_field in row_message.fields:
            if proto_field.source.startswith("$") or "." in proto_field.source:
                continue
            direct_sources.setdefault(proto_field.source, []).append(proto_field)

        conflicts = []
        for source, fields in direct_sources.items():
            if len(fields) > 1 and any(item.rule == "optional" for item in fields):
                conflicts.append(
                    f"{source} -> {', '.join(item.name for item in fields)}"
                )
        if conflicts:
            raise ProtoSchemaError(
                "optional字段的direct source不能与其他Row字段复用: "
                + "; ".join(conflicts)
            )

    @staticmethod
    def _validate_identifier(value: str, label: str, excel_row: int) -> None:
        if (
            not value
            or not PROTO_IDENTIFIER_RE.fullmatch(value)
            or value.lower() in PROTO_KEYWORDS
        ):
            raise ProtoSchemaError(
                f"PROTO第{excel_row}行 {label}不是合法Proto标识符: {value or '(空)'}"
            )

    @staticmethod
    def _parse_field_number(value: str, excel_row: int) -> int:
        if not re.fullmatch(r"[0-9]+", str(value or "")):
            raise ProtoSchemaError(
                f"PROTO第{excel_row}行 field number无效: {value or '(空)'}"
            )
        number = int(value)
        if number < 1 or number > PROTO_FIELD_NUMBER_MAX:
            raise ProtoSchemaError(
                f"PROTO第{excel_row}行 field number超出范围: {number}"
            )
        if PROTO_RESERVED_NUMBER_START <= number <= PROTO_RESERVED_NUMBER_END:
            raise ProtoSchemaError(
                f"PROTO第{excel_row}行 field number位于Protobuf保留区间: {number}"
            )
        return number

    @staticmethod
    def _validate_package(package: str, csharp_namespace: str) -> None:
        if (
            not PROTO_PACKAGE_RE.fullmatch(package)
            or any(part.lower() in PROTO_KEYWORDS for part in package.split("."))
        ):
            raise ProtoSchemaError(f"PROTO package无效: {package}")
        if not PROTO_PACKAGE_RE.fullmatch(csharp_namespace):
            raise ProtoSchemaError(
                f"PROTO csharp_namespace无效: {csharp_namespace}"
            )


def extract_manifest(proto_text: str) -> Optional[Dict[str, Any]]:
    lines = proto_text.splitlines()
    try:
        start = lines.index(MANIFEST_BEGIN)
        end = lines.index(MANIFEST_END, start + 1)
    except ValueError:
        return None

    json_lines = []
    for line in lines[start + 1 : end]:
        if not line.startswith("//"):
            raise ProtoSchemaError("旧.proto中的兼容清单格式无效")
        content = line[2:]
        if content.startswith(" "):
            content = content[1:]
        json_lines.append(content)

    try:
        manifest = json.loads("\n".join(json_lines))
    except json.JSONDecodeError as exc:
        raise ProtoSchemaError(f"旧.proto中的兼容清单无法解析: {exc}")
    if manifest.get("format_version") != MANIFEST_VERSION:
        raise ProtoSchemaError(
            f"不支持的兼容清单版本: {manifest.get('format_version')}"
        )
    return manifest


def validate_compatible(
    old_manifest: Dict[str, Any], new_manifest: Dict[str, Any]
) -> None:
    errors = []
    for key in ("package", "csharp_namespace", "sheet", "root_message"):
        if old_manifest.get(key) != new_manifest.get(key):
            errors.append(
                f"{key} 从 {old_manifest.get(key)!r} 变更为 {new_manifest.get(key)!r}"
            )

    old_messages = {item["name"]: item for item in old_manifest.get("messages", [])}
    new_messages = {item["name"]: item for item in new_manifest.get("messages", [])}

    for message_name, old_message in old_messages.items():
        new_message = new_messages.get(message_name)
        if not new_message:
            errors.append(f"message被删除: {message_name}")
            continue

        new_fields_by_number = {
            item["number"]: item for item in new_message.get("fields", [])
        }
        new_fields_by_name = {
            item["name"]: item for item in new_message.get("fields", [])
        }
        reserved_numbers = set(new_message.get("reserved_numbers", []))
        reserved_names = set(new_message.get("reserved_names", []))

        for old_field in old_message.get("fields", []):
            new_field = new_fields_by_number.get(old_field["number"])
            if new_field:
                for key in ("name", "type", "rule", "oneof"):
                    if old_field.get(key, "") != new_field.get(key, ""):
                        errors.append(
                            f"{message_name}.{old_field['name']} 的{key}发生变更"
                        )
            else:
                same_name = new_fields_by_name.get(old_field["name"])
                if same_name:
                    errors.append(
                        f"{message_name}.{old_field['name']} 字段号从 "
                        f"{old_field['number']} 变更为 {same_name['number']}"
                    )
                elif (
                    old_field["number"] not in reserved_numbers
                    or old_field["name"] not in reserved_names
                ):
                    errors.append(
                        f"删除 {message_name}.{old_field['name']} 时必须同时reserved名称和字段号"
                    )

        old_reserved_numbers = set(old_message.get("reserved_numbers", []))
        old_reserved_names = set(old_message.get("reserved_names", []))
        if not old_reserved_numbers.issubset(reserved_numbers):
            errors.append(f"message {message_name} 不能移除已有reserved字段号")
        if not old_reserved_names.issubset(reserved_names):
            errors.append(f"message {message_name} 不能移除已有reserved字段名")

    if errors:
        raise ProtoSchemaError("Protobuf协议存在破坏性变更: " + "; ".join(errors))


def canonical_manifest(manifest: Dict[str, Any]) -> str:
    return json.dumps(manifest, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _render_message(message: ProtoMessage) -> List[str]:
    lines = [f"message {message.name} {{"]
    if message.reserved_numbers:
        values = ", ".join(str(value) for value in sorted(message.reserved_numbers))
        lines.append(f"  reserved {values};")
    if message.reserved_names:
        values = ", ".join(
            f'"{_escape_proto_string(value)}"'
            for value in sorted(message.reserved_names)
        )
        lines.append(f"  reserved {values};")

    fields = sorted(message.fields, key=lambda value: value.number)
    regular = [item for item in fields if not item.oneof]
    oneof_names = []
    for item in fields:
        if item.oneof and item.oneof not in oneof_names:
            oneof_names.append(item.oneof)

    if (message.reserved_numbers or message.reserved_names) and fields:
        lines.append("")
    for item in regular:
        lines.extend(_render_field(item, "  "))
    for oneof_name in oneof_names:
        if regular or lines[-1] != f"message {message.name} {{":
            lines.append("")
        lines.append(f"  oneof {oneof_name} {{")
        for item in fields:
            if item.oneof == oneof_name:
                lines.extend(_render_field(item, "    ", in_oneof=True))
        lines.append("  }")
    lines.append("}")
    return lines


def _render_field(
    proto_field: ProtoField, indent: str, in_oneof: bool = False
) -> List[str]:
    lines = []
    if proto_field.description:
        clean_description = proto_field.description.replace("\r", " ").replace("\n", " ")
        lines.append(f"{indent}// {clean_description}")
    prefix = ""
    if not in_oneof and proto_field.rule in ("optional", "repeated"):
        prefix = proto_field.rule + " "
    lines.append(
        f"{indent}{prefix}{proto_field.type_name} {proto_field.name} = {proto_field.number};"
    )
    return lines


def _cell_text(cell: Any) -> str:
    return _value_text(getattr(cell, "value", cell))


def _value_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _escape_proto_string(value: str) -> str:
    return value.replace("\\", "\\\\").replace('"', '\\"')


def _auto_identifier(value: str) -> str:
    text = re.sub(r"[^A-Za-z0-9_]", "_", str(value or ""))
    if not text:
        text = "Field"
    if text[0].isdigit():
        text = "Field_" + text
    if text.lower() in PROTO_KEYWORDS:
        text += "_field"
    return text


def _auto_field_name(source: str, index: int, used: set[str]) -> str:
    candidate = _auto_identifier(source)
    if candidate == "Field" and source:
        candidate = f"field_{index}"
    base = candidate
    suffix = 2
    while candidate in used:
        candidate = f"{base}_{suffix}"
        suffix += 1
    used.add(candidate)
    return candidate


def _auto_parse_conversion(func: str, registry: Any = None) -> tuple[str, int]:
    """Return (scalar base, list dimensions) from TypeDefinition syntax."""
    text = str(func or "").strip()
    if not text:
        return "str", 0
    match = re.match(r"^([A-Za-z_][A-Za-z0-9_]*(?:\[[^\]]*\])?)\((.*)\)$", text)
    if not match:
        lowered = text.lower()
        if lowered in ("int", "int32"):
            return "int", 0
        if lowered in ("float", "double"):
            return "float", 0
        if lowered in ("str", "string"):
            return "str", 0
        if lowered in ("bool", "bytes", "text_key"):
            return lowered, 0
        return lowered, 0
    name, args = match.groups()
    base_name = name.split("[", 1)[0].lower()
    if base_name in ("split_list", "split_list2", "split_list3"):
        inner_arg = _split_auto_args(args)[0] if _split_auto_args(args) else "str"
        inner, dims = _auto_parse_conversion(inner_arg, registry)
        extra = {"split_list": 1, "split_list2": 2, "split_list3": 3}[base_name]
        return inner, dims + extra
    if base_name in ("find_id", "find", "path"):
        if base_name in ("find_id", "find") and registry is not None:
            args = _split_auto_args(args)
            if len(args) >= 3:
                referenced = registry.resolve_referenced_scalar_type(
                    args[0].strip(), args[2].strip()
                )
                referenced = str(referenced or "str").lower()
                if referenced in {"int", "float", "bool", "bytes"}:
                    return referenced, 0
                if referenced in {"str", "string"}:
                    return "str", 0
                raise ProtoSchemaError(
                    f"find_id引用目标类型无法映射到Protobuf: {referenced}"
                )
        return ("int" if base_name in ("find_id", "find") else "path"), 0
    if base_name in ("split_dict", "dict", "award"):
        return "dict", 0
    inner_args = _split_auto_args(args)
    return _auto_parse_conversion(inner_args[0] if inner_args else "str")


def _split_auto_args(args: str) -> List[str]:
    """Split a TypeDefinition call without breaking nested find_id commas."""
    parts: List[str] = []
    depth = 0
    start = 0
    for index, char in enumerate(args):
        if char == "(":
            depth += 1
        elif char == ")":
            depth = max(0, depth - 1)
        elif char == "," and depth == 0:
            parts.append(args[start:index].strip())
            start = index + 1
    tail = args[start:].strip()
    if tail or not parts:
        parts.append(tail)
    return parts
